"""
ms_excel.py — Integrazione file Excel elaborati Fineco (Fase 2)
             ExcelIntegrator  — confronto e scrittura file elaborati
             leggi_excel_elaborato_cc / _cc_mensile
             MESI_IT_NOMI
Parte di: Moneyspire Reconciler
"""

import re
import json
import shutil
from pathlib import Path
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

from ms_constants import _to_date, _to_float, fmt_eur, DEFAULT_CONFIG, _amt_eq, _date_ok
from ms_parsers import (
    parse_fineco_conto_originale, parse_fineco_conto, parse_fineco_conto_usd,
    parse_fineco_cc,
)

# ─────────────────────────────────────────────────────────────────────────────
# FASE 2: INTEGRAZIONE FILE EXCEL ELABORATI
# ─────────────────────────────────────────────────────────────────────────────
#
# I file elaborati sono archivi storici tenuti aggiornati manualmente:
#   2026.xlsx      → fogli Movimenti (CC Fineco), Lombard, USD
#   2026 mc.xlsx   → fogli mensili (gennaio…dicembre) per MC Fineco
#   2026 visa.xlsx → fogli mensili per Visa (Ignazio + Silvia)
#
# Obiettivo Fase 2: dopo la riconciliazione, aggiornare i file elaborati
# con le transazioni del periodo analizzato che sono già presenti nel file
# originale Fineco ma potrebbero mancare nei file elaborati, oppure
# aggiornare la colonna Moneymap con le categorie riconciliate.
#
# Struttura dati:
#   File CC elaborato  → foglio con header: Data, Entrate, Uscite,
#                        Descrizione, Descrizione_Completa, Moneymap, Saldo
#   File MC elaborato  → foglio mensile: Data operazione, Data registrazione,
#                        Descrizione, Circuito, Importo
#   File Visa elaborato → stesso schema MC + colonna Intestatario carta
# ─────────────────────────────────────────────────────────────────────────────

MESI_IT_NOMI = ["gennaio","febbraio","marzo","aprile","maggio","giugno",
                "luglio","agosto","settembre","ottobre","novembre","dicembre"]


def leggi_excel_elaborato_cc(xlsx_path: str, sheet_name: str = "Movimenti",
                              mese: int | None = None,
                              anno: int | None = None) -> list[dict]:
    """
    Legge il file CC elaborato (2026.xlsx, foglio Movimenti/Lombard/USD).
    Ritorna lista di dict con chiavi: row_idx, date, deposit, withdrawal,
    amount, descrizione, descrizione_completa, moneymap, saldo.
    row_idx è l'indice 1-based nella riga del foglio (usato per aggiornamenti).
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    result = []
    header_row = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {str(h or "").strip(): i for i, h in enumerate(header_row) if h}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        dt = _to_date(row[col_idx.get("Data", 0)] if "Data" in col_idx else row[0])
        if dt is None:
            continue
        if mese and dt.month != mese:
            continue
        if anno and dt.year != anno:
            continue
        entr = abs(_to_float(row[col_idx.get("Entrate", 1)] if "Entrate" in col_idx else row[1]))
        usc  = abs(_to_float(row[col_idx.get("Uscite",  2)] if "Uscite"  in col_idx else row[2]))
        if entr == 0 and usc == 0:
            continue
        desc  = str(row[col_idx.get("Descrizione", 3)]          or "").strip()
        desc2 = str(row[col_idx.get("Descrizione_Completa", 4)] or "").strip()
        mmap  = str(row[col_idx.get("Moneymap", 5)]              or "").strip()
        saldo = _to_float(row[col_idx.get("Saldo", 6)]           if "Saldo" in col_idx else None)
        result.append({
            "row_idx":              row_num,
            "date":                 dt,
            "deposit":              entr,
            "withdrawal":           usc,
            "amount":               entr - usc,
            "descrizione":          desc,
            "descrizione_completa": desc2,
            "raw_text":             f"{desc} {desc2}".strip(),
            "moneymap":             mmap,
            "saldo":                saldo,
            "source_file":          xlsx_path,
            "source_sheet":         sheet_name,
        })
    wb.close()
    return result


def leggi_excel_elaborato_cc_mensile(xlsx_path: str, mese: int, anno: int,
                                      sheet_map: dict | None = None) -> list[dict]:
    """
    Legge il file MC o Visa elaborato per il mese richiesto.
    sheet_map: {numero_mese: nome_foglio} — default usa MESI_IT_NOMI.
    Ritorna lista di dict con row_idx, date, date_registrazione,
    amount, descrizione, intestatario.
    """
    from openpyxl import load_workbook
    if sheet_map is None:
        sheet_map = {i+1: MESI_IT_NOMI[i] for i in range(12)}
    nome_foglio = sheet_map.get(mese, MESI_IT_NOMI[mese-1])

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    # Ricerca foglio case-insensitive
    found = None
    for s in wb.sheetnames:
        if s.lower().strip() == nome_foglio.lower():
            found = s
            break
    if found is None:
        wb.close()
        return []
    ws = wb[found]
    header_row = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {str(h or "").strip(): i for i, h in enumerate(header_row) if h}
    result = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None for v in row):
            continue
        # Colonne flessibili: Visa ha Intestatario carta in più
        c_data_op = col_idx.get("Data operazione", col_idx.get("Data", 0))
        c_data_reg = col_idx.get("Data registrazione", None)
        c_desc = col_idx.get("Descrizione", None)
        c_imp  = col_idx.get("Importo", None)
        c_int  = col_idx.get("Intestatario carta", None)
        dt_op = _to_date(row[c_data_op] if c_data_op is not None else None)
        if dt_op is None:
            continue
        # Filtra per mese/anno della data operazione o registrazione
        dt_reg = _to_date(row[c_data_reg]) if c_data_reg is not None else None
        # Considera la riga se appartiene al mese estratto
        # (alcune righe di inizio mese hanno data op del mese precedente)
        dt_check = dt_reg or dt_op
        if anno and dt_check.year != anno:
            continue
        importo = _to_float(row[c_imp]) if c_imp is not None else 0.0
        desc = str(row[c_desc] or "").strip() if c_desc is not None else ""
        intestatario = str(row[c_int] or "").strip() if c_int is not None else ""
        result.append({
            "row_idx":            row_num,
            "date":               dt_op,
            "date_registrazione": dt_reg,
            "amount":             importo,
            "deposit":            max(importo, 0),
            "withdrawal":         abs(min(importo, 0)),
            "descrizione":        desc,
            "raw_text":           desc,
            "intestatario":       intestatario,
            "source_file":        xlsx_path,
            "source_sheet":       found,
        })
    wb.close()
    return result


class ExcelIntegrator:
    """
    Fase 2: confronto e allineamento tra file elaborati annuali e
    transazioni riconciliate con Moneyspire.

    Funzionalità:
    - confronta_cc():   identifica transazioni presenti in Fineco originale
                        ma assenti nel file elaborato (gap da colmare)
    - aggiorna_moneymap(): aggiorna colonna Moneymap nel file elaborato
                           con le categorie Money riconciliate
    - riepilogo_mensile(): totali entrate/uscite/saldo per mese
    """

    def __init__(self, xlsx_elaborato: str, tipo: str = "cc"):
        """
        xlsx_elaborato: path al file elaborato (2026.xlsx, 2026 mc.xlsx, 2026 visa.xlsx)
        tipo: "cc" | "mc" | "visa"
        """
        self.xlsx_path = xlsx_elaborato
        self.tipo = tipo

    def confronta_cc(self, bank_txns_originale: list[dict],
                     sheet_name: str = "Movimenti",
                     mese: int | None = None,
                     anno: int | None = None) -> dict:
        """
        Confronta le transazioni del file originale Fineco con quelle
        del file elaborato. Identifica:
        - presenti_solo_originale: transazioni mancanti nel file elaborato
        - presenti_solo_elaborato: righe nel file elaborato senza corrispondenza
        - moneymap_aggiornabile: righe elaborato con Moneymap vuota/diversa

        NOTA: il filtro mese/anno viene applicato SOLO al file elaborato per
        limitare le righe da cercare, ma se il mese non esiste ancora nel file
        elaborato (es. aprile non ancora inserito) legge l'intero file per
        trovare comunque eventuali abbinamenti parziali già presenti.
        Il matching per data ±1gg impedisce falsi abbinamenti tra mesi diversi.
        """
        # Leggi elaborato: prima prova con filtro mese (più veloce),
        # se vuoto rilegge senza filtro (mese ancora assente nel file elaborato)
        if mese is not None or anno is not None:
            elab = leggi_excel_elaborato_cc(self.xlsx_path, sheet_name, mese, anno)
        else:
            elab = leggi_excel_elaborato_cc(self.xlsx_path, sheet_name)

        # Se non ci sono righe nel mese richiesto, è probabile che il mese
        # non sia ancora nel file elaborato → confronto su tutto il file
        # (il matching data ±1gg evita comunque falsi abbinamenti)
        if not elab and bank_txns_originale:
            elab = leggi_excel_elaborato_cc(self.xlsx_path, sheet_name)
            # Filtra solo le righe dello stesso mese per "solo elaborato"
            # (non vogliamo mostrare righe di altri mesi come "extra")
            elab_per_solo = []
        else:
            elab_per_solo = elab  # usato per calcolare "presenti_solo_elaborato"

        usati_elab = set()
        solo_orig  = []
        aggiornabili = []

        for bt in bank_txns_originale:
            trovato = False
            for i, et in enumerate(elab):
                if i in usati_elab:
                    continue
                if (_amt_eq(bt["amount"], et["amount"], 0.02) and
                        _date_ok(bt["date"], et["date"], 1)):
                    usati_elab.add(i)
                    trovato = True
                    # Controlla se Moneymap è aggiornabile
                    cat_money = bt.get("_categoria_money", "")
                    if cat_money and cat_money != et["moneymap"]:
                        aggiornabili.append({
                            "row_idx":          et["row_idx"],
                            "date":             et["date"],
                            "amount":           et["amount"],
                            "descrizione":      et["descrizione"],
                            "moneymap_attuale": et["moneymap"],
                            "moneymap_nuova":   cat_money,
                        })
                    break
            if not trovato:
                solo_orig.append(bt)

        # "Solo elaborato" = righe del mese nel file elaborato senza corrispondenza
        # Se il mese non aveva righe nell'elaborato, lista vuota (non mostriamo altri mesi)
        elab_ref = elab if elab_per_solo is elab else elab_per_solo
        solo_elab = [elab_ref[i] for i in range(len(elab_ref)) if i not in usati_elab]

        return {
            "presenti_solo_originale": solo_orig,
            "presenti_solo_elaborato": solo_elab,
            "moneymap_aggiornabile":   aggiornabili,
            "n_originale":             len(bank_txns_originale),
            "n_elaborato":             len(elab_ref),
            "n_abbinati":              len(bank_txns_originale) - len(solo_orig),
        }

    def aggiorna_moneymap(self, aggiornamenti: list[dict],
                          sheet_name: str = "Movimenti",
                          col_moneymap: int = 5) -> int:
        """
        Aggiorna la colonna Moneymap nel file elaborato per le righe indicate.
        aggiornamenti: lista di dict con row_idx e moneymap_nuova.
        Ritorna il numero di righe aggiornate.
        ATTENZIONE: modifica il file Excel in-place — fare backup prima.
        """
        from openpyxl import load_workbook
        wb = load_workbook(self.xlsx_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return 0
        ws = wb[sheet_name]
        n  = 0
        for ag in aggiornamenti:
            row_idx = ag.get("row_idx")
            nuova   = ag.get("moneymap_nuova", "")
            if row_idx is None:
                continue
            ws.cell(row=row_idx, column=col_moneymap + 1).value = nuova
            n += 1
        wb.save(self.xlsx_path)
        wb.close()
        return n

    def riepilogo_mensile(self, sheet_name: str = "Movimenti",
                          anno: int | None = None) -> list[dict]:
        """
        Calcola totali mensili con saldo iniziale e finale dal file elaborato.
        Il saldo è letto direttamente dalla colonna saldo del foglio
        (ultima riga del mese = saldo finale; riga precedente alla prima
        del mese = saldo iniziale del mese).
        Ritorna lista di dict:
          {mese, anno, nome_mese, entrate, uscite, netto,
           saldo_iniziale, saldo_finale, n_transazioni}
        """
        from openpyxl import load_workbook
        from collections import defaultdict

        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]

        # Rileva colonna saldo (F o G) dall'header
        hdr = [ws.cell(1, c).value for c in range(1, 10)]
        hdr_str = [str(v or "").strip().lower() for v in hdr]
        has_moneymap = "moneymap" in hdr_str
        col_saldo = 7 if has_moneymap else 6   # G=7 per Movimenti, F=6 per Lombard/USD
        # Saldo iniziale anno: nella stessa riga 1, colonna successiva al nome "Saldo"
        # (es. Movimenti: H1, Lombard/USD: G1)
        col_saldo_iniziale = col_saldo + 1

        saldo_anno = ws.cell(1, col_saldo_iniziale).value
        try:
            saldo_anno = float(saldo_anno or 0)
        except (TypeError, ValueError):
            saldo_anno = 0.0

        # Leggi tutte le righe dati con data e saldo
        righe = []
        for r in range(2, ws.max_row + 1):
            dt_raw  = ws.cell(r, 1).value
            saldo_v = ws.cell(r, col_saldo).value
            entr_v  = ws.cell(r, 2).value
            usc_v   = ws.cell(r, 3).value
            if dt_raw is None:
                continue
            dt = _to_date(dt_raw)
            if dt is None:
                continue
            if anno and dt.year != anno:
                continue
            try:
                saldo_f = float(saldo_v) if saldo_v is not None else None
            except (TypeError, ValueError):
                saldo_f = None
            righe.append({
                "row": r,
                "date": dt,
                "deposit":    abs(_to_float(entr_v)),
                "withdrawal": abs(_to_float(usc_v)),
                "saldo": saldo_f,
            })
        wb.close()

        if not righe:
            return []

        # Raggruppa per mese
        from collections import defaultdict as _dd
        per_mese: dict[tuple, dict] = _dd(lambda: {
            "entrate": 0.0, "uscite": 0.0, "n": 0,
            "righe": []
        })
        for t in righe:
            k = (t["date"].year, t["date"].month)
            per_mese[k]["entrate"]  += t["deposit"]
            per_mese[k]["uscite"]   += t["withdrawal"]
            per_mese[k]["n"]        += 1
            per_mese[k]["righe"].append(t)

        result = []
        saldo_corrente = saldo_anno   # saldo prima della prima riga del file
        for (y, m), v in sorted(per_mese.items()):
            righe_mese = sorted(v["righe"], key=lambda t: t["date"])

            # Saldo iniziale del mese = saldo_corrente (saldo fine mese precedente)
            saldo_inizio = saldo_corrente

            # Saldo finale: ultima riga del mese con saldo non None
            saldo_fine = next(
                (t["saldo"] for t in reversed(righe_mese) if t["saldo"] is not None),
                None)
            if saldo_fine is None:
                # Fallback: calcola da saldo iniziale + variazione
                saldo_fine = round(saldo_inizio + v["entrate"] - v["uscite"], 2)

            result.append({
                "anno":           y,
                "mese":           m,
                "nome_mese":      MESI_IT_NOMI[m-1].capitalize(),
                "entrate":        round(v["entrate"], 2),
                "uscite":         round(v["uscite"], 2),
                "netto":          round(v["entrate"] - v["uscite"], 2),
                "saldo_iniziale": round(saldo_inizio, 2),
                "saldo_finale":   round(saldo_fine, 2),
                "n_transazioni":  v["n"],
            })
            saldo_corrente = saldo_fine   # diventa saldo iniziale del mese successivo

        return result

    def prepara_excel_scrittura(self, profilo: str = "") -> tuple[str, str]:
        """
        Prepara il file Excel per la scrittura sicura:
        1. Crea backup timestampato dentro la sottocartella _Backup_<PROFILO>
           (creata se non esiste) accanto al file originale.
           Se profilo è vuoto, usa "_Backup" senza suffisso.
        2. Crea copia di lavoro su cui scrivere (nella cartella originale).
        Ritorna (path_copia_lavoro, path_backup).
        """
        import shutil, datetime
        p  = Path(self.xlsx_path)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        backup_dir_name = f"_Backup_{profilo}" if profilo else "_Backup"
        backup_dir = p.parent / backup_dir_name
        backup_dir.mkdir(exist_ok=True)

        backup = backup_dir / f"{p.stem}_backup_{ts}{p.suffix}"
        work   = p.parent  / f"{p.stem}_work_{ts}{p.suffix}"
        shutil.copy2(self.xlsx_path, backup)
        shutil.copy2(self.xlsx_path, work)
        return str(work), str(backup)

    def scrivi_transazioni_cc(self, work_path: str,
                               transazioni: list[dict],
                               sheet_name: str = "Movimenti",
                               rules=None) -> tuple[int, str]:
        """
        Scrive le transazioni mancanti nel file elaborato CC.
        Gestisce tre strutture di colonne:

          Movimenti: A:Data B:Entr C:Usc D:Desc E:Desc2 F:Moneymap G:Saldo
          Lombard:   A:Data B:Entr C:Usc D:Desc E:Desc2 F:Saldo  (no Moneymap)
          USD:       uguale a Lombard

        La struttura (con/senza Moneymap, colonna saldo) viene rilevata
        automaticamente dall ultima riga esistente nel foglio.
        """
        from openpyxl import load_workbook
        from datetime import date as _date
        import re as _re

        wb = load_workbook(work_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return 0, f"Foglio {sheet_name!r} non trovato"

        ws = wb[sheet_name]

        # ── Rileva struttura colonne dal foglio ───────────────────────────
        # Header riga 1: cerca "Moneymap" per sapere se c e la colonna
        hdr = [ws.cell(1, c).value for c in range(1, 9)]
        hdr_str = [str(v or "").strip().lower() for v in hdr]
        has_moneymap = "moneymap" in hdr_str

        if has_moneymap:
            # Movimenti: A Data B Entr C Usc D Desc E Desc2 F Moneymap G Saldo
            col_desc      = 4   # D
            col_desc2     = 5   # E
            col_moneymap  = 6   # F
            col_saldo     = 7   # G
        else:
            # Lombard / USD: A Data B Entr C Usc D Desc E Desc2 F Saldo
            col_desc      = 4   # D
            col_desc2     = 5   # E
            col_moneymap  = None
            col_saldo     = 6   # F

        # ── Trova ultima riga con dati (colonna A non vuota) ─────────────
        last_row = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value is not None:
                last_row = r

        # ── Rileva la lettera di colonna saldo dall ultima formula ────────
        # Necessario perche la prima riga usa =G1+B2+C2 anche per Lombard/USD
        # (saldo iniziale e in G1), poi dalla R3 usa =F{prev}+...
        # Per sicurezza leggiamo la formula dell ultima riga effettiva.
        last_saldo_formula = ws.cell(last_row, col_saldo).value or ""
        m_col = _re.match(r"=([A-Z]+)\d+\+", str(last_saldo_formula))
        if m_col:
            # Formula tipo "=F14+B15+C15" — usa la colonna trovata
            saldo_col_letter = m_col.group(1)
        else:
            # Fallback: usa la colonna dichiarata dalla struttura header
            from openpyxl.utils import get_column_letter
            saldo_col_letter = get_column_letter(col_saldo)

        # ── Ordina per data crescente ─────────────────────────────────────
        def _get_dt(t):
            d = t.get("date")
            return (d if isinstance(d, _date) else _to_date(d)) or _date(2000, 1, 1)

        transazioni_ord = sorted(transazioni, key=_get_dt)

        # ── Scrittura righe ───────────────────────────────────────────────
        inserite = 0
        for t in transazioni_ord:
            dt     = _get_dt(t)
            amount = t.get("amount", 0.0) or 0.0

            entrate = amount if amount > 0 else None
            uscite  = amount if amount < 0 else None   # già negativo

            desc  = (t.get("descrizione") or t.get("Descrizione") or "")
            desc2 = (t.get("descrizione_completa") or t.get("Descrizione_Completa") or "")

            moneymap = ""
            if has_moneymap and rules:
                sg = rules.apply(f"{desc} {desc2}".strip())
                if sg:
                    moneymap = sg.get("category", "")

            nr = last_row + 1

            ws.cell(nr, 1).value = dt
            ws.cell(nr, 1).number_format = "DD/MM/YYYY"

            ws.cell(nr, 2).value = entrate
            if entrate is not None:
                ws.cell(nr, 2).number_format = "#,##0.00"

            ws.cell(nr, 3).value = uscite
            if uscite is not None:
                ws.cell(nr, 3).number_format = "#,##0.00"

            ws.cell(nr, col_desc).value  = desc
            ws.cell(nr, col_desc2).value = desc2

            if col_moneymap:
                ws.cell(nr, col_moneymap).value = moneymap

            # Formula saldo: =<colonna_prev>{last_row}+B{nr}+C{nr}
            ws.cell(nr, col_saldo).value = (
                f"={saldo_col_letter}{last_row}+B{nr}+C{nr}")
            ws.cell(nr, col_saldo).number_format = "#,##0.00"

            # Aggiorna la lettera colonna per la prossima riga
            saldo_col_letter = get_column_letter(col_saldo)                 if "get_column_letter" in dir() else saldo_col_letter

            last_row    = nr
            inserite   += 1

        wb.save(work_path)
        wb.close()
        return inserite, f"{inserite} righe inserite, ultima riga dati: {last_row}"

    def finalizza_excel(self, xlsx_originale: str, work_path: str) -> bool:
        """Sostituisce il file originale con la copia di lavoro verificata."""
        Path(work_path).replace(xlsx_originale)
        return True

    def scrivi_transazioni_cc_mensile(self, work_path: str,
                                       transazioni: list[dict],
                                       mese: int,
                                       sheet_map: dict | None = None) -> tuple[int, str]:
        """
        Scrive le transazioni mancanti nel foglio mensile del file elaborato
        MC o Visa (2026 mc.xlsx / 2026 visa.xlsx).

        Struttura colonne foglio mensile MC:
          A: Data operazione   B: Data registrazione   C: Descrizione
          D: Circuito          E: Importo

        Struttura colonne foglio mensile Visa (ha Intestatario in più):
          A: Intestatario carta   B: Data operazione   C: Data registrazione
          D: Descrizione          E: Importo

        La struttura viene rilevata automaticamente dall header del foglio.
        Non ci sono formule di saldo — solo dati grezzi da appendere.
        """
        from openpyxl import load_workbook
        from datetime import date as _date

        if sheet_map is None:
            sheet_map = {i+1: MESI_IT_NOMI[i] for i in range(12)}

        nome_foglio = sheet_map.get(mese, MESI_IT_NOMI[mese-1])

        wb = load_workbook(work_path)
        # Ricerca foglio case-insensitive
        found = None
        for s in wb.sheetnames:
            if s.lower().strip() == nome_foglio.lower():
                found = s
                break

        if found is None:
            # Il foglio del mese non esiste ancora — lo creiamo copiando
            # la struttura header da un foglio esistente dello stesso file,
            # oppure usando la struttura standard MC/Visa.
            ws_new = wb.create_sheet(title=nome_foglio)
            # Cerca un foglio modello (qualsiasi mese già presente)
            template = None
            for s in wb.sheetnames:
                if s != nome_foglio and s.lower() in MESI_IT_NOMI:
                    template = wb[s]
                    break
            if template is not None:
                # Copia header dal foglio modello
                hdr_vals = [template.cell(1, c).value for c in range(1, 8)]
                for c, v in enumerate(hdr_vals, 1):
                    ws_new.cell(1, c).value = v
                    # Copia larghezza colonna se disponibile
                    col_letter = template.cell(1, c).column_letter
                    if col_letter in template.column_dimensions:
                        ws_new.column_dimensions[ws_new.cell(1, c).column_letter].width = \
                            template.column_dimensions[col_letter].width
            else:
                # Struttura standard MC (fallback)
                for c, h in enumerate(["Data operazione", "Data registrazione",
                                       "Descrizione", "Circuito", "Importo"], 1):
                    ws_new.cell(1, c).value = h
            wb.save(work_path)
            found = nome_foglio

        ws = wb[found]

        # Rileva struttura dall header (riga 1)
        hdr = [str(ws.cell(1,c).value or "").strip().lower() for c in range(1,8)]
        has_intestatario = "intestatario carta" in hdr

        if has_intestatario:
            # Visa: A Intestatario  B DataOp  C DataReg  D Desc  E Importo
            col_intestatario = 1
            col_data_op      = 2
            col_data_reg     = 3
            col_desc         = 4
            col_importo      = 5
        else:
            # MC:  A DataOp  B DataReg  C Desc  D Circuito  E Importo
            col_intestatario = None
            col_data_op      = 1
            col_data_reg     = 2
            col_desc         = 3
            col_circuito     = 4
            col_importo      = 5

        # Trova ultima riga con dati
        last_row = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, col_data_op).value is not None:
                last_row = r

        # Ordina per data operazione crescente
        def _get_dt(t):
            d = t.get("date")
            return (d if isinstance(d, _date) else _to_date(d)) or _date(2000,1,1)

        transazioni_ord = sorted(transazioni, key=_get_dt)

        inserite = 0
        for t in transazioni_ord:
            dt_op  = _get_dt(t)
            dt_reg = t.get("date_registrazione")
            if dt_reg and not isinstance(dt_reg, _date):
                dt_reg = _to_date(dt_reg)

            amount = t.get("amount", 0.0) or 0.0
            desc   = (t.get("descrizione") or t.get("Descrizione") or "")
            intest = (t.get("intestatario") or "")
            nr     = last_row + 1

            if col_intestatario:
                ws.cell(nr, col_intestatario).value = intest
            ws.cell(nr, col_data_op).value  = dt_op
            ws.cell(nr, col_data_op).number_format = "DD/MM/YYYY"
            ws.cell(nr, col_data_reg).value = dt_reg
            if dt_reg:
                ws.cell(nr, col_data_reg).number_format = "DD/MM/YYYY"
            ws.cell(nr, col_desc).value    = desc
            if not col_intestatario:
                # MC: scrivi circuito (dal campo originale, default MASTERCARD)
                ws.cell(nr, col_circuito).value = (
                    t.get("circuito") or "MASTERCARD")
            ws.cell(nr, col_importo).value         = amount
            ws.cell(nr, col_importo).number_format = "#,##0.00"

            last_row  = nr
            inserite += 1

        wb.save(work_path)
        wb.close()
        return inserite, f"{inserite} righe inserite nel foglio '{found}', ultima riga: {last_row}"

    def confronta_cc_mensile(self, bank_txns_originale: list[dict],
                              mese: int, anno: int,
                              sheet_map: dict | None = None) -> dict:
        """
        Come confronta_cc() ma per i file mensili MC/Visa (12 fogli).
        """
        elab = leggi_excel_elaborato_cc_mensile(
            self.xlsx_path, mese, anno, sheet_map)
        usati = set()
        solo_orig = []
        for bt in bank_txns_originale:
            trovato = False
            for i, et in enumerate(elab):
                if i in usati:
                    continue
                if (_amt_eq(bt["amount"], et["amount"], 0.02) and
                        _date_ok(bt["date"], et["date"], 2)):
                    usati.add(i)
                    trovato = True
                    break
            if not trovato:
                solo_orig.append(bt)
        solo_elab = [elab[i] for i in range(len(elab)) if i not in usati]
        return {
            "presenti_solo_originale": solo_orig,
            "presenti_solo_elaborato": solo_elab,
            "n_originale":             len(bank_txns_originale),
            "n_elaborato":             len(elab),
            "n_abbinati":              len(bank_txns_originale) - len(solo_orig),
        }