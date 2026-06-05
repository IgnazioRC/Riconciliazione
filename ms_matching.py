"""
ms_matching.py — Motore di riconciliazione e regole
                ReconcileEngine  — algoritmo di matching a 7 livelli
                RulesEngine      — regole pattern→categoria
                raggruppa_*      — raggruppamento cedole e transazioni correlate
                costruisci_transazioni_da_risultati
                export_full_csv / export_missing_csv
Parte di: Moneyspire Reconciler
"""

import re
import csv
import json
import sqlite3
from pathlib import Path
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from collections import Counter

from ms_constants import (
    MATCH_EXACT, MATCH_FUZZY, MATCH_SPLIT, MATCH_MERGE,
    MATCH_NONE, MATCH_SKIP, MATCH_PENDING,
    STATO_LABELS, DEFAULT_CONFIG,
    _to_date, _to_float, fmt_eur,
    _amt_eq, _date_ok,
)
import ms_db
from ms_db import MoneyspireDB
from ms_parsers import (
    parse_fineco_conto_originale, parse_fineco_conto, parse_fineco_conto_usd,
    parse_fineco_cc, leggi_numero_conto_fineco,
)

# ─────────────────────────────────────────────────────────────────────────────
# MATCHING ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def _sim(a: str, b: str) -> float:
    a = re.sub(r'\s+', ' ', (a or "").lower().strip())
    b = re.sub(r'\s+', ' ', (b or "").lower().strip())
    return SequenceMatcher(None, a, b).ratio() if a and b else 0.0


def _correggi_fuzzy_invertiti(results: list[dict], cfg: dict) -> list[dict]:
    """
    Post-processing: verifica le coppie di match fuzzy con importi quasi identici
    e le corregge se le regole suggeriscono la combinazione opposta.

    Algoritmo:
    1. Raccoglie tutti i risultati fuzzy (MATCH_FUZZY e MATCH_EXACT con Δ importo > 0)
    2. Per ogni coppia con importi quasi identici (Δ ≤ soglia), calcola un punteggio
       di "compatibilità" basato sulla somiglianza tra:
       - descrizione banca → categoria suggerita dalla regola
       - categoria Money già presente
    3. Se la combinazione inversa ha punteggio migliore, scambia gli abbinamenti
    """
    from ms_constants import MATCH_FUZZY, MATCH_EXACT
    _fz = cfg.get("tolleranze_fuzzy", {})
    soglia_delta = _fz.get("importo_max_eur", 1.0)  # max Δ per considerare "quasi identici"

    # Indici dei risultati fuzzy con una sola transazione Money abbinata
    candidati = [
        i for i, r in enumerate(results)
        if r.get("match_type") in (MATCH_FUZZY, MATCH_EXACT)
        and len(r.get("money_txns", [])) == 1
        and not r.get("annullata")
    ]

    def _score(bank_desc: str, money_cat: str, money_memo: str, money_payee: str) -> float:
        """Similarità testo banca vs contesto Money (categoria + memo + payee)."""
        ctx = f"{money_cat} {money_memo} {money_payee}".lower()
        return SequenceMatcher(None, bank_desc.lower(), ctx).ratio()

    # Controlla tutte le coppie di candidati
    visited = set()
    for ii in range(len(candidati)):
        if candidati[ii] in visited:
            continue
        for jj in range(ii + 1, len(candidati)):
            if candidati[jj] in visited:
                continue

            ri = results[candidati[ii]]
            rj = results[candidati[jj]]

            amt_i = abs(ri["bank_txn"]["amount"])
            amt_j = abs(rj["bank_txn"]["amount"])

            # Considera solo coppie con importi quasi identici
            if abs(amt_i - amt_j) > soglia_delta:
                continue

            mt_i = ri["money_txns"][0]
            mt_j = rj["money_txns"][0]

            raw_i = ri["bank_txn"].get("raw_text", ri["bank_txn"].get("descrizione", ""))
            raw_j = rj["bank_txn"].get("raw_text", rj["bank_txn"].get("descrizione", ""))

            # Punteggio configurazione attuale
            score_cur = (
                _score(raw_i, mt_i["category"], mt_i["memo"], mt_i["payee"]) +
                _score(raw_j, mt_j["category"], mt_j["memo"], mt_j["payee"])
            )

            # Punteggio configurazione inversa
            score_inv = (
                _score(raw_i, mt_j["category"], mt_j["memo"], mt_j["payee"]) +
                _score(raw_j, mt_i["category"], mt_i["memo"], mt_i["payee"])
            )

            # Scambia solo se il miglioramento è significativo (≥ 10%)
            if score_inv > score_cur * 1.10:
                # Scambia le transazioni Money tra i due risultati
                results[candidati[ii]]["money_txns"] = [mt_j]
                results[candidati[ii]]["money_ids"]  = [mt_j["id"]]
                results[candidati[jj]]["money_txns"] = [mt_i]
                results[candidati[jj]]["money_ids"]  = [mt_i["id"]]
                # Aggiorna la nota
                results[candidati[ii]]["note"] += " [corretto inversione]"
                results[candidati[jj]]["note"] += " [corretto inversione]"
                visited.add(candidati[ii])
                visited.add(candidati[jj])

    return results


class ReconcileEngine:

    def __init__(self, db: MoneyspireDB, config: dict | None = None):
        self.db  = db
        self.cfg = config or DEFAULT_CONFIG
        self.tol_days   = self.cfg.get("date_tolerance_days", 3)
        self.tol_amount = self.cfg.get("amount_tolerance", 0.01)
        self._account_id = 0  # impostato in reconcile()

    def reconcile(self, bank_txns: list[dict], money_txns: list[dict],
                  account_id: int = 0) -> list[dict]:
        self._account_id = account_id
        used_ids: set[int] = set()
        results = []
        for bt in bank_txns:
            gruppo = bt.get("gruppo", "")

            # Transazioni annullate (Canone+Sconto, ecc.) → ignora silenziosamente
            if gruppo == "annullato":
                r_ann = self._make(bt, [], MATCH_SKIP, 1.0,
                    bt.get("gruppo_note","Transazione annullata dalla controparte"))
                r_ann["annullata"] = True
                results.append(r_ann)
                continue

            # Accessori CBILL (commissione/abbuono) → Solo Money con nota
            if gruppo == "accessorio":
                r_acc = self._make(bt, [], MATCH_SKIP, 1.0,
                    bt.get("gruppo_note","Accessorio — gestito con la transazione principale"))
                results.append(r_acc)
                continue

            # Split interessi portafoglio: cerca in Money come coppia
            # Se non trovata, la segnala come Mancante con nota
            if gruppo == "split_interessi":
                # Tenta matching normale (potrebbero essere già in Money come split)
                r = self._find(bt, money_txns, used_ids)
                if r["match_type"] == MATCH_NONE:
                    # Non trovata — è davvero mancante, ma con contesto
                    r["note"] = bt.get("gruppo_note", r["note"])
                results.append(r)
                for mid in r["money_ids"]:
                    used_ids.add(mid)
                continue

            # Merge proventi: usa la somma del gruppo per il matching
            if gruppo == "merge_proventi":
                totale = bt.get("gruppo_totale", bt["amount"])
                # Cerca in Money la transazione con importo = totale gruppo
                mt_found = None
                for mt in [m for m in money_txns if m["id"] not in used_ids]:
                    if (_amt_eq(totale, mt["amount"], self.tol_amount) and
                            _date_ok(bt["date"], mt["date"], self.tol_days)):
                        mt_found = mt
                        break
                if mt_found:
                    results.append(self._make(bt, [mt_found], MATCH_MERGE, 0.90,
                        f"Parte di gruppo proventi (totale {fmt_eur(totale)}) "
                        f"→ trovata in Money come unica transazione"))
                    used_ids.add(mt_found["id"])
                    continue
                # Se non trovata, cade nel matching normale

            r = self._find(bt, money_txns, used_ids)
            results.append(r)
            for mid in r["money_ids"]:
                used_ids.add(mid)
        # Transazioni solo in Money (caricate manualmente)
        for mt in money_txns:
            if mt["id"] not in used_ids:
                results.append(self._make(None, [mt], MATCH_SKIP, 1.0,
                                          "Caricata manualmente in Moneyspire"))

        # Identifica coppie Solo Money che si annullano (es. Canone+Sconto)
        # Le marca come MATCH_SKIP con nota "annullate" — utente può filtrarle
        import re as _re3
        skip_results = [r for r in results if r["match_type"] == MATCH_SKIP
                        and r["money_txns"]]
        used_annull = set()
        for i, ra in enumerate(skip_results):
            if i in used_annull:
                continue
            mt_a = ra["money_txns"][0]
            for j, rb in enumerate(skip_results):
                if j <= i or j in used_annull:
                    continue
                mt_b = rb["money_txns"][0]
                # Stessa data ±1gg, importi opposti, somma = 0
                if (abs((mt_a["date"] - mt_b["date"]).days) <= 1 and
                        _amt_eq(mt_a["amount"] + mt_b["amount"], 0.0, 0.02)):
                    nota = (f"Transazione annullata dalla controparte "
                            f"({fmt_eur(mt_b['amount'])} del {mt_b['date'].strftime('%d/%m')})")
                    ra["note"] = nota
                    rb["note"] = (f"Transazione annullata dalla controparte "
                                  f"({fmt_eur(mt_a['amount'])} del {mt_a['date'].strftime('%d/%m')})")
                    # Marca entrambe come "annullate" nel tag gruppo
                    ra["annullata"] = True
                    rb["annullata"] = True
                    used_annull.update([i, j])
                    break

        # ── Post-processing cedole/dividendi ────────────────────────────────
        # Situazione attuale dopo il matching principale:
        #
        # CEDOLE BTP (già ok):
        #   Lordo banca → MERGE (trova Money netto con splits)
        #   Ritenuta banca → MATCH_NONE → qui la promuoviamo a SPLIT
        #
        # DIVIDENDI AZIONARI (problema inverso):
        #   Ritenuta banca → MERGE (il livello 4 abbina la ritenuta allo split)
        #   Lordo banca → MATCH_NONE → dobbiamo INVERTIRE: lordo=MERGE, ritenuta=SPLIT
        #
        import re as _re2
        pat_rit   = _re2.compile(r'ritenuta|rit[.]', _re2.IGNORECASE)
        pat_lordo = _re2.compile(r'dividendo italia|dividendo estero|div\.su|stacco cedol',
                                  _re2.IGNORECASE)

        # Passo A: correggi i dividendi (ritenuta→MERGE, lordo→NONE)
        # Trova coppie lordo+ritenuta dove la ritenuta è già MERGE
        for r_rit in results:
            if r_rit["match_type"] != MATCH_MERGE:
                continue
            bt_rit = r_rit.get("bank_txn")
            if not bt_rit or bt_rit["amount"] >= 0:
                continue
            if not pat_rit.search(bt_rit.get("raw_text", "")):
                continue
            # Cerca il lordo corrispondente (MATCH_NONE, stessa data ±3gg)
            mt_rit = r_rit["money_txns"][0] if r_rit["money_txns"] else None
            if not mt_rit or not mt_rit["has_splits"]:
                continue
            rit_amt = abs(bt_rit["amount"])
            for r_lordo in results:
                if r_lordo["match_type"] != MATCH_NONE:
                    continue
                bt_lordo = r_lordo.get("bank_txn")
                if not bt_lordo or bt_lordo["amount"] <= 0:
                    continue
                if not _date_ok(bt_rit["date"], bt_lordo["date"], self.tol_days):
                    continue
                if not pat_lordo.search(bt_lordo.get("raw_text", "")):
                    continue
                # Verifica che lordo - ritenuta ≈ netto Money
                lordo_amt = bt_lordo["amount"]
                netto_atteso = lordo_amt - rit_amt
                if _amt_eq(netto_atteso, mt_rit["amount"], 0.10):
                    # Inverti: lordo diventa MERGE, ritenuta diventa SPLIT
                    r_lordo["match_type"]         = MATCH_MERGE
                    r_lordo["money_txns"]         = [mt_rit]
                    r_lordo["money_ids"]          = [mt_rit["id"]]
                    r_lordo["confidence"]         = 0.95
                    r_lordo["note"]               = (f"Dividendo lordo — in Money come "
                                                     f"transazione split "
                                                     f"(netto {fmt_eur(mt_rit['amount'])})")
                    r_rit["match_type"]           = MATCH_SPLIT
                    r_rit["note"]                 = (f"Ritenuta inclusa nello split "
                                                     f"del {mt_rit['date'].strftime('%d/%m')} "
                                                     f"(netto {fmt_eur(mt_rit['amount'])})")
                    r_rit["suggested_category"]   = next(
                        (sp["category"] for sp in mt_rit["splits"]
                         if _amt_eq(sp["withdrawal"], rit_amt, 0.10)), "")
                    break

        # Passo B: ritenute cedole (MATCH_NONE) → MATCH_SPLIT
        merges = [(r2, r2["money_txns"][0]) for r2 in results
                  if r2["match_type"] == MATCH_MERGE
                  and r2["money_txns"]
                  and r2["money_txns"][0]["has_splits"]]
        for r in results:
            if r["match_type"] != MATCH_NONE:
                continue
            bt = r["bank_txn"]
            if not bt or bt["amount"] >= 0:
                continue
            if not pat_rit.search(bt.get("raw_text", "")):
                continue
            rit_amt = abs(bt["amount"])
            for r2, mt2 in merges:
                if not _date_ok(bt["date"], mt2["date"], self.tol_days):
                    continue
                for sp in mt2["splits"]:
                    if _amt_eq(sp["withdrawal"], rit_amt, self.tol_amount):
                        r["match_type"]         = MATCH_SPLIT
                        r["money_txns"]         = [mt2]
                        r["money_ids"]          = [mt2["id"]]
                        r["confidence"]         = 0.95
                        r["note"]               = (f"Ritenuta inclusa nello split "
                                                   f"del {mt2['date'].strftime('%d/%m')} "
                                                   f"(netto {fmt_eur(mt2['amount'])})")
                        r["suggested_category"] = sp["category"]
                        break
                if r["match_type"] == MATCH_SPLIT:
                    break

        # ── Post-processing: correggi abbinamenti fuzzy invertiti ────────
        # Quando due transazioni banca con importi quasi identici vengono
        # abbinate a due transazioni Money in modo invertito rispetto a quello
        # suggerito dalle regole di categorizzazione, le scambia.
        # Esempio: 251,20 (fattura Farina) abbinata a "Acquisto titoli" in Money
        #          250,95 (acquisto fondi) abbinata a "BdS:lavori25" in Money
        # → le regole suggeriscono la combinazione opposta, quindi le scambiamo.
        results = _correggi_fuzzy_invertiti(results, self.cfg)

        return results

    def _find(self, bt: dict, money_txns: list[dict], used_ids: set) -> dict:
        avail  = [m for m in money_txns if m["id"] not in used_ids]
        bt_amt = bt["amount"]
        bt_dt  = bt["date"]

        # 1. Esatto
        for mt in avail:
            # Guard: scarta candidati con importo zero o segno opposto
            if mt["amount"] == 0 or (bt_amt * mt["amount"] < 0):
                continue
            if _amt_eq(bt_amt, mt["amount"], self.tol_amount) and bt_dt == mt["date"]:
                return self._make(bt, [mt], MATCH_EXACT, 1.0, "Importo e data identici")

        # 1b. Stessa data, importo quasi identico (tolleranza configurabile)
        #     Copre arrotondamenti e micro-differenze su carte di credito
        #     (es. 25,89 banca vs 25,80 Money = commissione 0,09€)
        #     Configurabile in ms_config.json → "tolleranze_fuzzy"
        _fz = self.cfg.get("tolleranze_fuzzy", {})
        _fz_pct = _fz.get("importo_pct", 0.5) / 100.0
        _fz_max = _fz.get("importo_max_eur", 1.0)
        for mt in avail:
            if bt_dt != mt["date"]:
                continue
            # Guard: segno deve coincidere (entrambi entrata o entrambi uscita)
            if mt["amount"] == 0 or (bt_amt * mt["amount"] < 0):
                continue
            ref  = abs(bt_amt) if bt_amt else 1.0
            diff = abs(abs(bt_amt) - abs(mt["amount"]))
            tol  = max(ref * _fz_pct, _fz_max)
            if diff <= tol and diff > self.tol_amount:
                nota = (f"Importo quasi identico (Δ {diff:.2f} €), "
                        f"data identica")
                return self._make(bt, [mt], MATCH_FUZZY, 0.92, nota)

        # 1c. Data ±N giorni + differenza importo ≤ soglia commissione bancaria
        #     Copre il caso tipico bonifico Unicredit: banca addebita 0,31€ commissione
        #     che non compare in Money (es. -5.000,31 banca vs -5.000,00 Money, data ±1gg)
        #     Soglia configurabile in ms_config.json → "tolleranze_fuzzy.commissione_max_eur"
        _comm_max = _fz.get("commissione_max_eur", 0.50)
        _comm_days = _fz.get("commissione_giorni", 2)
        for mt in avail:
            days = abs((bt_dt - mt["date"]).days)
            if days > _comm_days:
                continue
            # Guard: segno deve coincidere (entrambi entrata o entrambi uscita)
            if mt["amount"] == 0 or (bt_amt * mt["amount"] < 0):
                continue
            diff = abs(abs(bt_amt) - abs(mt["amount"]))
            if self.tol_amount < diff <= _comm_max:
                nota = (f"Importo quasi identico (Δ {diff:.2f} €, prob. commissione), "
                        f"data sfasata di {days} gg")
                return self._make(bt, [mt], MATCH_FUZZY, 0.88, nota)

        # 2. Fuzzy (importo + data ±N gg)
        # Guard segno: scarta candidati con segno opposto (es. +105 entrata vs -105 uscita)
        fuzzy = [m for m in avail
                 if _amt_eq(bt_amt, m["amount"], self.tol_amount)
                 and _date_ok(bt_dt, m["date"], self.tol_days)
                 and m["amount"] != 0
                 and bt_amt * m["amount"] >= 0]
        if fuzzy:
            fuzzy.sort(
                key=lambda m: _sim(bt["raw_text"],
                                   f"{m['memo']} {m['category']} {m['payee']}"),
                reverse=True)
            best = fuzzy[0]
            days = abs((bt_dt - best["date"]).days)
            conf = min(0.80 + 0.15 * _sim(bt["raw_text"],
                                           f"{best['memo']} {best['category']}"), 0.95)
            return self._make(bt, [best], MATCH_FUZZY, conf,
                              f"Importo coincide, data sfasata di {days} gg")

        # 3. Split: banca = netto di transazione splittata
        for mt in avail:
            if not mt["has_splits"]:
                continue
            net = sum(s["deposit"] - s["withdrawal"] for s in mt["splits"])
            # Guard: il netto split deve avere lo stesso segno dell'importo banca
            if net == 0 or (bt_amt * net < 0):
                continue
            if _amt_eq(bt_amt, net, self.tol_amount) and _date_ok(bt_dt, mt["date"], self.tol_days):
                return self._make(bt, [mt], MATCH_SPLIT, 0.88,
                                  f"Corrisponde a transazione con {len(mt['splits'])} splits")

        # 4. Merge: banca = uno split di transazione composta
        for mt in avail:
            if not mt["has_splits"] or not _date_ok(bt_dt, mt["date"], self.tol_days):
                continue
            for sp in mt["splits"]:
                if _amt_eq(bt_amt, sp["deposit"] - sp["withdrawal"], self.tol_amount):
                    return self._make(bt, [mt], MATCH_MERGE, 0.72,
                                      "Corrisponde a una parte di transazione composta")

        # 5. Match per categoria + data, tolleranza importo percentuale
        #    Per transazioni a importo variabile mese/mese (es. pensione, stipendio)
        #    Configurabile in ms_config.json → "amount_tolerance_pct_by_category"
        #    Guard aggiuntivo: l'importo deve comunque essere "ragionevole", cioè
        #    non può essere zero o di segno opposto rispetto alla transazione banca.
        cat_tol_pct = self.cfg.get("amount_tolerance_pct_by_category", {})
        for mt in avail:
            if not _date_ok(bt_dt, mt["date"], self.tol_days):
                continue
            if not mt["category"]:
                continue
            # Guard: segno deve coincidere (entrambi entrata o entrambi uscita)
            if mt["amount"] == 0 or (bt_amt * mt["amount"] < 0):
                continue
            cat_root = mt["category"].split(":")[0]
            tol_pct = cat_tol_pct.get(mt["category"],
                      cat_tol_pct.get(cat_root, None))
            if tol_pct is None:
                continue
            ref_amt = abs(mt["amount"]) if mt["amount"] else 1
            diff_pct = abs(abs(bt_amt) - abs(mt["amount"])) / ref_amt * 100
            if diff_pct <= tol_pct:
                diff_eur = abs(abs(bt_amt) - abs(mt["amount"]))
                cat_label = mt["category"]
                return self._make(
                    bt, [mt], MATCH_FUZZY,
                    max(0.60, 0.85 - diff_pct / 100),
                    f"Categoria '{cat_label}':"
                    f" importo simile (Δ {diff_eur:.2f} €, {diff_pct:.1f}%)"
                )

        # 6. Trasferimento tra conti / Cambio valuta
        #    Cerca su tutti gli altri conti una transazione con CategoryID = account corrente
        #    che abbia data compatibile e importo compatibile (esatto per giroconti EUR,
        #    tolleranza % per cambi valuta).
        #    NON richiede testo specifico nella descrizione — copre qualsiasi trasferimento.
        if self._account_id:   # guard: salta se account_id non è stato passato
            try:
                conn6 = sqlite3.connect(self.db.path)
                cur6  = conn6.cursor()
                tol_pct_cambio = self.cfg.get("amount_tolerance_pct_by_category",
                                              {}).get("Cambio valuta", 5.0)
                tol_days = self.tol_days
                d_from = (bt["date"] - __import__("datetime").timedelta(days=tol_days)).isoformat()
                d_to   = (bt["date"] + __import__("datetime").timedelta(days=tol_days)).isoformat()
                cur6.execute("""
                    SELECT t.ID, DATE(t.TransactionDate) as dt,
                           COALESCE(t.Withdrawal,0) as w,
                           COALESCE(t.Deposit,0)    as d,
                           a.Name as acc_name,
                           COALESCE(t.Rate, 1.0)    as rate,
                           t.Memo
                    FROM Transactions t
                    JOIN Accounts a ON a.ID = t.AccountID
                    WHERE t.CategoryID = ?
                      AND DATE(t.TransactionDate) BETWEEN ? AND ?
                """, (self._account_id, d_from, d_to))
                rows6 = cur6.fetchall()
                conn6.close()
                conn6.close()

                for r6 in rows6:
                    w6   = r6[2]
                    rate = r6[5] if r6[5] else 1.0
                    amt_eur    = w6 * rate if rate != 1.0 else w6
                    diff_eur   = abs(abs(bt_amt) - abs(amt_eur))
                    ref        = abs(bt_amt) if bt_amt else 1.0
                    diff_pct   = diff_eur / ref * 100
                    is_cambio  = (rate != 1.0)
                    tol = tol_pct_cambio if is_cambio else 0.05
                    if diff_pct <= tol:
                        tipo = "Cambio valuta" if is_cambio else "Giroconto"
                        mt_giro = {
                            "id":          r6[0],
                            "date":        bt["date"],
                            "amount":      bt["amount"],
                            "withdrawal":  0.0,
                            "deposit":     bt["amount"],
                            "memo":        r6[6] or f"{tipo} da {r6[4]}",
                            "category":    f"→ {r6[4]}",
                            "category_id": None,
                            "payee": "", "payee_id": None,
                            "status": 0, "splits": [], "has_splits": False
                        }
                        nota = (f"{tipo}: controparte su {r6[4]} "
                                f"(Δ {diff_eur:.2f} € = {diff_pct:.2f}%)")
                        return self._make(bt, [mt_giro], MATCH_EXACT, 0.95, nota)
            except Exception as _e6:
                print(f"[livello6] ERRORE: {_e6} — bt={bt.get('descrizione','')} "
                      f"account_id={self._account_id}")

        # 7. Nessun match
        return self._make(bt, [], MATCH_NONE, 0.0, "Non trovata in Moneyspire")

    @staticmethod
    def _make(bt, money_list, mtype, conf, note):
        mt = money_list[0] if money_list else None
        return {
            "bank_txn": bt, "money_txns": money_list,
            "money_ids": [m["id"] for m in money_list],
            "match_type": mtype, "confidence": conf, "note": note,
            "suggested_category": mt["category"] if mt else "",
            "suggested_payee":    mt["payee"]    if mt else ""
        }


# ─────────────────────────────────────────────────────────────────────────────
# REGOLE
# ─────────────────────────────────────────────────────────────────────────────

class RulesEngine:

    def __init__(self, rules_path: str):
        self.path  = Path(rules_path)
        self.rules: list[dict] = []
        self.load()

    def load(self):
        self.rules = json.loads(self.path.read_text(encoding="utf-8")) \
            if self.path.exists() else []

    def save(self):
        self.path.write_text(
            json.dumps(self.rules, ensure_ascii=False, indent=2),
            encoding="utf-8")

    def apply(self, raw_text: str) -> dict | None:
        text = (raw_text or "").lower()
        for rule in self.rules:
            pat = rule.get("pattern", "")
            if not pat:
                continue
            try:
                if rule.get("regex"):
                    matched = bool(re.search(pat, text, re.IGNORECASE))
                elif "*" in pat or "?" in pat:
                    # Wildcard semplice: * = qualsiasi sequenza, ? = un carattere
                    import fnmatch
                    matched = fnmatch.fnmatch(text, pat.lower())
                else:
                    matched = pat.lower() in text
            except re.error:
                matched = False
            if matched:
                rule["hits"] = rule.get("hits", 0) + 1
                # Restituisce tutti i campi utili della regola (incluso _transfer_to)
                return {
                    "category":    rule.get("category", ""),
                    "payee":       rule.get("payee", ""),
                    "_transfer_to": rule.get("_transfer_to", "")
                }
        return None

    def add_or_update(self, pattern: str, category: str,
                      payee: str = "", regex: bool = False, source: str = "manual"):
        for r in self.rules:
            if r.get("pattern", "").lower() == pattern.lower():
                r.update({"category": category, "payee": payee, "source": source})
                self.save()
                return
        self.rules.insert(0, {"pattern": pattern, "regex": regex,
                               "category": category, "payee": payee,
                               "source": source, "hits": 0})
        self.save()

    def remove(self, index: int):
        if 0 <= index < len(self.rules):
            self.rules.pop(index)
            self.save()

    def move_up(self, index: int):
        if index > 0:
            self.rules[index-1], self.rules[index] = self.rules[index], self.rules[index-1]
            self.save()

    def learn_from_history(self, money_txns: list[dict]) -> int:
        pairs = Counter((t["payee"], t["category"])
                        for t in money_txns if t["payee"] and t["category"])
        learned = 0
        for (payee, category), count in pairs.most_common(100):
            if count < 2:
                break
            if not any(r.get("pattern", "").lower() == payee.lower() for r in self.rules):
                self.rules.append({"pattern": payee, "regex": False,
                                   "category": category, "payee": payee,
                                   "source": "auto", "hits": count})
                learned += 1
        if learned:
            self.save()
        return learned




# ─────────────────────────────────────────────────────────────────────────────
# RICONOSCIMENTO COPPIE CEDOLA/RITENUTA
# ─────────────────────────────────────────────────────────────────────────────

def marca_in_attesa(bank_txns: list[dict],
                    giorni_fine_mese: int = 4) -> list[dict]:
    """
    Per le carte di credito, marca le transazioni banca negli ultimi N giorni
    del mese come "pending": hanno data operazione nel mese corrente ma data
    registrazione nel mese successivo, quindi appaiono nell'estratto del mese
    dopo. Vengono marcate per escluderle dal matching normale.
    """
    import calendar
    for t in bank_txns:
        dt = t["date"]
        ultimo_giorno = calendar.monthrange(dt.year, dt.month)[1]
        giorni_dalla_fine = ultimo_giorno - dt.day
        if giorni_dalla_fine < giorni_fine_mese:
            t["pending"] = True
    return bank_txns


def marca_solo_money_in_attesa(results: list[dict],
                                giorni_fine_mese: int = 4) -> list[dict]:
    """
    Per i risultati già calcolati: le transazioni "Solo Money" con data
    negli ultimi N giorni del mese potrebbero avere la controparte banca
    nell'estratto del mese successivo. Le rimarca come MATCH_PENDING
    con nota esplicativa.
    """
    import calendar
    for r in results:
        if r["match_type"] != MATCH_SKIP:
            continue
        mt = r["money_txns"][0] if r["money_txns"] else None
        if not mt:
            continue
        dt = mt["date"]
        ultimo_giorno = calendar.monthrange(dt.year, dt.month)[1]
        giorni_dalla_fine = ultimo_giorno - dt.day
        if giorni_dalla_fine < giorni_fine_mese:
            r["match_type"] = MATCH_PENDING
            r["note"] = (f"Caricata il {dt.strftime('%d/%m')} "
                         f"({giorni_dalla_fine+1}° giorno dalla fine mese) — "
                         f"la controparte banca è probabilmente nell'estratto "
                         f"del mese successivo")
    return results


def raggruppa_cedole_ritenute(bank_txns: list[dict],
                               finestra_giorni: int = 1) -> list[dict]:
    """
    Identifica coppie (lordo, ritenuta) nelle transazioni banca:
    - Cedole BTP/Estero: "Stacco Cedole" + "Ritenuta su Cedole" (12.5% o 26%)
    - Dividendi azionari: "Dividendo Italia/Estero" + "Ritenuta dividendo Italia"
      abbinati tramite titolo identico nella descrizione completa
    Ritorna la lista originale con le coppie marcate.
    """
    import re
    # Pattern cedole obbligazionarie
    pat_cedola   = re.compile(r'cedol|stacco cedol|ced\.su', re.IGNORECASE)
    pat_rit_ced  = re.compile(r'ritenuta su cedole|rit\.ced\.su', re.IGNORECASE)
    # Pattern dividendi azionari
    pat_dividendo = re.compile(r'dividendo italia|dividendo estero|div\.su', re.IGNORECASE)
    pat_rit_div   = re.compile(r'ritenuta dividendo|rit\.div\.su', re.IGNORECASE)

    def titolo_da_desc2(desc2: str) -> str:
        """Estrae il nome del titolo dalla descrizione completa per abbinamento."""
        # Es: "Div.su 600,000 UNICREDIT" → "UNICREDIT"
        # Es: "Rit.div.su 600,000 UNICREDIT" → "UNICREDIT"
        m = re.search(r'su\s+[\d\.,]+\s+(\S+)', desc2, re.IGNORECASE)
        return m.group(1).upper() if m else ""

    # ── Cedole obbligazionarie ──────────────────────────────────────────
    cedole   = [t for t in bank_txns
                if pat_cedola.search(t.get("raw_text","")) and t["amount"] > 0
                and not pat_rit_ced.search(t.get("raw_text",""))]
    ritenute_ced = [t for t in bank_txns
                    if pat_rit_ced.search(t.get("raw_text","")) and t["amount"] < 0]

    used = set()
    for ced in cedole:
        for i, rit in enumerate(ritenute_ced):
            if id(rit) in used:
                continue
            if abs((ced["date"] - rit["date"]).days) > finestra_giorni:
                continue
            effettiva    = abs(rit["amount"])
            cedola_lorda = abs(ced["amount"])
            match_12 = abs(effettiva - cedola_lorda * 0.125) / (cedola_lorda * 0.125) < 0.10
            match_26 = abs(effettiva - cedola_lorda * 0.26)  / (cedola_lorda * 0.26)  < 0.10
            if match_12 or match_26:
                ced["cedola_ruolo"] = "lorda"
                ced["cedola_proposta_split"] = True
                ced["cedola_coppia_idx"] = bank_txns.index(rit)
                rit["cedola_ruolo"] = "ritenuta"
                rit["cedola_coppia_lorda_idx"] = bank_txns.index(ced)
                used.add(id(rit))
                break

    # ── Dividendi azionari ──────────────────────────────────────────────
    dividendi    = [t for t in bank_txns
                    if pat_dividendo.search(t.get("raw_text","")) and t["amount"] > 0]
    ritenute_div = [t for t in bank_txns
                    if pat_rit_div.search(t.get("raw_text","")) and t["amount"] < 0]

    for div in dividendi:
        if div.get("cedola_ruolo"):  # già processato
            continue
        titolo_div = titolo_da_desc2(div.get("descrizione_completa",""))
        for rit in ritenute_div:
            if id(rit) in used:
                continue
            if abs((div["date"] - rit["date"]).days) > finestra_giorni:
                continue
            # Abbina per titolo identico nella descrizione completa
            titolo_rit = titolo_da_desc2(rit.get("descrizione_completa",""))
            if titolo_div and titolo_div == titolo_rit:
                div["cedola_ruolo"] = "lorda"
                div["cedola_proposta_split"] = True
                div["cedola_coppia_idx"] = bank_txns.index(rit)
                rit["cedola_ruolo"] = "ritenuta"
                rit["cedola_coppia_lorda_idx"] = bank_txns.index(div)
                used.add(id(rit))
                break
            # Fallback: proporzione 26%
            effettiva = abs(rit["amount"])
            lorda     = abs(div["amount"])
            if lorda > 0 and abs(effettiva - lorda * 0.26) / (lorda * 0.26) < 0.10:
                if abs((div["date"] - rit["date"]).days) <= finestra_giorni:
                    div["cedola_ruolo"] = "lorda"
                    div["cedola_proposta_split"] = True
                    div["cedola_coppia_idx"] = bank_txns.index(rit)
                    rit["cedola_ruolo"] = "ritenuta"
                    rit["cedola_coppia_lorda_idx"] = bank_txns.index(div)

    return bank_txns


def fondi_cedole_per_match(bank_txns: list[dict]) -> list[dict]:
    """
    Per il MATCHING: dopo raggruppa_cedole_ritenute(), collassa ogni coppia
    (lorda, ritenuta) marcata in un'unica transazione banca col valore NETTO
    (lordo - ritenuta), così da abbinarla alla riga unica già presente in
    Moneyspire (che registra dividendi/cedole al netto).

    Caso tipico conto USD Fineco: la banca espone due righe
        Dividendo estero  +88,23
        Ritenuta          -22,94
    mentre Money ha una sola riga +65,29.

    La riga netta conserva i riferimenti alle due righe originali nei campi
    `_lorda_orig` / `_ritenuta_orig`, in modo che — se NON trova corrispondenza
    in Money — il chiamante possa comunque ricostruire lo split a due righe.
    Le righe non appaiate restano invariate.
    """
    by_id = {id(t): t for t in bank_txns}
    consumati: set = set()
    out: list[dict] = []

    for t in bank_txns:
        if id(t) in consumati:
            continue

        # Riga lorda con coppia ritenuta marcata
        if t.get("cedola_ruolo") == "lorda" and t.get("cedola_proposta_split"):
            idx_rit = t.get("cedola_coppia_idx")
            rit = bank_txns[idx_rit] if (idx_rit is not None
                                         and 0 <= idx_rit < len(bank_txns)) else None
            if rit is not None and rit.get("cedola_ruolo") == "ritenuta":
                lordo    = abs(t.get("deposit") or t["amount"])
                ritenuta = abs(rit.get("withdrawal") or rit["amount"])
                netto    = round(lordo - ritenuta, 2)
                merged = dict(t)
                merged["amount"]     = netto
                merged["deposit"]    = netto
                merged["withdrawal"] = 0.0
                merged["_cedola_netta"]   = True
                merged["_lorda_orig"]     = t
                merged["_ritenuta_orig"]  = rit
                merged["_lordo_val"]      = lordo
                merged["_ritenuta_val"]   = ritenuta
                out.append(merged)
                consumati.add(id(t))
                consumati.add(id(rit))
                continue

        out.append(t)

    return out


def costruisci_transazioni_da_risultati(
        results: list[dict],
        account_id: int,
        rules: "RulesEngine",
        cat_map: dict[str, int],
        cfg: dict | None = None) -> list[dict]:
    """
    A partire dai risultati della riconciliazione, costruisce la lista
    di transazioni da inserire nel DB per le righe MATCH_NONE.

    Gestisce automaticamente:
    - Transazioni semplici (categoria da regole)
    - Coppie cedola+ritenuta (split automatico)
    - Trasferimenti verso altri conti (CategoryID = ID conto)

    Ritorna lista di dict pronti per MoneyWriter.inserisci_transazione()
    """
    cfg = cfg or {}
    contropartita_default = cfg.get("contropartita_default", "Da classificare")

    # Risoluzione ID categorie speciali (cedole / dividendi / ritenuta).
    # Priorità: (1) cat_map del DB corrente, (2) variabili di modulo ms_db
    # (aggiornate da MoneyWriter._resolve_special_cats), (3) fallback hardcoded.
    def _resolve_cat(nome_canonico: str, fallback: int) -> int:
        cid = cat_map.get(nome_canonico.lower()) if cat_map else None
        return cid if cid else fallback

    cat_id_cedole    = _resolve_cat(ms_db._CAT_CEDOLE_NAME,    ms_db._CAT_CEDOLE)
    cat_id_dividendi = _resolve_cat(ms_db._CAT_DIVIDENDI_NAME, ms_db._CAT_DIVIDENDI)
    cat_id_ritenuta  = _resolve_cat(ms_db._CAT_RITENUTA_NAME,  ms_db._CAT_RITENUTA)

    da_inserire = []
    processati = set()  # indici già gestiti come parte di una coppia

    for i, r in enumerate(results):
        if r["match_type"] != MATCH_NONE:
            continue
        if i in processati:
            continue

        bt = r["bank_txn"]
        deposit    = bt["deposit"]
        withdrawal = bt["withdrawal"]
        desc       = bt.get("descrizione", "")
        desc2      = bt.get("descrizione_completa", "")
        raw        = bt.get("raw_text", "")
        txn_date   = bt["date"]

        # ── Caso 0: riga cedola/dividendo già fusa al NETTO per il match,
        #            rimasta senza corrispondenza in Money → ricrea lo split
        #            a due righe (lordo + ritenuta) usando le righe originali.
        if bt.get("_cedola_netta"):
            lordo    = bt.get("_lordo_val", deposit)
            ritenuta = bt.get("_ritenuta_val", 0.0)
            netto    = round(lordo - ritenuta, 2)
            is_dividendo = any(k in raw.lower()
                               for k in ["dividendo", "div.su", "proventi"])
            cat_id_lordo = cat_id_dividendi if is_dividendo else cat_id_cedole
            memo_txn = _estrai_memo_cedola(desc2)
            da_inserire.append({
                "account_id":  account_id,
                "txn_date":    txn_date,
                "deposit":     netto,
                "withdrawal":  0.0,
                "memo":        memo_txn,
                "category_id": None,
                "_descrizione_banca":    desc,
                "_descrizione_completa": bt.get("descrizione_completa", desc2),
                "_categoria_suggerita":  "split_cedola",
                "splits": [
                    {"deposit":    lordo,    "withdrawal": 0.0,
                     "category_id": cat_id_lordo, "memo": ""},
                    {"deposit":    0.0,      "withdrawal": ritenuta,
                     "category_id": cat_id_ritenuta, "memo": "ritenuta"},
                ],
                "_tipo": "split_cedola"
            })
            continue

        # ── Caso 1: coppia cedola+ritenuta ──────────────────────────────
        if bt.get("cedola_proposta_split") and "cedola_coppia_idx" in bt:
            coppia_idx = bt["cedola_coppia_idx"]
            # Trova la riga mancante corrispondente alla ritenuta
            rit_result = next(
                (results[j] for j in range(len(results))
                 if results[j]["match_type"] == MATCH_NONE
                 and results[j]["bank_txn"] is not None
                 and j == coppia_idx),
                None)
            if rit_result:
                rit_bt = rit_result["bank_txn"]
                lordo    = deposit
                ritenuta = rit_bt["withdrawal"]
                netto    = lordo - ritenuta

                # Determina categoria cedole/dividendi dal testo
                is_dividendo = any(k in raw.lower()
                                   for k in ["dividendo", "div.su", "proventi"])
                cat_id_lordo = cat_id_dividendi if is_dividendo else cat_id_cedole

                # Memo dalla descrizione completa (titolo)
                memo_txn = _estrai_memo_cedola(desc2)

                da_inserire.append({
                    "account_id":  account_id,
                    "txn_date":    txn_date,
                    "deposit":     netto,
                    "withdrawal":  0.0,
                    "memo":        memo_txn,
                    "category_id": None,
                    "_descrizione_banca":    desc,
                    "_descrizione_completa": bt.get("descrizione_completa", desc2),
                    "_categoria_suggerita":  "split_cedola",
                    "splits": [
                        {"deposit":    lordo,    "withdrawal": 0.0,
                         "category_id": cat_id_lordo, "memo": ""},
                        {"deposit":    0.0,      "withdrawal": ritenuta,
                         "category_id": cat_id_ritenuta, "memo": "ritenuta"},
                    ],
                    "_tipo": "split_cedola"
                })
                # Segna la ritenuta come già processata
                processati.add(coppia_idx)
                continue

        # Se è una ritenuta di una coppia: la salta solo se la cedola lorda
        # è anch'essa MATCH_NONE (verrà gestita come split insieme alla lorda).
        # Se la lorda è già in Money (MATCH_MERGE/MATCH_EXACT), la ritenuta
        # va inserita come transazione singola separata.
        if bt.get("cedola_ruolo") == "ritenuta":
            idx_lorda = bt.get("cedola_coppia_lorda_idx")
            if idx_lorda is not None:
                lorda_match = results[idx_lorda]["match_type"]
                if lorda_match == MATCH_NONE:
                    continue  # gestita insieme alla lorda come split
                # else: lorda già in Money → inserisci ritenuta da sola
            else:
                continue  # no coppia nota → salta

        # ── Caso 2: transazione semplice ────────────────────────────────
        sg = rules.apply(raw) or {}
        cat_name = sg.get("category", "") or contropartita_default

        # Categoria di sicurezza.
        # Una transazione senza regola NON deve essere inserita zoppa in Moneyspire:
        # viene sempre assegnata alla categoria reale configurata come contropartita_default.
        cat_id = cat_map.get(cat_name.lower())
        if cat_id is None and cat_name in ("", "—", contropartita_default):
            cat_id = cat_map.get(contropartita_default.lower())
            cat_name = contropartita_default

        # Gestione trasferimenti verso carta di credito o altro conto
        cat_id_finale = cat_id
        if cat_name == "Trasferimento:carta":
            # Cerca numero carta nella descrizione completa
            for cifre, acc_name in [("6421", "6421 Visa Fineco"),
                                      ("6553", "6421 Visa Fineco"),
                                      ("5260", "5260 MC Fineco")]:
                if cifre in desc2:
                    cat_id_finale = None
                    sg["_transfer_to"] = acc_name
                    break
        elif cat_name == "Trasferimento:conto":
            # Giroconto: CategoryID = ID conto Lombard (o determinato dalla desc)
            transfer_to = sg.get("_transfer_to", "Fineco Lombard")
            cat_id_finale = None
            sg["_transfer_to"] = transfer_to

        da_inserire.append({
            "account_id":  account_id,
            "txn_date":    txn_date,
            "deposit":     deposit,
            "withdrawal":  withdrawal,
            "memo":        "",
            "category_id": cat_id_finale,
            "splits":      None,
            "_descrizione_banca": desc,
            "_descrizione_completa": desc2,
            "_categoria_suggerita": cat_name,
            "_transfer_to": sg.get("_transfer_to"),
            "_tipo": "semplice"
        })

    return da_inserire


def _estrai_memo_cedola(desc_completa: str) -> str:
    """
    Estrae il memo dalla descrizione completa Fineco preservando il prefisso originale.
    Fineco usa già "Ced.su", "Div.su", "Rit.ced.su", "Rit.div.su" — usiamo quello.
    Es: "Div.su 45,000 APPLE"        → "Div.su 45,000 APPLE"
    Es: "Ced.su 200.000,000 BTP..."  → "Ced.su 200.000,000 BTP..."
    Es: "Rit.ced.su 400.000,000 BTP" → "Ced.su 400.000,000 BTP" (rimuove Rit.)
    """
    import re
    d = desc_completa.strip()

    # Cerca i pattern Fineco nativi: prefisso + quantità + titolo
    # Cattura tutto fino a eventuale "Cliente:" o fine stringa
    m = re.match(
        r'^(Div\.su|Ced\.su|Rit\.ced\.su|Rit\.div\.su)\s+([\d\.,]+)\s+(\S+.*?)(?:\s+Cliente:.*)?$',
        d, re.IGNORECASE)
    if m:
        prefisso = m.group(1)
        quantita = m.group(2)
        titolo   = m.group(3).strip()
        # Per le ritenute (Rit.xxx.su) usa il prefisso senza Rit.
        if prefisso.lower().startswith("rit."):
            prefisso = prefisso[4:]   # "Rit.ced.su" → "ced.su", "Rit.div.su" → "div.su"
            prefisso = prefisso[0].upper() + prefisso[1:]  # capitalizza
        return f"{prefisso} {quantita} {titolo}"

    # Fallback: restituisce la desc completa troncata (già leggibile da Fineco)
    return d[:50]




# ─────────────────────────────────────────────────────────────────────────────
# RAGGRUPPAMENTO TRANSAZIONI CORRELATE (N banca → 1 Money)
# ─────────────────────────────────────────────────────────────────────────────

def raggruppa_transazioni_correlate(bank_txns: list[dict],
                                     finestra_giorni: int = 1) -> list[dict]:
    """
    Identifica gruppi di transazioni banca che in Money vengono registrate
    come un'unica transazione (netta o split):

    1. Canone Mensile + Sconto Canone → si annullano, importo netto = 0
       → entrambe marcate come "gruppo_annullato"

    2. Interessi Portaf.Remunerato + Ritenuta Portaf.Remunerato
       → split attivo/passivo, netto = interessi - ritenuta
       → marcate come "gruppo_split_interessi"

    3. CBILL principale + Commissione CBILL + Abbuono CBILL
       → la commissione e l'abbuono si annullano, netto = importo principale
       → commissione e abbuono marcati come "gruppo_accessorio" della principale

    4. Proventi fondo: N righe con stessa descrizione e data → somma = Money
       → marcate come "gruppo_merge" (da abbinare a un'unica Money)
    """
    import re

    # Pattern per ogni gruppo
    pat_canone   = re.compile(r'canone mensile conto', re.IGNORECASE)
    pat_sconto   = re.compile(r'sconto canone mensile', re.IGNORECASE)
    # Bollo dossier e bollo C/C: addebito + storno si azzerano
    pat_bollo    = re.compile(r'imposta bollo|bollo dossier|bollo.*c/c|storno imposta', re.IGNORECASE)
    pat_int_port = re.compile(r'interessi portaf', re.IGNORECASE)
    pat_rit_port = re.compile(r'ritenuta portaf', re.IGNORECASE)
    pat_cbill    = re.compile(r'pagamento bollettino cbill', re.IGNORECASE)
    pat_comm_cb  = re.compile(r'commiss\.bollettino cbill', re.IGNORECASE)
    pat_abbuono  = re.compile(r'abbuono comm\. cbill', re.IGNORECASE)
    pat_proventi = re.compile(r'proventi fondo|dividendo.*pimco|dividendo.*sic', re.IGNORECASE)

    used = set()

    for i, t in enumerate(bank_txns):
        if i in used:
            continue
        raw = t.get("raw_text", "")

        # ── 1. Canone + Sconto (si annullano) ────────────────────────────
        if pat_canone.search(raw) and t["amount"] < 0:
            for j, t2 in enumerate(bank_txns):
                if j == i or j in used: continue
                if abs((t["date"] - t2["date"]).days) > finestra_giorni: continue
                if pat_sconto.search(t2.get("raw_text","")) and t2["amount"] > 0:
                    if abs(abs(t["amount"]) - abs(t2["amount"])) < 0.02:
                        t["gruppo"] = "annullato"
                        t2["gruppo"] = "annullato"
                        t["gruppo_note"] = "Canone annullato dallo sconto — entrambi ignorabili"
                        t2["gruppo_note"] = "Sconto annulla il canone — entrambi ignorabili"
                        used.update([i, j])
                        break

        # ── 1b. Bollo dossier / Bollo C/C: addebito + storno si azzerano ──
        elif pat_bollo.search(raw):
            for j, t2 in enumerate(bank_txns):
                if j == i or j in used: continue
                if abs((t["date"] - t2["date"]).days) > 5: continue  # finestra più larga
                if not pat_bollo.search(t2.get("raw_text","")): continue
                # Stessa descrizione, importi opposti
                if abs(t["amount"] + t2["amount"]) < 0.02:
                    t["gruppo"]       = "annullato"
                    t2["gruppo"]      = "annullato"
                    t["gruppo_note"]  = "Bollo annullato dallo storno — entrambi ignorabili"
                    t2["gruppo_note"] = "Storno annulla il bollo — entrambi ignorabili"
                    used.update([i, j])
                    break

        # ── 2. Interessi + Ritenuta Portafoglio Remunerato ───────────────        # ── 2. Interessi + Ritenuta Portafoglio Remunerato ───────────────
        elif pat_int_port.search(raw) and t["amount"] > 0:
            for j, t2 in enumerate(bank_txns):
                if j == i or j in used: continue
                if abs((t["date"] - t2["date"]).days) > finestra_giorni: continue
                if pat_rit_port.search(t2.get("raw_text","")) and t2["amount"] < 0:
                    netto = t["amount"] + t2["amount"]
                    t["gruppo"]       = "split_interessi"
                    t2["gruppo"]      = "split_interessi"
                    t["gruppo_pari"]  = j
                    t2["gruppo_pari"] = i
                    t["gruppo_note"]  = f"Interessi lordi — netto con ritenuta: {fmt_eur(netto)}"
                    t2["gruppo_note"] = "Ritenuta sugli interessi — abbinata alla riga precedente"
                    used.update([i, j])
                    break

        # ── 3. CBILL principale + Commissione + Abbuono ──────────────────
        elif pat_cbill.search(raw):
            # Cerca commissione e abbuono nella stessa data ±1gg
            for j, t2 in enumerate(bank_txns):
                if j == i or j in used: continue
                if abs((t["date"] - t2["date"]).days) > finestra_giorni: continue
                raw2 = t2.get("raw_text","")
                if pat_comm_cb.search(raw2) or pat_abbuono.search(raw2):
                    t2["gruppo"]      = "accessorio"
                    t2["gruppo_pari"] = i
                    t2["gruppo_note"] = f"Commissione/abbuono CBILL — accessorio a {fmt_eur(t['amount'])}"
                    if "gruppo_note" not in t:
                        t["gruppo_note"] = f"Principale CBILL con commissioni accessorie"
                    t["gruppo"] = t.get("gruppo", "principale_cbill")
                    used.add(j)

        # ── 4. Proventi fondo: più righe stessa data → merge ─────────────
        elif pat_proventi.search(raw) and i not in used:
            gruppo_proventi = [i]
            for j, t2 in enumerate(bank_txns):
                if j == i or j in used: continue
                if abs((t["date"] - t2["date"]).days) > finestra_giorni: continue
                if pat_proventi.search(t2.get("raw_text","")):
                    gruppo_proventi.append(j)
            if len(gruppo_proventi) > 1:
                totale = sum(bank_txns[k]["amount"] for k in gruppo_proventi)
                for k in gruppo_proventi:
                    bank_txns[k]["gruppo"]      = "merge_proventi"
                    bank_txns[k]["gruppo_note"] = (f"Provento fondo — somma gruppo: "
                                                    f"{fmt_eur(totale)}")
                    bank_txns[k]["gruppo_totale"] = totale
                used.update(gruppo_proventi)

    return bank_txns


# ─────────────────────────────────────────────────────────────────────────────
# FASE 2: INTEGRAZIONE FILE EXCEL ELABORATI


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_full_csv(results: list[dict], out_path: str,
                    rules: "RulesEngine | None" = None) -> int:
    """
    Esporta risultati riconciliazione in CSV con formato ottimizzato per revisione:
    - Date e importi affiancati (banca | Money) per confronto immediato
    - Ordinamento per tipologia: trovate → fuzzy → split/merge → mancanti → solo Money
    - Flag IMPORTO_DIVERSO quando importo banca ≠ importo Money (oltre tolleranza)
    - Flag DATA_DIVERSA con delta giorni
    """

    # Ordine tipologie
    ORDINE = {
        MATCH_EXACT:    0,
        MATCH_FUZZY:    1,
        MATCH_SPLIT:    2,
        MATCH_MERGE:    3,
        MATCH_NONE:     4,
        MATCH_SKIP:     5,
        MATCH_PENDING:  6,   # ultime: le vedremo il mese dopo
    }

    # Ordina: prima per tipologia, poi per data
    def sort_key(r):
        bt = r.get("bank_txn")
        mt = r["money_txns"][0] if r["money_txns"] else None
        dt = (bt["date"] if bt else mt["date"] if mt else date(2000,1,1))
        return (ORDINE.get(r["match_type"], 9), dt)

    sorted_results = sorted(results, key=sort_key)

    # ── Identifica coppie Mancante ↔ Solo Money con importo simile ──────────
    # Per ogni Mancante, cerca un Solo Money con stessa data ±5gg e importo ±5%
    mancanti  = [(i, r) for i, r in enumerate(sorted_results)
                 if r["match_type"] == MATCH_NONE and r.get("bank_txn")]
    solo_money = [(i, r) for i, r in enumerate(sorted_results)
                  if r["match_type"] == MATCH_SKIP and r["money_txns"]]
    coppie: dict[int, int] = {}  # indice mancante → indice solo money
    used_sm = set()
    for i_m, rm in mancanti:
        bt_amt  = rm["bank_txn"]["amount"]
        bt_date = rm["bank_txn"]["date"]
        for i_s, rs in solo_money:
            if i_s in used_sm:
                continue
            mt      = rs["money_txns"][0]
            sm_amt  = mt["amount"]
            sm_date = mt["date"]
            delta_gg  = abs((bt_date - sm_date).days)
            delta_imp = abs(abs(bt_amt) - abs(sm_amt))
            ref       = abs(bt_amt) if abs(bt_amt) > 0.01 else 1
            if delta_gg <= 5 and delta_imp / ref <= 0.05:
                coppie[i_m] = i_s
                used_sm.add(i_s)
                break

    rows = []
    for idx, r in enumerate(sorted_results):
        bt = r.get("bank_txn")
        mt = r["money_txns"][0] if r["money_txns"] else None
        # Se annullata (coppia che si azzera), usa etichetta speciale
        if r.get("annullata"):
            label = "🔄 Annullata"
        else:
            label, _ = STATO_LABELS.get(r["match_type"], ("?", ""))

        # Importi
        imp_banca = bt["amount"] if bt else None
        imp_money = mt["amount"] if mt else None

        # Flag discrepanza importo (oltre 0.02 €)
        # Non segnalare per MERGE/SPLIT: il delta è atteso (è la ritenuta nello split)
        flag_importo = ""
        if (imp_banca is not None and imp_money is not None
                and r["match_type"] not in (MATCH_SPLIT, MATCH_MERGE)):
            diff_imp = abs(abs(imp_banca) - abs(imp_money))
            if diff_imp > 0.02:
                flag_importo = f"Δ {diff_imp:.2f} €"

        # Flag discrepanza data
        flag_data = ""
        if bt and mt:
            delta_gg = abs((bt["date"] - mt["date"]).days)
            if delta_gg > 0:
                flag_data = f"+{delta_gg}gg" if mt["date"] > bt["date"] else f"-{delta_gg}gg"

        # Categoria: da Money oppure suggerita da regole
        cat_display   = (mt["category"] if mt else "") or r.get("suggested_category", "")
        payee_display = (mt["payee"]    if mt else "") or r.get("suggested_payee", "")

        def fmt_imp(v):
            if v is None: return ""
            return f"{v:.2f}".replace(".", ",")

        # Flag coppia Mancante ↔ Solo Money
        flag_coppia = ""
        if r["match_type"] == MATCH_NONE and idx in coppie:
            i_s = coppie[idx]
            sm = sorted_results[i_s]["money_txns"][0]
            d_imp = abs(abs(imp_banca or 0) - abs(sm["amount"]))
            flag_coppia = f"⚠️ in Money={sm['amount']:.2f}€ Δ{d_imp:.2f}€" if d_imp > 0.02 else "⚠️ possibile coppia"
        elif r["match_type"] == MATCH_SKIP:
            for i_m, i_s in coppie.items():
                if i_s == idx:
                    bm = sorted_results[i_m]["bank_txn"]
                    d_imp = abs(abs(bm["amount"]) - abs(imp_money or 0))
                    flag_coppia = f"⚠️ in banca={bm['amount']:.2f}€ Δ{d_imp:.2f}€" if d_imp > 0.02 else "⚠️ possibile coppia"
                    break

        rows.append({
            "Tipologia":     label,
            "Data_banca":    bt["date"].strftime("%d/%m/%Y") if bt else "",
            "Data_Money":    mt["date"].strftime("%d/%m/%Y") if mt else "",
            "Delta_data":    flag_data,
            "Importo_banca": fmt_imp(imp_banca),
            "Importo_Money": fmt_imp(imp_money),
            "Delta_importo": flag_importo,
            "Descrizione":   bt.get("descrizione", "") if bt else "",
            "Descr_completa":bt.get("descrizione_completa", "") if bt else "",
            "Categoria":     cat_display,
            "Payee":         payee_display,
            "Memo_Money":    mt["memo"] if mt else "",
            "Splits":        str(len(mt["splits"])) if mt and mt["has_splits"] else "",
            "Note":          r.get("note", ""),
            "Coppia":        flag_coppia,
            "ID_Money":      ",".join(str(i) for i in r["money_ids"])
        })

    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        if rows:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=";")
            w.writeheader()
            w.writerows(rows)
    return len(rows)

def export_missing_csv(results: list[dict], out_path: str,
                       rules: "RulesEngine | None" = None) -> int:
    rows = []
    for r in results:
        if r["match_type"] != MATCH_NONE:
            continue
        bt  = r["bank_txn"]
        cat = payee = ""
        if rules:
            sg = rules.apply(bt.get("raw_text", ""))
            if sg:
                cat, payee = sg.get("category", ""), sg.get("payee", "")
        rows.append({
            "Data":                 bt["date"].strftime("%d/%m/%Y"),
            "Importo":              f"{bt['amount']:.2f}".replace(".", ","),
            "Descrizione":          bt.get("descrizione", ""),
            "Descrizione_Completa": bt.get("descrizione_completa", ""),
            "Categoria_suggerita":  cat or "Da classificare",
            "Payee_suggerito":      payee,
            "Fonte":                bt.get("source", "")
        })
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        if rows:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=";")
            w.writeheader()
            w.writerows(rows)
    return len(rows)



def valida_file_banca(xlsx_path: str, conto_cfg: dict) -> tuple[bool, str]:
    """
    Verifica che il file Excel corrisponda al conto selezionato.
    Ritorna (ok, messaggio).
    """
    val = conto_cfg.get("validazione", {})
    tipo = val.get("tipo")

    if tipo == "foglio_excel":
        foglio_atteso = val.get("foglio_atteso", "")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)
            fogli = wb.sheetnames
            wb.close()
            if foglio_atteso in fogli:
                return True, f"Foglio '{foglio_atteso}' trovato ✓"
            else:
                return False, (f"Il file non contiene il foglio '{foglio_atteso}'.\n"
                               f"Fogli trovati: {', '.join(fogli)}\n"
                               f"Hai selezionato il file giusto per il conto '{conto_cfg.get('excel_sheet', '')}'?")
        except Exception as e:
            return False, f"Errore lettura file: {e}"

    elif tipo == "numero_carta":
        cifre_raw = val.get("cifre_attese", "")
        # Accetta sia stringa singola che lista
        cifre_list = cifre_raw if isinstance(cifre_raw, list) else [cifre_raw]
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            carte_trovate = set()
            header_found = False
            for row in ws.iter_rows(max_row=200, values_only=True):
                if not header_found:
                    if any(c == "Numero carta" for c in row if c):
                        header_found = True
                    continue
                num = str(row[1] or "").strip() if len(row) > 1 else ""
                if num:
                    carte_trovate.add(num)
            wb.close()
            trovate_ok = [c for c in cifre_list
                          if any(c in carta for carta in carte_trovate)]
            if trovate_ok:
                desc = ", ".join(f"...{c}" for c in trovate_ok)
                return True, f"Carte trovate: {desc} ✓"
            else:
                desc_file = ", ".join(carte_trovate) if carte_trovate else "nessuna carta"
                attese = ", ".join(f"...{c}" for c in cifre_list)
                return False, (f"Nessuna carta attesa ({attese}) trovata nel file.\n"
                               f"Carte nel file: {desc_file}\n"
                               f"Hai selezionato il file giusto?")
        except Exception as e:
            return False, f"Errore lettura file: {e}"

    elif tipo == "intestazione_conto":
        # Validazione per file Unicredit: controlla le prime righe del file
        testo_atteso = val.get("testo_atteso", "Unicredit")
        try:
            import pandas as pd
            # Legge le prime 5 righe senza header per trovare testo identificativo
            df = pd.read_excel(xlsx_path, header=None, nrows=5, engine="xlrd")
            # Cerca "Rapporto" o il testo atteso in qualsiasi cella delle prime righe
            testo_trovato = ""
            for _, row in df.iterrows():
                for val_cell in row:
                    s = str(val_cell or "").strip()
                    if s and s.lower() != "nan":
                        testo_trovato = s
                        break
                if testo_trovato:
                    break
            parole_unicredit = ["rapporto", "unicredit", "saldo contabile", "saldo disponibile"]
            if any(p in testo_trovato.lower() for p in parole_unicredit):
                return True, f"File Unicredit riconosciuto ✓"
            else:
                return False, (f"Il file non sembra un estratto Unicredit.\n"
                               f"Prima riga trovata: '{testo_trovato[:80]}'\n"
                               f"Atteso testo contenente: '{testo_atteso}'")
        except Exception as e:
            return False, f"Errore lettura file: {e}"

    elif tipo == "intestazione_bper":
        # Validazione file BPER .xls: usa xlrd, non openpyxl
        try:
            import xlrd
            wb = xlrd.open_workbook(xlsx_path)
            ws = wb.sheet_by_index(0)
            # Controlla le prime 20 righe per parole chiave BPER/Popolare di Sondrio
            testo = " ".join(
                str(ws.cell_value(r, c))
                for r in range(min(ws.nrows, 20))
                for c in range(min(ws.ncols, 4))
            ).lower()
            parole = ["saldo disponibile", "saldo contabile", "intestatari",
                      "banca popolare", "bper", "iban"]
            if any(p in testo for p in parole):
                return True, "File BPER riconosciuto ✓"
            else:
                return False, ("Il file non sembra un export BPER.\n"
                               "Verifica di aver esportato dal portale BPER.")
        except ImportError:
            return False, "Libreria 'xlrd' non installata. Esegui: pip install xlrd"
        except Exception as e:
            return False, f"Errore lettura file BPER: {e}"

    elif tipo == "intestazione_nexi":
        # Validazione file Nexi .xlsx
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            testo = " ".join(
                str(v) for row in ws.iter_rows(max_row=10, values_only=True)
                for v in row if v is not None
            ).lower()
            wb.close()
            parole = ["movimenti carta", "carta di credito", "nexi", "importo (€)"]
            if any(p in testo for p in parole):
                return True, "File Nexi riconosciuto ✓"
            else:
                return False, ("Il file non sembra un export Nexi Pay.\n"
                               "Verifica di aver esportato dal portale Nexi Pay.")
        except Exception as e:
            return False, f"Errore lettura file Nexi: {e}"

    # Nessuna validazione configurata → passa sempre
    return True, "Nessuna validazione configurata"

