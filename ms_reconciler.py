"""
ms_reconciler.py — GUI principale Moneyspire Reconciler
Riconciliazione Moneyspire ↔ Fineco / BPER / Nexi  |  Versione 2.1.3 — Multi-profilo IRC / SC

Struttura:
  Costanti e percorsi           — APP_TITLE, APP_NAME, CFG_PATH, …
  Log di audit                  — write_audit_log(), _audit_log_path()
  Utility config                — load_config(), save_config()
  Dialoghi di supporto          — PeriodDialog, DetailDialog
  Tab 1 — Riconcilia            — TabRiconcilia
  Dialog revisione inserimenti  — DialogRevisione
  Dialog aggiornamento fuzzy    — DialogAggiornaFuzzy
  Dialog modifica categoria     — DialogModificaCategoria
  Tab 2 — Regole                — TabRegole, RuleDialog
  Tab 3 — Log                   — TabLog
  Tab 4 — Fase 2 Excel          — TabFase2
  Applicazione principale       — App
"""

# --- IRC shared bootstrap ---
# Rende disponibili i moduli in Python/shared/ senza dipendere da PYTHONPATH.
# Saltato se eseguito da bundle PyInstaller (sys.frozen=True): in quel caso
# i moduli sono gia' inclusi nel bundle.
import sys as _sys
from pathlib import Path as _Path
if not getattr(_sys, 'frozen', False):
    _shared = _Path.home() / "Library/CloudStorage/Dropbox/Documenti_IRC/Python/shared"
    if str(_shared) not in _sys.path:
        _sys.path.insert(0, str(_shared))
# --- end IRC shared bootstrap ---


VERSION = "2.1.4"

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import sys
import calendar
from pathlib import Path
from datetime import date, datetime
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# Aggiunge la cartella shared di progetto al path per trovare path_widgets.py
_SHARED = Path.home() / "Library/CloudStorage/Dropbox/Documenti_IRC/Python/shared"
if _SHARED.exists() and str(_SHARED) not in sys.path:
    sys.path.insert(0, str(_SHARED))

from ms_engine import (
    MoneyspireDB, ReconcileEngine, RulesEngine,
    parse_fineco_conto_originale, parse_fineco_conto, parse_fineco_cc,
    parse_fineco_cc_cartella,
    parse_unicredit_ccm, leggi_intestazione_unicredit,
    parse_bper, leggi_intestazione_bper,
    parse_nexi_xlsx, leggi_intestazione_nexi,
    raggruppa_cedole_ritenute, fondi_cedole_per_match,
    valida_file_banca, leggi_numero_conto_fineco,
    export_missing_csv, export_full_csv, fmt_eur,
    costruisci_transazioni_da_risultati, MoneyWriter,
    prepara_db_scrittura, verifica_scrittura, finalizza_db,
    marca_in_attesa, marca_solo_money_in_attesa,
    raggruppa_transazioni_correlate,
    leggi_saldo_fineco, leggi_saldo_money,
    leggi_variazione_mensile_money, leggi_variazione_mensile_fineco,
    MATCH_EXACT, MATCH_FUZZY, MATCH_SPLIT, MATCH_MERGE, MATCH_NONE, MATCH_SKIP, MATCH_PENDING,
    STATO_LABELS, DEFAULT_CONFIG,
    # Fase 2
    ExcelIntegrator, MESI_IT_NOMI,
    leggi_excel_elaborato_cc, leggi_excel_elaborato_cc_mensile,
)
try:
    from path_widgets import PathVar, PathEntry, PathLabel
    _HAS_PATH_WIDGETS = True
except ImportError:
    _HAS_PATH_WIDGETS = False
    # Fallback: PathVar = StringVar, PathEntry = Entry readonly normale
    class PathVar(tk.StringVar):  # type: ignore
        pass
    class PathEntry(ttk.Frame):  # type: ignore
        def __init__(self, parent, pathvar, **kw):
            super().__init__(parent)
            ttk.Entry(self, textvariable=pathvar, state="readonly"
                      ).pack(fill="x", expand=True)

# ─────────────────────────────────────────────────────────────────────────────
APP_NAME   = "Riconciliazione Moneyspire"   # usato da AppBuilder per la cartella _Config
SCRIPT_DIR = Path(__file__).parent
# Cartella config canonica: Dropbox/Documenti_IRC/Python/_Config/Riconciliazione Moneyspire/
_DROPBOX   = Path.home() / "Library/CloudStorage/Dropbox/Documenti_IRC"
_CFG_DIR   = _DROPBOX / "Python/_Config/Riconciliazione Moneyspire"

# ── Selezione profilo ─────────────────────────────────────────────────────────
# Determinata da ProfiloDialog all'avvio. Le variabili globali vengono
# impostate da _init_profilo() prima che App() venga istanziata.
_PROFILO   = "IRC"          # "IRC" o "SC" — impostato da _init_profilo()
APP_TITLE  = f"Moneyspire Reconciler  v{VERSION}"   # aggiornato dopo selezione profilo
CFG_PATH   = _CFG_DIR / "IRC_config.json"    # placeholder — aggiornato da _init_profilo()
RULES_PATH = _CFG_DIR / "IRC_rules.json"     # placeholder — aggiornato da _init_profilo()
# Log di audit: Documenti_IRC/Log/ms_reconciler_YYYY.log  (append-only, uno per anno)
# Usa la cartella Log condivisa tra tutte le app, non quella specifica del progetto.
_LOG_DIR   = Path.home() / "Documents/log"


# ══════════════════════════════════════════════════════════════════════
# LOG DI AUDIT
# Scrive su _Config/Logs/ms_reconciler_YYYY.log dopo ogni operazione di
# scrittura sul DB. Permette tracciabilità completa e roll-back manuale
# tramite gli ID transazione SQLite registrati per ogni riga.
# ══════════════════════════════════════════════════════════════════════

def _audit_log_path() -> Path:
    """Ritorna il percorso del file di log per l'anno corrente."""
    from datetime import date as _date
    _LOG_DIR.mkdir(parents=True, exist_ok=True)
    return _LOG_DIR / f"ms_reconciler_{_date.today().year}.log"


def write_audit_log(
        operazione: str,
        ffd_path: str,
        backup_path: str,
        conto_nome: str,
        account_id: int,
        righe: list[dict],
        riepilogo: str,
        periodo: str = "") -> None:
    """
    Scrive una sessione di audit nel file di log annuale.

    operazione : "INSERIMENTO" | "AGGIORNA_DATA" | "AGGIORNA_IMPORTO" | "AGGIORNA_FUZZY"
    ffd_path   : percorso del file .ffd originale
    backup_path: percorso del backup creato prima della scrittura
    conto_nome : es. "Fineco", "Visa Fineco"
    account_id : ID conto nel DB
    righe      : lista di dict — una per ogni contabile toccata (vedi sotto)
    riepilogo  : stringa riassuntiva (es. "12 inserite  0 errori  Verifica: OK")
    periodo    : es. "Aprile 2026" (solo per inserimenti)

    Struttura di ogni dict in `righe`:
        tipo        : "INS" | "UPD_DATA" | "UPD_IMP" | "SKIP" | "ERR"
        txn_id      : ID transazione nel DB (int, None se nuovo inserimento non confermato)
        data        : date o str
        importo     : float (deposit - withdrawal)
        categoria   : str
        payee       : str
        memo        : str
        note        : str  (es. "data: 2026-04-14 → 2026-04-15", "split — importo bloccato")
    """
    from datetime import datetime as _dt

    sep_heavy = "═" * 72
    sep_light = "─" * 72
    ts = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = [
        "",
        sep_heavy,
        f"SESSIONE  {ts}   {APP_TITLE}",
        f"File .ffd : {Path(ffd_path).name}",
        f"Backup    : {Path(backup_path).name}",
        f"Operazione: {operazione}   Conto: {conto_nome} (ID {account_id})"
        + (f"   Periodo: {periodo}" if periodo else ""),
        sep_light,
    ]

    for r in righe:
        tipo    = r.get("tipo", "?")
        txn_id  = r.get("txn_id")
        data    = r.get("data", "")
        importo = r.get("importo")
        cat     = r.get("categoria", "") or "—"
        payee   = r.get("payee", "") or "—"
        memo    = r.get("memo", "") or "—"
        note    = r.get("note", "")

        # Formatta importo
        if importo is not None:
            sign   = "+" if importo >= 0 else "-"
            imp_s  = f"{sign}{abs(importo):>11,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
        else:
            imp_s  = " " * 14 + "—"

        # Formatta data
        if hasattr(data, "strftime"):
            data_s = data.strftime("%Y-%m-%d")
        else:
            data_s = str(data)[:10]

        id_s  = f"ID:{txn_id:<7}" if txn_id else " " * 10
        note_s = f"  [{note}]" if note else ""

        line = (f"  {tipo:<8} {data_s}  {imp_s}  "
                f"Cat: {cat[:28]:<28}  Payee: {payee[:20]:<20}  "
                f"{id_s}  Memo: {str(memo)[:40]}{note_s}")
        lines.append(line)

    lines += [
        sep_light,
        f"Riepilogo: {riepilogo}",
        sep_heavy,
    ]

    log_path = _audit_log_path()
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
    except Exception as e:
        # Il log di audit non deve mai bloccare l'operazione principale
        print(f"[AVVISO] Impossibile scrivere audit log: {e}")

COL = {
    MATCH_EXACT:    "#d4edda",
    MATCH_FUZZY:    "#fff3cd",
    MATCH_SPLIT:    "#cce5ff",
    MATCH_MERGE:    "#e2ccff",
    MATCH_NONE:     "#f8d7da",
    MATCH_SKIP:     "#e2e3e5",
    MATCH_PENDING:  "#fde8c8",
}
COL_ANNULLATA = "#eeeeee"   # grigio chiaro per transazioni annullate
MESI_IT = ["Gennaio","Febbraio","Marzo","Aprile","Maggio","Giugno",
           "Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"]


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — carica/salva impostazioni inclusi ultimi percorsi usati
# ──────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# UTILITY ORDINAMENTO TREEVIEW
# ══════════════════════════════════════════════════════════════════════

def treeview_sort(tree: "ttk.Treeview", col: str,
                  sort_state: dict,
                  date_cols:   tuple = (),
                  number_cols: tuple = ()) -> None:
    """
    Ordinamento generico con toggle asc/desc per qualsiasi Treeview.

    sort_state  : dict condiviso {"col": str, "asc": bool} — modificato in place.
    date_cols   : nomi colonna da trattare come date gg/mm/aaaa.
    number_cols : nomi colonna da trattare come numeri (rimuove €, virgole, punti).

    Aggiorna l'header con ▲/▼ e riposiziona le righe.
    """
    if sort_state.get("col") == col:
        sort_state["asc"] = not sort_state.get("asc", True)
    else:
        sort_state["col"] = col
        sort_state["asc"] = True

    asc = sort_state["asc"]
    items = []
    for k in tree.get_children(""):
        val = tree.set(k, col)
        if col in date_cols and val and len(val) == 10 and val[2] == "/":
            try:
                d, m, y = val.split("/")
                sort_key = f"{y}-{m}-{d}"
            except ValueError:
                sort_key = val
        elif col in number_cols and val:
            try:
                clean = val.replace("€", "").replace("$", "").replace(" ", "")
                clean = clean.replace(".", "").replace(",", ".")
                sort_key = float(clean)
            except ValueError:
                sort_key = 0.0
        else:
            sort_key = val
        items.append((sort_key, k))

    try:
        items.sort(reverse=not asc)
    except TypeError:
        items.sort(key=lambda x: str(x[0]), reverse=not asc)

    for idx, (_, k) in enumerate(items):
        tree.move(k, "", idx)

    # Aggiorna frecce negli header
    for c in tree["columns"]:
        txt = tree.heading(c, "text").rstrip(" ▲▼")
        if c == col:
            tree.heading(c, text=txt + (" ▲" if asc else " ▼"))
        else:
            tree.heading(c, text=txt)


# ══════════════════════════════════════════════════════════════════════
# UTILITY CONFIGURAZIONE
# ══════════════════════════════════════════════════════════════════════

def _norm_path(p: str) -> str:
    """
    Normalizza un path per macOS+Dropbox:
    1. Rimuove double slash iniziali (//Users → /Users) prodotte da symlink Dropbox
    2. Risolve alias username macOS: il sistema ha sia /Users/ignaziorusconiclerici
       che /Users/ignazio (shortname). Se il path usa il nome lungo ma la home
       attuale è il nome corto (o viceversa), sostituisce il prefisso /Users/XXX
       con il prefisso corretto basato su Path.home().
    """
    if not p:
        return p
    # 1. Rimuovi double slash
    while p.startswith("//"):
        p = p[1:]
    # 2. Normalizza username se il path non esiste ma esiste con l'altro alias
    from pathlib import Path as _Path
    if not _Path(p).exists() and p.startswith("/Users/"):
        home = str(_Path.home())  # es. /Users/ignazio
        parts = p.split("/", 3)   # ['', 'Users', 'ignaziorusconiclerici', 'rest...']
        if len(parts) >= 4:
            alt = "/".join(["", "Users", home.split("/")[-1], parts[3]])
            if _Path(alt).exists():
                p = alt
    return p

# ─────────────────────────────────────────────────────────────────────────────
# SELEZIONE PROFILO
# ─────────────────────────────────────────────────────────────────────────────

class ProfiloDialog(tk.Toplevel):
    """Finestra modale di selezione profilo all'avvio (IRC = Ignazio, SC = Silvia)."""

    PROFILI = {
        "IRC": ("Ignazio", "IRC_config.json", "IRC_rules.json"),
        "SC":  ("Silvia",  "SC_config.json",  "SC_rules.json"),
    }

    def __init__(self, parent: tk.Tk):
        super().__init__(parent)
        self.title("Seleziona profilo")
        self.resizable(False, False)
        self.profilo: str | None = None

        # Centra sulla root
        self.update_idletasks()
        pw, ph = parent.winfo_screenwidth(), parent.winfo_screenheight()
        self.geometry(f"+{pw//2 - 160}+{ph//2 - 80}")

        ttk.Label(self, text="Scegli il profilo da caricare:",
                  font=("Arial", 13, "bold")).pack(pady=(22, 10), padx=36)

        frm = ttk.Frame(self)
        frm.pack(pady=(6, 18), padx=36)
        for col, (codice, (nome, _, _)) in enumerate(self.PROFILI.items()):
            ttk.Button(frm, text=f"  {nome}  ({codice})  ", width=18,
                       command=lambda c=codice: self._scegli(c)
                       ).grid(row=0, column=col, padx=12, pady=8)

        self.protocol("WM_DELETE_WINDOW", self._annulla)
        self.transient(parent)
        self.grab_set()
        self.wait_window()

    def _scegli(self, codice: str):
        self.profilo = codice
        self.destroy()

    def _annulla(self):
        self.profilo = None
        self.destroy()


def _init_profilo(root: tk.Tk) -> str:
    """
    Mostra ProfiloDialog, aggiorna le variabili globali CFG_PATH / RULES_PATH /
    APP_TITLE / _PROFILO e restituisce il codice profilo ("IRC" o "SC").
    Termina il processo se l'utente chiude la finestra senza scegliere.
    """
    global CFG_PATH, RULES_PATH, APP_TITLE, _PROFILO

    dlg = ProfiloDialog(root)
    if dlg.profilo is None:
        root.destroy()
        sys.exit(0)

    codice = dlg.profilo
    _, cfg_file, rules_file = ProfiloDialog.PROFILI[codice]

    # Percorsi nella cartella _Config canonica; fallback alla cartella script
    if _CFG_DIR.exists():
        CFG_PATH   = _CFG_DIR / cfg_file
        RULES_PATH = _CFG_DIR / rules_file
    else:
        CFG_PATH   = SCRIPT_DIR / cfg_file
        RULES_PATH = SCRIPT_DIR / rules_file

    _PROFILO  = codice
    nome      = ProfiloDialog.PROFILI[codice][0]
    APP_TITLE = f"Moneyspire Reconciler  v{VERSION}  —  {nome} ({codice})"

    return codice


def _init_profilo_da_argomento(codice: str):
    """
    Inizializza il profilo direttamente da un codice stringa ("IRC" o "SC"),
    senza mostrare alcuna dialog. Usata quando il profilo è passato via sys.argv
    da ms_launcher.py.
    """
    global CFG_PATH, RULES_PATH, APP_TITLE, _PROFILO

    if codice not in ProfiloDialog.PROFILI:
        codice = "IRC"   # fallback sicuro

    _, cfg_file, rules_file = ProfiloDialog.PROFILI[codice]

    if _CFG_DIR.exists():
        CFG_PATH   = _CFG_DIR / cfg_file
        RULES_PATH = _CFG_DIR / rules_file
    else:
        CFG_PATH   = SCRIPT_DIR / cfg_file
        RULES_PATH = SCRIPT_DIR / rules_file

    _PROFILO  = codice
    nome      = ProfiloDialog.PROFILI[codice][0]
    APP_TITLE = f"Moneyspire Reconciler  v{VERSION}  —  {nome} ({codice})"



def _path_to_cfg(p) -> str:
    """Salva il path relativo alla home, per portabilità tra Mac."""
    try:
        return str(Path(str(p)).expanduser().resolve().relative_to(Path.home()))
    except ValueError:
        return str(p)

def _path_from_cfg(s: str) -> str:
    """Ricostruisce il path assoluto: se relativo, prepende Path.home()."""
    if not s:
        return s
    p = Path(s)
    if p.is_absolute():
        # retrocompatibilità: path assoluto vecchio stile
        return str(p)
    return str(Path.home() / p)

def load_config() -> dict:
    if CFG_PATH.exists():
        cfg = json.loads(CFG_PATH.read_text(encoding="utf-8"))
        # Ricostruisce i path assoluti da path relativi salvati
        lp = cfg.get("last_paths", {})
        for k, v in lp.items():
            if isinstance(v, str):
                lp[k] = _norm_path(_path_from_cfg(v))
        return cfg
    # Nessun fallback silenzioso: un config mancante è un errore di configurazione,
    # non una condizione da mascherare con DEFAULT_CONFIG (che contiene gli account
    # IRC e produrrebbe comportamento sbagliato se caricato per il profilo SC).
    raise FileNotFoundError(
        f"File di configurazione non trovato: {CFG_PATH}\n"
        f"Verificare che il profilo sia stato selezionato correttamente e che "
        f"il file esista in: {CFG_PATH.parent}"
    )

def save_config(cfg: dict):
    import copy
    out = copy.deepcopy(cfg)
    # Normalizza e rende relativi tutti i path in last_paths prima di salvare
    lp = out.get("last_paths", {})
    for k, v in list(lp.items()):
        if isinstance(v, str) and ("/" in v):
            lp[k] = _path_to_cfg(_norm_path(v))
    CFG_PATH.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# DIALOGO PERIODO
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# DIALOGHI DI SUPPORTO
#
# PeriodDialog         — selezione mese/anno per l'analisi
# DetailDialog         — dettaglio riga + impara regola (doppio click)
# ══════════════════════════════════════════════════════════════════════

class PeriodDialog(tk.Toplevel):
    def __init__(self, parent, con_opzione_anno: bool = False):
        super().__init__(parent)
        self.title("Seleziona periodo")
        self.resizable(False, False)
        # result: (mese, anno) oppure (None, anno) se tutto l'anno
        self.result: tuple[int | None, int] | None = None
        # Default mese:
        # - giorni 1-10 del mese: propone il mese precedente (sto chiudendo il
        #   mese appena finito)
        # - dal giorno 11: propone il mese corrente (sono ormai a regime del
        #   mese in corso e voglio riconciliare quello che sta arrivando)
        now = date.today()
        if now.day <= 10:
            def_mese = 12 if now.month == 1 else now.month - 1
            def_anno = now.year - 1 if now.month == 1 else now.year
        else:
            def_mese = now.month
            def_anno = now.year

        frm = ttk.Frame(self, padding=16)
        frm.pack(fill="both", expand=True)

        row = 0
        # Checkbox "Tutto l'anno" (solo per conti che la supportano)
        self._var_tutto_anno = tk.BooleanVar(value=False)
        if con_opzione_anno:
            ttk.Checkbutton(frm, text="Tutto l'anno",
                            variable=self._var_tutto_anno,
                            command=self._toggle_tutto_anno
                            ).grid(row=row, column=0, columnspan=2,
                                   sticky="w", pady=(0, 8))
            row += 1

        ttk.Label(frm, text="Mese:").grid(row=row, column=0, sticky="w", pady=4)
        self._cb = ttk.Combobox(frm,
                                values=[f"{i+1} – {m}" for i, m in enumerate(MESI_IT)],
                                state="readonly", width=18)
        self._cb.current(def_mese - 1)
        self._cb.grid(row=row, column=1, padx=8, pady=4)
        row += 1

        ttk.Label(frm, text="Anno:").grid(row=row, column=0, sticky="w", pady=4)
        self._var_anno = tk.IntVar(value=def_anno)
        ttk.Spinbox(frm, from_=2018, to=2035, textvariable=self._var_anno,
                    width=7).grid(row=row, column=1, padx=8, pady=4, sticky="w")
        row += 1

        ttk.Button(frm, text="  OK  ", command=self._ok).grid(
            row=row, column=0, columnspan=2, pady=10)

        # Centra il dialogo sul parent
        self.update_idletasks()
        px = parent.winfo_rootx() + parent.winfo_width() // 2 - self.winfo_width() // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2 - self.winfo_height() // 2
        self.geometry(f"+{px}+{py}")

        self.transient(parent)
        self.grab_set()
        self.wait_window()

    def _toggle_tutto_anno(self):
        """Abilita/disabilita il selettore mese in base al checkbox."""
        stato = "disabled" if self._var_tutto_anno.get() else "readonly"
        self._cb.configure(state=stato)

    def _ok(self):
        anno = self._var_anno.get()
        if self._var_tutto_anno.get():
            self.result = (None, anno)   # None = tutto l'anno
        else:
            self.result = (self._cb.current() + 1, anno)
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# DIALOGO DETTAGLIO / IMPARA REGOLA
# ─────────────────────────────────────────────────────────────────────────────

class DetailDialog(tk.Toplevel):
    def __init__(self, parent, result: dict, rules: RulesEngine, categories: list[str]):
        super().__init__(parent)
        self.title("Dettaglio transazione")
        self._rules = rules
        self._cats  = categories
        self._build(result)
        self.resizable(True, False)
        self.transient(parent)
        self.grab_set()

    def _build(self, r: dict):
        frm = ttk.Frame(self, padding=16)
        frm.pack(fill="both", expand=True)
        bt    = r.get("bank_txn")
        mt    = r["money_txns"][0] if r["money_txns"] else None
        mtype = r["match_type"]
        label, bg = STATO_LABELS.get(mtype, ("?", "#fff"))

        tk.Label(frm, text=label, background=bg,
                 font=("Helvetica", 13, "bold"), padx=8, pady=4
                 ).grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))

        row = [1]
        def pair(lbl, val, bold=False):
            tk.Label(frm, text=lbl, anchor="e", width=24,
                     font=("Helvetica", 10, "bold" if bold else "normal")
                     ).grid(row=row[0], column=0, sticky="e", pady=1)
            tk.Label(frm, text=val or "—", anchor="w", wraplength=430,
                     font=("Helvetica", 10)
                     ).grid(row=row[0], column=1, sticky="w", padx=8)
            row[0] += 1

        if bt:
            pair("── BANCA ──", "", bold=True)
            pair("Data operazione:",   bt["date"].strftime("%d/%m/%Y"))
            pair("Importo:",           fmt_eur(bt["amount"]))
            pair("Descrizione:",       bt.get("descrizione", ""))
            pair("Descr. completa:",   bt.get("descrizione_completa", ""))
            pair("Fonte:",             bt.get("source", ""))

        if mt:
            pair("── MONEYSPIRE ──", "", bold=True)
            pair("Data:",        mt["date"].strftime("%d/%m/%Y"))
            pair("Importo:",     fmt_eur(mt["amount"]))
            pair("Categoria:",   mt["category"])
            pair("Payee:",       mt["payee"])
            pair("Memo:",        mt["memo"])
            if mt["has_splits"]:
                pair("Splits:", f"{len(mt['splits'])} parti")
                for i, sp in enumerate(mt["splits"], 1):
                    amt = sp["deposit"] - sp["withdrawal"]
                    pair(f"  Parte {i}:",
                         f"{fmt_eur(amt)}   {sp['category']}   {sp['memo']}")

        pair("Nota match:",   r.get("note", ""))
        pair("Confidenza:",   f"{r['confidence']:.0%}")

        # ── Sezione impara regola (solo MATCH_NONE) ──
        if mtype == MATCH_NONE and bt:
            ttk.Separator(frm, orient="horizontal").grid(
                row=row[0], column=0, columnspan=2, sticky="ew", pady=10)
            row[0] += 1
            pair("── IMPARA REGOLA ──", "", bold=True)

            tk.Label(frm, text="Pattern:", anchor="e", width=24
                     ).grid(row=row[0], column=0, sticky="e", pady=3)
            self._var_pat = tk.StringVar(value=bt.get("descrizione", "")[:50])
            pat_frame = ttk.Frame(frm)
            pat_frame.grid(row=row[0], column=1, sticky="w", padx=8)
            ttk.Entry(pat_frame, textvariable=self._var_pat, width=42).pack(side="left")
            tk.Label(pat_frame, text="  (* = qualsiasi testo)",
                     fg="gray", font=("", 9)).pack(side="left")
            row[0] += 1

            tk.Label(frm, text="Categoria:", anchor="e", width=24
                     ).grid(row=row[0], column=0, sticky="e", pady=3)
            self._var_cat = tk.StringVar()
            cat_frame = ttk.Frame(frm)
            cat_frame.grid(row=row[0], column=1, sticky="w", padx=8)
            # Campo filtro
            self._var_filtro = tk.StringVar()
            ttk.Entry(cat_frame, textvariable=self._var_filtro, width=34
                      ).pack(fill="x")
            self._var_filtro.trace_add("write", self._aggiorna_filtro_cat)
            # Listbox con scrollbar
            lb_frame = ttk.Frame(cat_frame)
            lb_frame.pack(fill="both", expand=True)
            self._lb_cat = tk.Listbox(lb_frame, height=6, width=42,
                                      exportselection=False)
            vsb = ttk.Scrollbar(lb_frame, orient="vertical",
                                 command=self._lb_cat.yview)
            self._lb_cat.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y")
            self._lb_cat.pack(side="left", fill="both", expand=True)
            self._lb_cat.bind("<<ListboxSelect>>", self._on_cat_select)
            self._popola_listbox_cat(self._cats)
            row[0] += 1

            tk.Label(frm, text="Payee:", anchor="e", width=24
                     ).grid(row=row[0], column=0, sticky="e", pady=3)
            self._var_payee = tk.StringVar()
            ttk.Entry(frm, textvariable=self._var_payee, width=42
                      ).grid(row=row[0], column=1, sticky="w", padx=8)
            row[0] += 1

            ttk.Button(frm, text="💾  Salva regola",
                       command=self._save_rule
                       ).grid(row=row[0], column=1, sticky="w", padx=8, pady=6)
            row[0] += 1

        ttk.Button(frm, text="Chiudi", command=self.destroy
                   ).grid(row=row[0], column=0, columnspan=2, pady=10)

    def _popola_listbox_cat(self, cats: list[str]):
        self._lb_cat.delete(0, "end")
        for c in cats:
            self._lb_cat.insert("end", c)

    def _aggiorna_filtro_cat(self, *_):
        filtro = self._var_filtro.get().lower()
        filtrati = [c for c in self._cats if filtro in c.lower()]
        self._popola_listbox_cat(filtrati)
        if filtrati:
            self._lb_cat.selection_set(0)
            self._var_cat.set(filtrati[0])

    def _on_cat_select(self, *_):
        sel = self._lb_cat.curselection()
        if sel:
            self._var_cat.set(self._lb_cat.get(sel[0]))

    def _save_rule(self):
        pat   = self._var_pat.get().strip()
        cat   = self._var_cat.get().strip()
        payee = self._var_payee.get().strip()
        if not pat or not cat:
            messagebox.showwarning("Attenzione",
                                   "Pattern e Categoria sono obbligatori", parent=self)
            return
        self._rules.add_or_update(pat, cat, payee, source="manual")
        messagebox.showinfo("Regola salvata",
                            f"Regola aggiunta:\n\"{pat}\" → {cat}", parent=self)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1: RICONCILIAZIONE
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# TAB 1 — RICONCILIA
#
# Gestisce il flusso principale:
#   selezione file → analisi → visualizzazione risultati → scrittura su Money
# Apre DialogRevisione (inserimenti) e DialogAggiornaFuzzy (aggiornamenti fuzzy).
# ══════════════════════════════════════════════════════════════════════

class TabRiconcilia(ttk.Frame):
    def __init__(self, parent, app: "App"):
        super().__init__(parent)
        self.app            = app
        self.results:  list[dict] = []
        self._last_ffd:      str  = ""
        self._last_account_id: int = 0
        self._sort_state:    dict = {}   # stato sort {"col": str, "asc": bool}
        self._build()

    def _build(self):
        # ── Toolbar ambiente: combobox profilo (allineata col contenuto) ──
        toolbar = ttk.Frame(self, padding=(10, 4, 10, 0))
        toolbar.pack(fill="x")

        ttk.Label(toolbar, text="Ambiente:", font=("Arial", 9, "bold")
                  ).pack(side="left", padx=(0, 4))

        self.app._var_profilo = tk.StringVar()
        self.app._cb_profilo  = ttk.Combobox(
            toolbar,
            textvariable=self.app._var_profilo,
            values=list(self.app.PROFILI.values()),
            state="readonly",
            width=20,
        )
        self.app._cb_profilo.pack(side="left", padx=(0, 4))
        self.app._cb_profilo.bind("<<ComboboxSelected>>",
                                   self.app._on_profilo_change)

        ttk.Label(toolbar, text="  Prefisso output:", font=("Arial", 9)
                  ).pack(side="left", padx=(8, 2))
        self.app._lbl_prefisso = ttk.Label(toolbar, text="—",
                                            font=("Arial", 9, "bold"),
                                            foreground="#1a5a1a")
        self.app._lbl_prefisso.pack(side="left")

        ttk.Separator(toolbar, orient="vertical").pack(side="left", fill="y",
                                                        padx=12, pady=2)
        self.app._lbl_regole = ttk.Label(toolbar, text="",
                                          font=("Arial", 9), foreground="#555555")
        self.app._lbl_regole.pack(side="left")

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=8, pady=(4, 0))

        # ── Sezione file ─────────────────────────────────────────────────
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x", padx=8, pady=6)
        top.columnconfigure(1, weight=1)

        def file_row(row_n, label, pathvar, filetypes, persist_key):
            ttk.Label(top, text=label, anchor="w", width=30
                      ).grid(row=row_n, column=0, sticky="w", pady=3)
            pe = PathEntry(top, pathvar)
            pe.grid(row=row_n, column=1, sticky="ew", padx=4)
            def browse(pv=pathvar, ft=filetypes, pk=persist_key):
                # Riposiziona sulla cartella dell'ultimo file usato per questa chiave.
                # Priorità: (1) valore attuale nel PathVar, (2) last_paths nel config,
                # (3) home directory.
                current = pv.get()
                if not current:
                    current = self.app.cfg.get("last_paths", {}).get(pk, "")
                if current:
                    parent = Path(current).expanduser().parent
                    start_dir = str(parent) if parent.exists() else str(Path.home())
                else:
                    start_dir = str(Path.home())
                p = filedialog.askopenfilename(filetypes=ft, initialdir=start_dir)
                if p:
                    pv.set(p)
                    self.app.cfg.setdefault("last_paths", {})[pk] = p
                    # Salva anche il path specifico per il conto corrente
                    if pk == "xls_cc":
                        conto_now = self.var_conto.get()
                        self.app.cfg["last_paths"][f"xls_cc_{conto_now}"] = p
                    save_config(self.app.cfg)
                    # Carica categorie dal .ffd se è quello il file
                    if pk == "ffd":
                        self.app.load_db_cats(p)
            ttk.Button(top, text="📂", width=3, command=browse
                       ).grid(row=row_n, column=2)

        last = self.app.cfg.get("last_paths", {})

        self.pv_ffd    = PathVar()
        self.pv_xls_cc = PathVar()
        self._cartella_cc = ""   # cartella multi-file (alternativa a file singolo)

        def _norm(p): return _norm_path(p) if p else ""
        if last.get("ffd"):    self.pv_ffd.set(_norm(last["ffd"]))
        if last.get("xls_cc"): self.pv_xls_cc.set(_norm(last["xls_cc"]))

        file_row(0, "File Moneyspire (.ffd):",
                 self.pv_ffd,    [("Moneyspire", "*.ffd"), ("Tutti", "*.*")], "ffd")
        file_row(1, "Movimenti originali banca:",
                 self.pv_xls_cc, [("Excel/PDF", "*.xlsx *.xls *.pdf"),
                                   ("Excel", "*.xlsx *.xls"),
                                   ("PDF", "*.pdf"),
                                   ("Tutti", "*.*")], "xls_cc")

        # Pulsante cartella multi-file (accanto al pulsante file singolo)
        def _scegli_cartella():
            current = self._cartella_cc or self.pv_xls_cc.get()
            start = str(Path(current).expanduser().parent) if current else str(Path.home())
            cartella = filedialog.askdirectory(
                title="Seleziona cartella con gli estratti conto",
                initialdir=start)
            if cartella:
                self._cartella_cc = cartella
                # Mostra nella PathVar con prefisso 📁 per distinguerla da file singolo
                self.pv_xls_cc.set(f"📁 {cartella}")
                self.app.cfg.setdefault("last_paths", {})["xls_cc_cartella"] = cartella
                save_config(self.app.cfg)

        ttk.Button(top, text="📁", width=3, command=_scegli_cartella,
                   ).grid(row=1, column=3, padx=(2, 0))

        # ── Conto + pulsanti ─────────────────────────────────────────────
        ctrl = ttk.Frame(top)
        ctrl.grid(row=2, column=0, columnspan=3, pady=8, sticky="w")

        ttk.Label(ctrl, text="Conto:").pack(side="left")
        conti = [
            nome for nome, cfg in self.app.cfg.get("conti", {}).items()
            if cfg.get("includi_in_riconciliazione", True)
        ]
        self.var_conto = tk.StringVar(value=last.get("conto", conti[0] if conti else ""))
        self._cb_conto = ttk.Combobox(ctrl, textvariable=self.var_conto,
                                      values=conti, state="readonly", width=20)
        self._cb_conto.pack(side="left", padx=6)
        self._cb_conto.bind("<<ComboboxSelected>>", self._on_conto_change)

        ttk.Button(ctrl, text="▶  Analizza",
                   command=self._run).pack(side="left", padx=10)
        self.btn_aggiorna = ttk.Button(ctrl, text="✏️  Aggiorna Money",
                   command=self._apri_revisione, state="disabled")
        self.btn_aggiorna.pack(side="left", padx=4)
        self.btn_fuzzy = ttk.Button(ctrl, text="🔧  Aggiorna Fuzzy",
                   command=self._apri_fuzzy, state="disabled")
        self.btn_fuzzy.pack(side="left", padx=4)
        ttk.Button(ctrl, text="💾  Esporta tutto (CSV)",
                   command=self._export_full).pack(side="left", padx=4)
        ttk.Button(ctrl, text="📋  Esporta mancanti (CSV)",
                   command=self._export_missing).pack(side="left", padx=4)

        # ── Sommario ─────────────────────────────────────────────────────
        self.lbl_summary = ttk.Label(self, text="", font=("Helvetica", 11))
        self.lbl_summary.pack(padx=8, pady=2, anchor="w")

        # ── Filtri ───────────────────────────────────────────────────────
        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=8, pady=2)
        ttk.Label(filt, text="Mostra:").pack(side="left")
        self._fvars: dict[str, tk.BooleanVar] = {}
        for label, key in [
            ("✅ Trovate",    MATCH_EXACT),
            ("🔶 Fuzzy",      MATCH_FUZZY),
            ("🔀 Split",      MATCH_SPLIT),
            ("🔗 Merge",      MATCH_MERGE),
            ("❌ Mancanti",   MATCH_NONE),
            ("📝 Solo Money", MATCH_SKIP),
            ("⏳ In attesa",  MATCH_PENDING),
            ("🔄 Annullate",  "annullata"),
        ]:
            v = tk.BooleanVar(value=True)
            self._fvars[key] = v
            ttk.Checkbutton(filt, text=label, variable=v,
                            command=self._refresh_table).pack(side="left", padx=4)
        ttk.Separator(filt, orient="vertical").pack(side="left", fill="y", padx=6)
        ttk.Button(filt, text="Tutti",    width=6,
                   command=lambda: self._set_filtri(True)).pack(side="left", padx=2)
        ttk.Button(filt, text="Nessuno",  width=7,
                   command=lambda: self._set_filtri(False)).pack(side="left", padx=2)

        # ── Tabella ───────────────────────────────────────────────────────
        tbl = ttk.Frame(self)
        tbl.pack(fill="both", expand=True, padx=8, pady=4)
        cols = ("Stato", "Data", "Importo", "Descrizione banca",
                "Categoria Money", "Payee Money", "Memo Money", "Spl", "Nota")
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings",
                                 selectmode="browse")
        for col, w, mw in zip(cols,
                              (115,  88,  92, 300, 200, 140, 180, 35, 280),
                              (115,  88,  92, 200, 150, 100, 120, 35, 200)):
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort(c))
            self.tree.column(col, width=w, minwidth=mw,
                             anchor="w" if w > 90 else "center", stretch=False)
        vsb = ttk.Scrollbar(tbl, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tbl, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        for k, c in COL.items():
            self.tree.tag_configure(k, background=c)
        self.tree.tag_configure(MATCH_PENDING,  background=COL.get(MATCH_PENDING, "#fde8c8"))
        self.tree.tag_configure("annullata", background=COL_ANNULLATA,
                                foreground="#999999")   # grigio tenue
        self.tree.bind("<Double-1>", self._on_dbl)

        # Categorie caricate da App dopo che tutti i tab sono costruiti

    def _apri_revisione(self):
        """Apre il dialogo di revisione e scrittura su Money."""
        if not self.results:
            messagebox.showinfo("Nessun dato", "Esegui prima un'analisi")
            return
        if not self._last_ffd or not os.path.exists(self._last_ffd):
            messagebox.showerror("Errore", "File .ffd non trovato")
            return
        dlg = DialogRevisione(self, self.app, self._last_ffd,
                              self._last_account_id, self.results,
                              mese=self._last_mese, anno=self._last_anno)
        # Riesegui analisi automaticamente solo se la scrittura è andata a buon fine
        self.wait_window(dlg)
        if getattr(dlg, "_scrittura_ok", False):
            self.app.log("  🔄 Rieseguo analisi dopo aggiornamento...")
            self._run()

    def _apri_fuzzy(self):
        """Apre il dialogo di aggiornamento match fuzzy (data/importo) su Money."""
        if not self.results:
            messagebox.showinfo("Nessun dato", "Esegui prima un'analisi")
            return
        if not self._last_ffd or not os.path.exists(self._last_ffd):
            messagebox.showerror("Errore", "File .ffd non trovato")
            return
        fuzzy = [r for r in self.results if r["match_type"] == MATCH_FUZZY]
        if not fuzzy:
            messagebox.showinfo("Nessun fuzzy", "Nessun match fuzzy da aggiornare")
            return
        dlg = DialogAggiornaFuzzy(self, self.app, self._last_ffd,
                                  self._last_account_id, fuzzy)
        # Riesegui analisi automaticamente solo se l'aggiornamento è andato a buon fine
        self.wait_window(dlg)
        if getattr(dlg, "_scrittura_ok", False):
            self.app.log("  🔄 Rieseguo analisi dopo aggiornamento fuzzy...")
            self._run()

    def _on_conto_change(self, _=None):
        conto_name = self.var_conto.get()
        self.app.cfg.setdefault("last_paths", {})["conto"] = conto_name
        save_config(self.app.cfg)

        conto_cfg = self.app.cfg.get("conti", {}).get(conto_name, {})
        tipo_file = conto_cfg.get("tipo_file_analisi", "originale_fineco")
        last = self.app.cfg.get("last_paths", {})

        # Estensioni accettate per ogni tipo di file
        _ext_ok = {
            "originale_bper":       {".xls"},
            "originale_unicredit":  {".xls"},
            "originale_nexi_xlsx":  {".xlsx"},
            "originale_fineco":     {".xlsx"},
            "originale_fineco_usd": {".xlsx"},
            "estratto_cc":          {".xlsx"},
        }
        ext_accettate = _ext_ok.get(tipo_file, {".xlsx"})

        # Prima cerca il path specifico per questo conto
        path_key = f"xls_cc_{conto_name}"
        path_per_conto = last.get(path_key, "")
        if path_per_conto and Path(path_per_conto).exists():
            self.pv_xls_cc.set(_norm_path(path_per_conto))
            self._cartella_cc = ""
        else:
            # Nessun path specifico: pulisce se l'estensione corrente è incompatibile
            path_corrente = self.pv_xls_cc.get().lstrip("📁 ").strip()
            if path_corrente and not path_corrente.startswith("/"):
                # era una cartella — la lascia solo se il tipo la supporta
                pass
            elif path_corrente:
                ext_corrente = Path(path_corrente).suffix.lower()
                if ext_corrente not in ext_accettate:
                    self.pv_xls_cc.set("")
                    self._cartella_cc = ""

    def _aggiorna_conti(self):
        """Ripopola il combobox conti dopo la scelta del profilo."""
        self._cartella_cc = ""   # reset — verrà ripristinato sotto se presente
        conti = [
            nome for nome, cfg in self.app.cfg.get("conti", {}).items()
            if cfg.get("includi_in_riconciliazione", True)
        ]
        last = self.app.cfg.get("last_paths", {})
        self._cb_conto["values"] = conti
        val = last.get("conto", conti[0] if conti else "")
        self.var_conto.set(val if val in conti else (conti[0] if conti else ""))
        def _norm(p): return _norm_path(p) if p else ""
        if last.get("ffd"): self.pv_ffd.set(_norm(last["ffd"]))

        # Ripristina cartella o file per il conto correntemente selezionato
        conto_attuale  = self.var_conto.get()
        cartella_saved = last.get("xls_cc_cartella", "")
        path_per_conto = last.get(f"xls_cc_{conto_attuale}", "")
        path_generico  = last.get("xls_cc", "")

        if path_per_conto and Path(path_per_conto).exists():
            self.pv_xls_cc.set(_norm(path_per_conto))
        elif cartella_saved and Path(cartella_saved).is_dir():
            self._cartella_cc = cartella_saved
            self.pv_xls_cc.set(f"📁 {cartella_saved}")
        elif path_generico:
            self.pv_xls_cc.set(_norm(path_generico))

    # ── Analisi ───────────────────────────────────────────────────────────

    def _run(self):
        ffd        = self.pv_ffd.get()
        conto_name = self.var_conto.get()

        # Normalizza il path (rimuove doppia slash e altri artefatti)
        if ffd:
            ffd = _norm_path(ffd)
            if ffd != self.pv_ffd.get():
                self.pv_ffd.set(ffd)

        if not ffd or not os.path.exists(ffd):
            messagebox.showerror("Errore", "Seleziona il file Moneyspire (.ffd)")
            return
        conto_cfg = self.app.cfg.get("conti", {}).get(conto_name)
        if not conto_cfg:
            messagebox.showerror("Errore", f"Conto '{conto_name}' non configurato")
            return

        # Letto qui perché serve già per il PeriodDialog (opzione "Tutto l'anno")
        # Fallback per config vecchi: riconosce Unicredit per nome conto
        tipo_file_analisi = conto_cfg.get("tipo_file_analisi")
        if not tipo_file_analisi:
            if "unicredit" in conto_name.lower():
                tipo_file_analisi = "originale_unicredit"
            elif conto_cfg.get("tipo") == "carta_credito":
                tipo_file_analisi = "estratto_cc"
            elif conto_cfg.get("valuta") == "USD":
                tipo_file_analisi = "originale_fineco_usd"
            else:
                tipo_file_analisi = "originale_fineco"

        # Opzione "tutto anno" disponibile per tutti i conti correnti
        _tipi_cc = ("originale_fineco", "originale_fineco_usd",
                    "originale_unicredit", "originale_bper", "originale_nexi_xlsx")
        # Abilita "tutto anno" anche per le carte se è selezionata una cartella multi-file
        # Controlla modalità cartella: _cartella_cc impostato, PathVar con 📁,
        # oppure PathVar che punta direttamente a una directory (senza prefisso)
        _pv_val = self.pv_xls_cc.get()
        _pv_path = _pv_val.lstrip("📁 ").strip()
        _con_cartella = (
            bool(self._cartella_cc and os.path.isdir(self._cartella_cc)) or
            (_pv_path and os.path.isdir(_pv_path))
        )
        # Se la cartella è identificata ma _cartella_cc non è ancora impostato, lo imposta
        if _con_cartella and not self._cartella_cc and _pv_path:
            self._cartella_cc = _pv_path
        dlg = PeriodDialog(self, con_opzione_anno=(
            tipo_file_analisi in _tipi_cc or _con_cartella))
        if dlg.result is None:
            return
        mese, anno = dlg.result

        if mese is None:
            # Tutto l'anno: finestra money 1 gen → 31 dic
            date_from = date(anno, 1, 1)
            date_to   = date(anno, 12, 31)
            label_periodo = f"Anno {anno}"
        else:
            date_from = date(anno, mese, 1)
            date_to   = date(anno, mese, calendar.monthrange(anno, mese)[1])
            label_periodo = f"{MESI_IT[mese-1]} {anno}"

        self.app.log(f"▶ Analisi {conto_name} — {label_periodo}")
        try:
            db    = MoneyspireDB(ffd)
            acct  = conto_cfg["ffd_account_id"]
            tipo  = conto_cfg.get("tipo", "conto_corrente")

            # Per le carte di credito, le transazioni di fine mese precedente
            # (es. 26-28 feb data operazione) appaiono nell'estratto del mese
            # successivo (data registrazione marzo). Estendiamo la finestra
            # money_txns solo indietro (per trovare le transazioni del mese
            # precedente già in Money) ma NON in avanti (evita di includere
            # transazioni del mese successivo come Solo Money).
            giorni_extra = self.app.cfg.get("giorni_fine_mese_pending", 4)
            if tipo == "carta_credito":
                from datetime import timedelta
                date_from_money = date_from - timedelta(days=giorni_extra)
                date_to_money   = date_to   # NON estendere in avanti
            else:
                date_from_money = date_from
                date_to_money   = date_to

            money_txns = db.get_transactions(acct, date_from_money, date_to_money)
            self.app.log(f"  Moneyspire: {len(money_txns)} transazioni nel periodo")

            # Un solo campo file per CC e carte — pv_xls_cc
            # In modalità cartella, _cartella_cc è già validata; altrimenti serve un file
            xls_banca = self.pv_xls_cc.get().lstrip("📁 ").strip()
            # Modalità cartella: path è una directory (con o senza prefisso 📁)
            if not self._cartella_cc and xls_banca and os.path.isdir(xls_banca):
                self._cartella_cc = xls_banca
            _usa_cartella = bool(self._cartella_cc and os.path.isdir(self._cartella_cc))

            if not _usa_cartella:
                if not xls_banca or not os.path.exists(xls_banca):
                    messagebox.showerror("Errore",
                        "Seleziona il file movimenti banca (📂) oppure una cartella (📁)")
                    db.close(); return
                # Validazione: verifica che il file corrisponda al conto selezionato
                ok_val, msg_val = valida_file_banca(xls_banca, conto_cfg)
                if not ok_val:
                    messagebox.showerror("File non corretto", msg_val)
                    db.close(); return

            if tipo == "carta_credito":
                if _usa_cartella:
                    self.app.log(f"  Modalità cartella: {self._cartella_cc}")
                    if tipo_file_analisi == "originale_nexi_xlsx":
                        # Cartella con xlsx Nexi — unifica e deduplicca
                        import glob as _glob
                        files_nexi = sorted(_glob.glob(
                            os.path.join(self._cartella_cc, "*.xlsx")))
                        bank_txns = []
                        seen_n: set = set()
                        for f in files_nexi:
                            try:
                                for t in parse_nexi_xlsx(f, mese, anno):
                                    key = (t["date"], t["amount"], t["descrizione"])
                                    if key not in seen_n:
                                        seen_n.add(key)
                                        bank_txns.append(t)
                            except Exception:
                                pass
                        bank_txns.sort(key=lambda t: t["date"])
                    else:
                        num_carta = conto_cfg.get("numeri_carta", conto_cfg.get("numero_carta"))
                        bank_txns = parse_fineco_cc_cartella(
                            self._cartella_cc, num_carta, mese, anno)
                    self.app.log(f"  {len(bank_txns)} transazioni trovate")
                elif tipo_file_analisi == "originale_nexi_xlsx":
                    bank_txns = parse_nexi_xlsx(xls_banca, mese, anno)
                else:
                    num_carta = conto_cfg.get("numeri_carta", conto_cfg.get("numero_carta"))
                    bank_txns = parse_fineco_cc(xls_banca, num_carta, mese, anno)

            elif tipo_file_analisi == "originale_unicredit":
                bank_txns = parse_unicredit_ccm(xls_banca, mese, anno)

            elif tipo_file_analisi == "originale_bper":
                bank_txns = parse_bper(xls_banca, mese, anno)

            elif tipo_file_analisi == "originale_fineco_usd" \
                    or conto_cfg.get("valuta") == "USD":
                from ms_engine import parse_fineco_conto_usd
                bank_txns = parse_fineco_conto_usd(xls_banca, mese, anno)

            else:
                bank_txns = parse_fineco_conto_originale(xls_banca, mese, anno)
            # Raggruppa coppie cedola/ritenuta e transazioni correlate
            bank_txns = raggruppa_cedole_ritenute(bank_txns)
            # Conto USD: Money registra dividendi/cedole al NETTO (una riga),
            # mentre la banca espone lordo + ritenuta separati. Fonde le coppie
            # al netto così il matching trova la riga unica già presente in Money.
            if tipo_file_analisi == "originale_fineco_usd" \
                    or conto_cfg.get("valuta") == "USD":
                bank_txns = fondi_cedole_per_match(bank_txns)
            if tipo != "carta_credito":
                bank_txns = raggruppa_transazioni_correlate(bank_txns)

            self.app.log(f"  Banca: {len(bank_txns)} transazioni nel periodo")

            engine = ReconcileEngine(db, self.app.cfg)
            self.results = engine.reconcile(bank_txns, money_txns, account_id=acct)
            db.close()
            # Per le carte: le Solo Money a fine mese sono probabilmente
            # nell'estratto del mese successivo → rimarca come "In attesa"
            if tipo == "carta_credito":
                giorni = self.app.cfg.get("giorni_fine_mese_pending", 4)
                self.results = marca_solo_money_in_attesa(self.results, giorni)

            # Applica suggerimenti regole sui mancanti
            for r in self.results:
                if r["match_type"] == MATCH_NONE and r["bank_txn"]:
                    sg = self.app.rules.apply(r["bank_txn"].get("raw_text", ""))
                    if sg:
                        r["suggested_category"] = sg.get("category", "")
                        r["suggested_payee"]    = sg.get("payee", "")

            self._refresh_table()
            n_miss = sum(1 for r in self.results if r["match_type"] == MATCH_NONE)
            self.app.log(f"  Completato. {n_miss} mancanti su {len(bank_txns)} transazioni banca")

            # ── Confronto saldi (solo Fineco CC per mese singolo, non carte, non tutto l'anno) ──
            _tipi_con_saldo = ("originale_fineco", "originale_fineco_usd")
            if tipo != "carta_credito" and mese is not None \
                    and tipo_file_analisi in _tipi_con_saldo:
                try:
                    s_banca = leggi_saldo_fineco(xls_banca, mese, anno)
                    var_banca = leggi_variazione_mensile_fineco(xls_banca, mese, anno)
                    var_money = leggi_variazione_mensile_money(db, acct, mese, anno)
                    soglia = self.app.cfg.get("soglia_aggiustamento_saldo", 5.0)

                    if s_banca is not None:
                        self.app.log(f"  Saldo Fineco fine {mese:02d}/{anno}: {fmt_eur(s_banca)}")

                    if var_banca is not None and var_money is not None:
                        diff_var = round(var_banca - var_money, 2)
                        self.app.log(
                            f"  Variazione mensile — Banca: {fmt_eur(var_banca)}"
                            f"  Money: {fmt_eur(var_money)}"
                            f"  Δ: {fmt_eur(diff_var)}")
                        if abs(diff_var) > soglia:
                            self.app.log(
                                f"  ⚠️  Differenza variazione mensile > {soglia}€ — "
                                f"verifica le transazioni mancanti "
                                f"({fmt_eur(diff_var)})")
                        else:
                            self.app.log(f"  ✅ Variazione mensile OK (Δ < {soglia}€)")
                except Exception as e:
                    self.app.log(f"  (confronto saldi non disponibile: {e})")

            # Abilita il bottone Aggiorna Money se ci sono mancanti
            self.btn_aggiorna.config(state="normal" if n_miss > 0 else "disabled")
            # Abilita il bottone Aggiorna Fuzzy se ci sono match fuzzy
            n_fuzzy = sum(1 for r in self.results if r["match_type"] == MATCH_FUZZY)
            self.btn_fuzzy.config(state="normal" if n_fuzzy > 0 else "disabled")
            self._last_ffd = ffd
            self._last_account_id = acct
            self._last_mese = mese
            self._last_anno = anno

            # Salva i path correnti nel config dopo ogni analisi riuscita
            lp = self.app.cfg.setdefault("last_paths", {})
            lp["ffd"]   = ffd
            lp["xls_cc"] = xls_banca
            lp["conto"]  = conto_name
            lp[f"xls_cc_{conto_name}"] = xls_banca
            save_config(self.app.cfg)

        except Exception as e:
            import traceback
            self.app.log(f"  ERRORE: {e}")
            self.app.log(traceback.format_exc())
            messagebox.showerror("Errore durante l'analisi", str(e))

    def _refresh_table(self):
        self.tree.delete(*self.tree.get_children())
        for r in self.results:
            mtype = r["match_type"]
            # Filtro speciale per annullate — distinto da Solo Money
            if r.get("annullata"):
                if not self._fvars.get("annullata", tk.BooleanVar(value=True)).get():
                    continue
            elif not self._fvars.get(mtype, tk.BooleanVar(value=True)).get():
                continue
            bt = r.get("bank_txn")
            mt = r["money_txns"][0] if r["money_txns"] else None
            label, _ = STATO_LABELS.get(mtype, ("?", ""))
            if r.get("annullata"):
                label = "🔄 Annullata"
            # Categoria: preferisce quella Money, poi il suggerimento da regola
            cat   = (mt["category"] if mt else "") or r.get("suggested_category", "")
            payee = (mt["payee"]    if mt else "") or r.get("suggested_payee", "")
            # Per MATCH_SKIP (Solo Money): bt=None, i dati vengono da mt
            data_str   = (bt["date"].strftime("%d/%m/%Y") if bt
                          else mt["date"].strftime("%d/%m/%Y") if mt else "")
            importo_str = (fmt_eur(bt["amount"]) if bt
                           else fmt_eur(mt["amount"]) if mt else "")
            desc_str    = bt.get("descrizione", "") if bt else ""
            # Le righe annullate usano un tag/colore speciale
            tag = "annullata" if r.get("annullata") else mtype
            self.tree.insert("", "end", tags=(tag,), values=(
                label,
                data_str,
                importo_str,
                desc_str,
                cat, payee,
                mt["memo"] if mt else "",
                str(len(mt["splits"])) if mt and mt["has_splits"] else "",
                r.get("note", "")
            ))
        self._update_summary()

    def _update_summary(self):
        if not self.results:
            self.lbl_summary.config(text="")
            return
        cnt: dict[str, int] = {}
        for r in self.results:
            cnt[r["match_type"]] = cnt.get(r["match_type"], 0) + 1
        parts = [f"{STATO_LABELS[k][0]}: {cnt[k]}"
                 for k in (MATCH_EXACT, MATCH_FUZZY, MATCH_SPLIT,
                            MATCH_MERGE, MATCH_NONE, MATCH_SKIP)
                 if cnt.get(k, 0)]
        self.lbl_summary.config(text="   ".join(parts))

    def _set_filtri(self, valore: bool):
        """Seleziona o deseleziona tutti i filtri."""
        for v in self._fvars.values():
            v.set(valore)
        self._refresh_table()

    def _sort(self, col: str):
        """Ordinamento colonna: primo click ascendente, secondo discendente."""
        treeview_sort(self.tree, col, self._sort_state,
                      date_cols=("Data",),
                      number_cols=("Importo",))
        self._reapply_row_colors()

    def _reapply_row_colors(self):
        """Riapplica i tag colore alle righe dopo un riordinamento.
        I colori sono determinati dal match_type della riga, non dalla posizione,
        quindi il tag è già corretto. Forziamo un ridisegno aggiornando ogni riga.
        """
        for k in self.tree.get_children(""):
            tags = self.tree.item(k, "tags")
            if tags:
                self.tree.item(k, tags=tags)

    def _on_dbl(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        visible = [r for r in self.results
                   if self._fvars.get(r["match_type"],
                                      tk.BooleanVar(value=True)).get()]
        if idx < len(visible):
            cats = sorted(set(self.app.db_cats.values())) if self.app.db_cats else []
            DetailDialog(self, visible[idx], self.app.rules, cats)
            self._refresh_table()

    # ── Export ───────────────────────────────────────────────────────────

    def _export_full(self):
        if not self.results:
            messagebox.showinfo("Nessun dato", "Esegui prima un'analisi"); return
        prefisso = self.app.cfg.get("prefisso_output", "")
        conto    = self.var_conto.get().replace(" ", "_").replace("/", "-")
        init_dir = str(Path.home() / "Documents" / "download")
        p = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialdir=init_dir,
            initialfile=f"{prefisso}{conto}_riconciliazione_completa.csv")
        if p:
            n = export_full_csv(self.results, p)
            messagebox.showinfo("Esportato", f"{n} righe esportate in:\n{p}")
            self.app.log(f"  Export completo: {n} righe → {p}")

    def _export_missing(self):
        if not self.results:
            messagebox.showinfo("Nessun dato", "Esegui prima un'analisi"); return
        prefisso = self.app.cfg.get("prefisso_output", "")
        conto    = self.var_conto.get().replace(" ", "_").replace("/", "-")
        init_dir = str(Path.home() / "Documents" / "download")
        p = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialdir=init_dir,
            initialfile=f"{prefisso}{conto}_transazioni_mancanti.csv")
        if p:
            n = export_missing_csv(self.results, p, self.app.rules)
            messagebox.showinfo("Esportato", f"{n} transazioni mancanti in:\n{p}")
            self.app.log(f"  Export mancanti: {n} righe → {p}")



# ─────────────────────────────────────────────────────────────────────────────
# DIALOGO REVISIONE E SCRITTURA SU MONEY
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# DIALOG REVISIONE INSERIMENTI
#
# Mostra le transazioni da inserire in Money con categoria suggerita.
# Permette modifica categoria via DialogModificaCategoria (doppio click).
# Flusso: backup → scrittura su copia → verifica forte → sostituzione.
# Al termine scrive il log di audit (operazione: INSERIMENTO).
# ══════════════════════════════════════════════════════════════════════

class DialogRevisione(tk.Toplevel):
    """
    Finestra di revisione prima della scrittura su .ffd.
    Mostra le transazioni da inserire, permette di modificare la categoria,
    poi scrive sul DB con backup automatico.
    """

    def __init__(self, parent, app: "App", ffd_path: str,
                 account_id: int, results: list[dict],
                 mese: int | None = None, anno: int | None = None):
        super().__init__(parent)
        self.title("Revisione e scrittura su Moneyspire")
        self.app        = app
        self.ffd_path   = ffd_path
        self.account_id = account_id
        self._scrittura_ok = False   # True dopo scrittura riuscita
        cat_map = {n.lower(): i for i, n in app.db_cats.items()}
        try:
            self.da_inserire = costruisci_transazioni_da_risultati(
                results, account_id, app.rules, cat_map, app.cfg)
            app.log(f"  Revisione: {len(self.da_inserire)} transazioni da elaborare")
        except Exception as e:
            self.da_inserire = []
            import traceback
            app.log(f"  ❌ Errore costruzione revisione: {e}")
            app.log(traceback.format_exc())

        # Categorie e payees per combobox (ordinati)
        self.cats_lista   = sorted(set(app.db_cats.values()))
        self.payees_lista = sorted(app.db_payees.values()) if hasattr(app, "db_payees") else []

        # Estrai mese/anno dall'ultima analisi (serve per verifica)
        import re as _re
        self._mese = self._anno = 0

        # BooleanVar per includere/escludere ogni transazione
        self._chk_ins: list[tk.BooleanVar] = [
            tk.BooleanVar(value=(t.get("_categoria_suggerita", "") != "Ignora"))
            for t in self.da_inserire
        ]
        for t in self.da_inserire:
            if t.get("txn_date"):
                self._mese = t["txn_date"].month
                self._anno = t["txn_date"].year
                break

        self.geometry("1100x700")
        self.minsize(900, 500)
        self._build()
        self.update_idletasks()   # forza rendering completo prima di grab
        self.transient(parent)
        self.grab_set()
        self.lift()               # porta la finestra in primo piano

    def _build(self):
        # ── Intestazione ──────────────────────────────────────────────
        hdr = ttk.Frame(self, padding=(10, 6))
        hdr.pack(fill="x")
        n_tot    = len(self.da_inserire)
        n_ignora = sum(1 for t in self.da_inserire
                       if t.get("_categoria_suggerita") == "Ignora")
        n_ins    = n_tot - n_ignora
        self._lbl_contatore = ttk.Label(hdr,
                  text=f"Da inserire: {n_ins}  (+ {n_ignora} da ignorare)",
                  font=("Helvetica", 12, "bold"))
        self._lbl_contatore.pack(side="left")
        ttk.Label(hdr,
                  text="Clicca su ▶Ins per includere/escludere  |  Doppio click per modificare categoria",
                  foreground="#666").pack(side="right")

        # ── Tabella transazioni ────────────────────────────────────────
        tbl_frame = ttk.Frame(self)
        tbl_frame.pack(fill="both", expand=True, padx=8, pady=4)

        cols = ("▶Ins", "Data", "Importo", "Descrizione banca",
                "Categoria", "Splits", "Memo")
        self.tree = ttk.Treeview(tbl_frame, columns=cols,
                                 show="headings", selectmode="browse")
        self._sort_state: dict = {}
        for col, w in zip(cols, (52, 90, 100, 260, 220, 50, 180)):
            anchor = "center" if col in ("▶Ins", "Splits") else "w"
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._toggle_tutti()
                              if c == "▶Ins" else treeview_sort(
                                  self.tree, c, self._sort_state,
                                  date_cols=("Data",), number_cols=("Importo",)))
            self.tree.column(col, width=w, anchor=anchor)
        vsb = ttk.Scrollbar(tbl_frame, orient="vertical",
                            command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        # Colori riga
        self.tree.tag_configure("inserisci", background="#d4edda")
        self.tree.tag_configure("ignora",    background="#e2e3e5")
        self.tree.tag_configure("esclusa",   background="#f0f0f0")
        self.tree.tag_configure("split",     background="#cce5ff")
        self.tree.tag_configure("da_class",  background="#fff3cd")

        self.tree.bind("<Double-1>",         self._on_dbl)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Button-1>",         self._on_click)

        self._popola_tabella()

        # ── Pannello dettaglio ────────────────────────────────────────
        det = ttk.LabelFrame(self, text="Descrizione completa", padding=6)
        det.pack(fill="x", padx=8, pady=2)
        self.lbl_det = tk.Text(det, height=3, wrap="word",
                               font=("Helvetica", 10),
                               state="disabled", bg="#f8f8f8")
        self.lbl_det.pack(fill="x")

        # ── Pulsanti ─────────────────────────────────────────────────
        btn_bar = ttk.Frame(self, padding=(8, 6))
        btn_bar.pack(fill="x")
        ttk.Button(btn_bar, text="✅  Esegui scrittura su Money",
                   command=self._esegui).pack(side="right", padx=4)
        ttk.Button(btn_bar, text="Annulla",
                   command=self.destroy).pack(side="right", padx=4)
        ttk.Button(btn_bar, text="☑ Seleziona tutti",
                   command=lambda: self._set_tutti(True)).pack(side="left", padx=4)
        ttk.Button(btn_bar, text="☐ Deseleziona tutti",
                   command=lambda: self._set_tutti(False)).pack(side="left", padx=4)
        self.lbl_stato = ttk.Label(btn_bar, text="", foreground="#666")
        self.lbl_stato.pack(side="left", padx=8)

    def _popola_tabella(self):
        self.tree.delete(*self.tree.get_children())
        if not self.da_inserire:
            self.tree.insert("", "end", values=(
                "", "⚠️ Nessuna transazione da inserire", "", "", "", "", ""))
            return
        for i, t in enumerate(self.da_inserire):
            try:
                _cp_default = self.app.cfg.get("contropartita_default", "Da classificare")
                cat   = t.get("_categoria_suggerita","") or _cp_default
                amt   = t["deposit"] - t["withdrawal"]
                inclusa = self._chk_ins[i].get()
                if not inclusa:
                    chk_v, tag = "☐", "esclusa"
                elif cat == "Ignora":
                    chk_v, tag = "⏭", "ignora"
                elif t.get("splits"):
                    chk_v, tag = "☑", "split"
                elif cat in (_cp_default, "—", ""):
                    chk_v, tag = "☑", "da_class"
                else:
                    chk_v, tag = "☑", "inserisci"
                self.tree.insert("", "end", iid=str(i), tags=(tag,), values=(
                    chk_v,
                    t["txn_date"].strftime("%d/%m/%Y"),
                    fmt_eur(amt),
                    t.get("_descrizione_banca","")[:45],
                    cat,
                    str(len(t["splits"])) if t.get("splits") else "",
                    t.get("memo","")
                ))
            except Exception as e:
                self.tree.insert("", "end", values=(
                    "❌", f"Errore riga {i}", "", str(e)[:40], "", "", ""))

    def _on_click(self, event):
        """Click su colonna ▶Ins → toggle includi/escludi riga."""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)
        iid = self.tree.identify_row(event.y)
        if col == "#1" and iid:   # colonna ▶Ins
            try:
                i = int(iid)
                self._chk_ins[i].set(not self._chk_ins[i].get())
                self._aggiorna_riga_ins(i)
                self._aggiorna_header_contatore()
            except (ValueError, IndexError):
                pass

    def _aggiorna_riga_ins(self, i: int):
        """Aggiorna la visualizzazione della riga i dopo toggle."""
        if i >= len(self.da_inserire):
            return
        t = self.da_inserire[i]
        _cp_default = self.app.cfg.get("contropartita_default", "Da classificare")
        cat = t.get("_categoria_suggerita", "") or _cp_default
        inclusa = self._chk_ins[i].get()
        if not inclusa:
            chk_v, tag = "☐", "esclusa"
        elif cat == "Ignora":
            chk_v, tag = "⏭", "ignora"
        elif t.get("splits"):
            chk_v, tag = "☑", "split"
        elif cat in (_cp_default, "—", ""):
            chk_v, tag = "☑", "da_class"
        else:
            chk_v, tag = "☑", "inserisci"
        vals = list(self.tree.item(str(i), "values"))
        vals[0] = chk_v
        self.tree.item(str(i), values=vals, tags=(tag,))

    def _toggle_tutti(self):
        """Click sull'header ▶Ins → seleziona/deseleziona tutte."""
        n_on = sum(1 for v in self._chk_ins if v.get())
        nuovo = n_on < len(self._chk_ins)
        self._set_tutti(nuovo)

    def _set_tutti(self, valore: bool):
        """Imposta tutte le checkbox a valore (True=includi, False=escludi)."""
        for i, v in enumerate(self._chk_ins):
            v.set(valore)
            self._aggiorna_riga_ins(i)
        self._aggiorna_header_contatore()

    def _aggiorna_header_contatore(self):
        """Aggiorna il contatore nell'intestazione."""
        n_ins = sum(1 for i, v in enumerate(self._chk_ins)
                    if v.get() and self.da_inserire[i].get("_categoria_suggerita","") != "Ignora")
        n_ign = sum(1 for i, v in enumerate(self._chk_ins)
                    if v.get() and self.da_inserire[i].get("_categoria_suggerita","") == "Ignora")
        n_exc = sum(1 for v in self._chk_ins if not v.get())
        txt = f"Da inserire: {n_ins}"
        if n_ign: txt += f"  (+ {n_ign} da ignorare)"
        if n_exc: txt += f"  — {n_exc} escluse"
        self._lbl_contatore.config(text=txt)

    def _on_select(self, _=None):
        """Mostra descrizione completa nel pannello dettaglio."""
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        t   = self.da_inserire[idx]
        desc2 = t.get("_descrizione_completa","") or ""
        self.lbl_det.configure(state="normal")
        self.lbl_det.delete("1.0", "end")
        self.lbl_det.insert("1.0", desc2)
        self.lbl_det.configure(state="disabled")

    def _on_dbl(self, _=None):
        """Apre dialogo modifica categoria sulla riga selezionata."""
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        t   = self.da_inserire[idx]
        if t.get("_categoria_suggerita") == "Ignora":
            return  # non si modifica

        dlg = DialogModificaCategoria(
            self, t, self.cats_lista,
            on_save=lambda cat, memo, ignora, payee="":                 self._aggiorna_riga(idx, cat, memo, ignora, payee),
            payees=self.payees_lista,
            payee_cats=getattr(self.app, "db_payee_cats", {}),
            cats_by_id=self.app.db_cats,
            contropartita_default=self.app.cfg.get("contropartita_default", "Da classificare"),
        )

    def _aggiorna_riga(self, idx: int, nuova_cat: str,
                       nuovo_memo: str, ignora: bool, payee: str = ""):
        t = self.da_inserire[idx]
        if ignora:
            t["_categoria_suggerita"] = "Ignora"
            t["category_id"]          = None
        else:
            cat_id = {n.lower(): i
                      for i, n in self.app.db_cats.items()
                      }.get(nuova_cat.lower())
            # Risolvi payee_id
            payee_id = None
            if payee and hasattr(self.app, "db_payees"):
                payee_id = {n.lower(): i
                            for i, n in self.app.db_payees.items()
                            }.get(payee.lower())
            t["_categoria_suggerita"] = nuova_cat
            t["category_id"]          = cat_id
            t["memo"]                 = nuovo_memo
            t["payee_id"]             = payee_id
            t["_payee_suggerito"]     = payee
        self._popola_tabella()
        self.tree.selection_set(str(idx))
        self.tree.see(str(idx))

    def _esegui(self):
        """Verifica Moneyspire chiuso, backup, scrittura, verifica."""
        import subprocess

        # 1. Verifica Moneyspire non in esecuzione
        r = subprocess.run(["pgrep", "-ix", "Moneyspire"],
                           capture_output=True, text=True)
        if r.returncode == 0:
            messagebox.showerror(
                "Moneyspire aperto",
                "Chiudi Moneyspire prima di procedere.",
                parent=self)
            return

        # 2. Riepilogo — considera solo le transazioni incluse (checkbox ☑)
        _cp_default = self.app.cfg.get("contropartita_default", "Da classificare")
        da_elaborare = [t for i, t in enumerate(self.da_inserire)
                        if self._chk_ins[i].get()]
        escluse = len(self.da_inserire) - len(da_elaborare)

        n_ins    = sum(1 for t in da_elaborare
                       if t.get("_categoria_suggerita") != "Ignora")
        n_ignora = len(da_elaborare) - n_ins
        n_da_cl  = sum(1 for t in da_elaborare
                       if t.get("_categoria_suggerita") in
                          (_cp_default, "", "—")
                       and t.get("_categoria_suggerita") != "Ignora")

        msg = (f"Stai per inserire {n_ins} transazioni in Moneyspire.\n"
               f"  - Con categoria:   {n_ins - n_da_cl}\n"
               f"  - {_cp_default}: {n_da_cl}\n"
               f"  - Ignorate:        {n_ignora}\n")
        if escluse:
            msg += f"  - Escluse (☐):     {escluse}\n"
        msg += f"\nVerrà creato un backup automatico del file .ffd.\n\nProcedere?"
        if not messagebox.askyesno("Conferma scrittura", msg, parent=self):
            return

        # 3. Crea backup + copia di lavoro
        try:
            work_path, backup_path = prepara_db_scrittura(self.ffd_path, _PROFILO)
        except RuntimeError as e:
            messagebox.showerror("Impossibile procedere", str(e), parent=self)
            return
        except Exception as e:
            messagebox.showerror("Errore backup", str(e), parent=self)
            return

        from pathlib import Path
        self.app.log(f"  Backup: {Path(backup_path).name}")
        self.app.log(f"  Copia di lavoro: {Path(work_path).name}")

        # 4. Scrittura sulla COPIA DI LAVORO (non sull'originale)
        try:
            writer   = MoneyWriter(work_path)
            inserite = ignorate = errori = 0
            txns_inserite: list[dict] = []   # per verifica forte
            righe_audit:   list[dict] = []   # per log di audit

            for t in da_elaborare:   # solo le transazioni incluse
                if t.get("_categoria_suggerita") == "Ignora":
                    ignorate += 1
                    righe_audit.append({
                        "tipo": "SKIP", "txn_id": None,
                        "data": t["txn_date"],
                        "importo": t["deposit"] - t["withdrawal"],
                        "categoria": t.get("_categoria_suggerita", ""),
                        "payee": "", "memo": t.get("memo", "") or t.get("_descrizione_banca", ""),
                        "note": "ignorata dall'utente"
                    })
                    continue

                cat_id      = t.get("category_id")
                transfer_to = t.get("_transfer_to","")
                if transfer_to and not cat_id:
                    cat_id = writer.account_id(transfer_to)
                    if cat_id is None:
                        self.app.log(
                            f"  ⚠️ Giroconto: conto '{transfer_to}' non trovato in "
                            f"Moneyspire — transazione inserita SENZA categoria "
                            f"(verificare transfer_default o _transfer_to in config)"
                        )

                try:
                    new_id = writer.inserisci_transazione(
                        account_id  = self.account_id,
                        txn_date    = t["txn_date"],
                        deposit     = t["deposit"],
                        withdrawal  = t["withdrawal"],
                        memo        = t.get("memo",""),
                        category_id = cat_id,
                        payee_id    = t.get("payee_id"),
                        splits      = t.get("splits")
                    )
                    inserite += 1
                    txns_inserite.append(t)
                    righe_audit.append({
                        "tipo": "INS", "txn_id": new_id,
                        "data": t["txn_date"],
                        "importo": t["deposit"] - t["withdrawal"],
                        "categoria": t.get("_categoria_suggerita", ""),
                        "payee": "",
                        "memo": t.get("memo", "") or t.get("_descrizione_banca", ""),
                        "note": ("split_cedola" if t.get("_tipo") == "split_cedola"
                                 else transfer_to if transfer_to else "")
                    })
                except Exception as e:
                    errori += 1
                    self.app.log(f"  ❌ {t['txn_date']} "
                                 f"{t['_descrizione_banca'][:30]}: {e}")
                    righe_audit.append({
                        "tipo": "ERR", "txn_id": None,
                        "data": t["txn_date"],
                        "importo": t["deposit"] - t["withdrawal"],
                        "categoria": t.get("_categoria_suggerita", ""),
                        "payee": "", "memo": t.get("_descrizione_banca", ""),
                        "note": str(e)[:60]
                    })

            writer.close()

        except Exception as e:
            messagebox.showerror("Errore scrittura", str(e), parent=self)
            import os; os.unlink(work_path)
            return

        # 5. Verifica copia di lavoro
        # Deriva il range di date direttamente dalle transazioni scritte
        # (robusto per mese singolo E tutto l'anno)
        from datetime import date as _date_cls
        import calendar as _cal
        txn_dates = [t["txn_date"] for t in self.da_inserire
                     # Nota: "_ignore" non viene mai impostato nel flusso attuale
                     # (l'esclusione reale vive in self._chk_ins). Il filtro è
                     # innocuo ma è un residuo di refactoring — da rimuovere
                     # se si unifica la logica di esclusione.
                     if t.get("txn_date") and not t.get("_ignore")]
        if txn_dates:
            ver_from = min(txn_dates)
            ver_to   = max(txn_dates)
        elif self._anno > 0 and self._mese and self._mese > 0:
            ver_from = _date_cls(self._anno, self._mese, 1)
            ver_to   = _date_cls(self._anno, self._mese,
                                 _cal.monthrange(self._anno, self._mese)[1])
        else:
            ver_from = _date_cls(self._anno, 1, 1)
            ver_to   = _date_cls(self._anno, 12, 31)
        ok_ver, msg_ver = verifica_scrittura(
            work_path, self.account_id,
            ver_from, ver_to,
            inserite,
            expected_txns=txns_inserite)
        self.app.log(f"  {msg_ver}")

        # 6. Riepilogo e conferma sostituzione
        self.app.log(f"  Scritte: {inserite}  Ignorate: {ignorate}  Errori: {errori}")
        msg_ris = (
            f"Scritte sulla copia di lavoro: {inserite}\n"
            f"Ignorate: {ignorate}   Errori: {errori}\n\n"
            f"Verifica: {msg_ver}\n\n"
            f"Backup originale: {Path(backup_path).name}\n\n"
            "Vuoi sostituire il file Moneyspire originale\n"
            "con la copia di lavoro verificata?")
        if errori == 0 and ok_ver:
            if messagebox.askyesno("Conferma sostituzione", msg_ris, parent=self):
                finalizza_db(self.ffd_path, work_path)
                self.app.log("  ✅ File .ffd sostituito — backup conservato")
                self._scrittura_ok = True   # segnale per riesegui analisi
                # ── Audit log ────────────────────────────────────────────
                if self._mese and self._mese > 0:
                    periodo = f"{MESI_IT[self._mese-1]} {self._anno}"
                elif self._anno > 0:
                    periodo = f"Anno {self._anno}"
                else:
                    periodo = f"{ver_from}–{ver_to}"
                riepilogo_audit = (f"{inserite} inserite  {ignorate} ignorate  "
                                   f"{errori} errori  — Verifica: {msg_ver}")
                conto_nome = self.app.cfg.get("last_paths", {}).get("conto", "?")
                write_audit_log(
                    operazione="INSERIMENTO",
                    ffd_path=self.ffd_path,
                    backup_path=backup_path,
                    conto_nome=conto_nome,
                    account_id=self.account_id,
                    righe=righe_audit,
                    riepilogo=riepilogo_audit,
                    periodo=periodo)
                self.app.log(f"  📋 Audit log: {_audit_log_path().name}")
                # Persiste i contatori hits delle regole usate in questa sessione
                self.app.rules.save()
                # ─────────────────────────────────────────────────────────
                messagebox.showinfo("Completato",
                    f"File Moneyspire aggiornato.\n"
                    f"Backup conservato: {Path(backup_path).name}\n\n"
                    f"Ora puoi aprire Moneyspire per verificare.",
                    parent=self)
            else:
                import os; os.unlink(work_path)
                self.app.log("  ⚠️ Sostituzione annullata — copia di lavoro eliminata")
        else:
            import os; os.unlink(work_path)
            messagebox.showerror("Errore",
                f"Verifica fallita o errori di scrittura.\n{msg_ver}\n"
                f"Originale non modificato. Copia di lavoro eliminata.",
                parent=self)
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# DIALOGO AGGIORNAMENTO MATCH FUZZY (data e/o importo)
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# DIALOG AGGIORNAMENTO MATCH FUZZY
#
# Mostra i match fuzzy (🔶 Trovata ±gg) e permette di aggiornare
# selettivamente data e/o importo delle transazioni in Money.
# Regole di sicurezza:
#   - Split: data aggiornabile, importo MAI modificabile (🔒)
#   - Segno invertito banca/Money: importo bloccato (⚠️)
#   - Campo non diverso: mostra — e non è cliccabile
# Click su riga → toggle data; click su colonna ▶Imp → toggle importo.
# Al termine scrive il log di audit (operazione: AGGIORNA_FUZZY).
# ══════════════════════════════════════════════════════════════════════

class DialogAggiornaFuzzy(tk.Toplevel):
    """
    Mostra i match fuzzy (data sfasata o importo leggermente diverso) e
    permette di selezionare quali aggiornare nel DB Moneyspire.

    Per ogni riga fuzzy vengono mostrati:
      - Data banca vs data Money (con checkbox per aggiornare la data)
      - Importo banca vs importo Money (con checkbox per aggiornare l'importo)
      - Descrizione banca e nota del match

    L'utente seleziona le righe desiderate tramite checkbox e clicca
    "Applica aggiornamenti" per scrivere le modifiche sul DB.
    """

    def __init__(self, parent, app, ffd_path: str, account_id: int,
                 fuzzy_results: list[dict]):
        super().__init__(parent)
        self.title("🔧 Aggiornamento match fuzzy")
        self.resizable(True, True)
        self.geometry("1050x540")

        self.app        = app
        self.ffd_path   = ffd_path
        self.account_id = account_id
        self.results    = fuzzy_results
        self._scrittura_ok = False   # True dopo aggiornamento riuscito
        # _chk_data[i]    = BooleanVar → aggiornare la data?
        # _chk_importo[i] = BooleanVar → aggiornare l'importo?
        self._chk_data:    list[tk.BooleanVar] = []
        self._chk_importo: list[tk.BooleanVar] = []

        self._build_ui()
        self.grab_set()

    # ── UI ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        # Intestazione
        ttk.Label(self, text=(
            "Seleziona le contabili fuzzy da aggiornare in Moneyspire.\n"
            "Clicca su una riga per attivare/disattivare l'aggiornamento data. "
            "Clicca sulla colonna ▶Imp per attivare/disattivare l'aggiornamento importo."
        ), justify="left").pack(padx=10, pady=(8, 4), anchor="w")

        # Frame tabella con scrollbar
        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, padx=10, pady=4)

        cols = ("sel_data", "sel_imp", "data_banca", "data_money",
                "importo_banca", "importo_money", "descrizione", "nota")
        hdrs = ("▶Data", "▶Imp", "Data Banca", "Data Money",
                "Importo Banca", "Importo Money", "Descrizione", "Nota match")
        widths = (55, 55, 88, 88, 105, 105, 270, 220)

        self.tree = ttk.Treeview(frm, columns=cols, show="headings",
                                 selectmode="none", height=16)
        self._sort_state: dict = {}
        _date_c   = ("Data Banca", "Data Money")
        _number_c = ("Importo Banca", "Importo Money")
        for col, hdr, w in zip(cols, hdrs, widths):
            self.tree.heading(col, text=hdr,
                              command=lambda c=col: treeview_sort(
                                  self.tree, c, self._sort_state,
                                  date_cols=_date_c, number_cols=_number_c))
            self.tree.column(col, width=w, anchor="center" if w < 150 else "w",
                             stretch=(w > 150))

        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        # Popolamento righe
        self._iids: list[str] = []
        self._is_split:    list[bool] = []   # importo bloccato (split o segno inv.)
        self._data_diversa: list[bool] = []  # True se la data è effettivamente diversa
        self._imp_diverso:  list[bool] = []  # True se l'importo è effettivamente diverso
        for i, r in enumerate(self.results):
            bt  = r["bank_txn"]
            mt  = r["money_txns"][0] if r["money_txns"] else None

            d_banca  = bt["date"].strftime("%d/%m/%Y") if bt["date"] else "—"
            d_money  = mt["date"].strftime("%d/%m/%Y") if (mt and mt["date"]) else "—"
            data_diversa = (mt and bt["date"] != mt["date"]) if mt else False

            # Importi con segno reale (deposit positivo, withdrawal negativo)
            amt_banca_signed = bt["deposit"] - bt["withdrawal"]
            amt_money_signed  = mt["amount"] if mt else 0.0

            # Inversione di segno: banca e Money hanno segno opposto → situazione
            # strutturale (es. pagamento carta, giroconto) — importo NON modificabile
            segno_invertito = (mt and amt_banca_signed != 0 and amt_money_signed != 0
                               and (amt_banca_signed > 0) != (amt_money_signed > 0))

            # Differenza di importo (solo se segno concorde)
            imp_diverso = (mt and not segno_invertito
                           and abs(abs(amt_banca_signed) - abs(amt_money_signed)) > 0.005) if mt else False

            # Regola split: la data si può aggiornare; l'importo NON si tocca mai
            is_split = bool(mt and mt.get("has_splits"))
            # Blocco importo: split, segno invertito, o segno zero
            imp_bloccato = is_split or segno_invertito
            self._is_split.append(imp_bloccato)   # flag generico blocco importo
            self._data_diversa.append(data_diversa)
            self._imp_diverso.append(imp_diverso)

            # Default checkbox:
            # - Se importo è diverso (e non bloccato) → proponi aggiornamento importo
            # - Se importo è uguale ma data è diversa → proponi aggiornamento data
            # - Se entrambi diversi → importo ha priorità (più critico per le carte)
            if imp_diverso and not imp_bloccato:
                v_data = tk.BooleanVar(value=False)
                v_imp  = tk.BooleanVar(value=True)
            else:
                v_data = tk.BooleanVar(value=data_diversa)
                v_imp  = tk.BooleanVar(value=False)
            self._chk_data.append(v_data)
            self._chk_importo.append(v_imp)

            # Simbolo cella iniziale coerente con i BooleanVar appena impostati
            chk_d = "☑" if v_data.get() else ("—" if not data_diversa else "☐")
            if imp_bloccato:
                chk_i = "🔒" if is_split else "⚠️"
            elif not imp_diverso:
                chk_i = "—"
            else:
                chk_i = "☑" if v_imp.get() else "☐"

            # Mostra importo Money con segno reale (non abs)
            imp_money_str = (fmt_eur(amt_money_signed) if mt else "—")
            if is_split and imp_diverso:
                imp_money_str += " [split]"
            elif segno_invertito:
                imp_money_str += " ⚠️segno"

            iid = self.tree.insert("", "end", values=(
                chk_d, chk_i,
                d_banca, d_money,
                fmt_eur(bt["deposit"] - bt["withdrawal"]),
                imp_money_str,
                bt.get("descrizione", "")[:45],
                r.get("note", "")[:40]
            ))
            self._iids.append(iid)
            # Tag colore
            if is_split:
                self.tree.item(iid, tags=("split_fuzzy",))
            elif segno_invertito:
                self.tree.item(iid, tags=("segno_inv",))
            elif data_diversa and imp_diverso:
                self.tree.item(iid, tags=("entrambi",))
            elif data_diversa:
                self.tree.item(iid, tags=("data",))
            elif imp_diverso:
                self.tree.item(iid, tags=("importo",))

        self.tree.tag_configure("data",        background="#fff3cd")
        self.tree.tag_configure("importo",     background="#cce5ff")
        self.tree.tag_configure("entrambi",    background="#fce5b5")
        self.tree.tag_configure("split_fuzzy", background="#e8d5f5")  # viola tenue
        self.tree.tag_configure("segno_inv",   background="#ffd5d5")  # rosso tenue

        # Click su riga → toggle checkbox
        self.tree.bind("<ButtonRelease-1>", self._on_click)

        # Legenda + pulsanti selezione rapida
        leg = ttk.Frame(self)
        leg.pack(fill="x", padx=10, pady=(0, 2))
        ttk.Label(leg, text="Legenda:", font=("Helvetica", 9)).pack(side="left")
        for txt, bg in [(" Data sfasata ", "#fff3cd"),
                        (" Importo diverso ", "#cce5ff"),
                        (" Entrambi ", "#fce5b5"),
                        (" Split (importo bloccato) ", "#e8d5f5"),
                        (" Segno invertito (bloccato) ", "#ffd5d5")]:
            lbl = tk.Label(leg, text=txt, bg=bg, relief="solid", bd=1,
                           font=("Helvetica", 9))
            lbl.pack(side="left", padx=4)
        ttk.Separator(leg, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(leg, text="✓ Seleziona tutti (data)",
                   command=lambda: self._sel_tutti("data", True)).pack(side="left", padx=2)
        ttk.Button(leg, text="✗ Deseleziona tutti (data)",
                   command=lambda: self._sel_tutti("data", False)).pack(side="left", padx=2)
        ttk.Button(leg, text="✓ Tutti (importo)",
                   command=lambda: self._sel_tutti("importo", True)).pack(side="left", padx=2)
        ttk.Button(leg, text="✗ Nessuno (importo)",
                   command=lambda: self._sel_tutti("importo", False)).pack(side="left", padx=2)

        # Barra pulsanti
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=10, pady=(4, 8))
        ttk.Button(bar, text="🔧  Applica aggiornamenti",
                   command=self._applica).pack(side="left", padx=4)
        ttk.Button(bar, text="Annulla",
                   command=self.destroy).pack(side="right", padx=4)

    # ── Logica toggle checkbox via click ────────────────────────────────

    def _on_click(self, event):
        """Toggle checkbox data/importo in base alla colonna cliccata.
        - Click su colonna ▶Imp (#2) → toggle importo (se non bloccato)
        - Click su qualsiasi altra colonna della riga → toggle data
        Questo rende intuitivo deselezionare singole righe: basta cliccarle.
        """
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)   # es. "#1", "#2"
        iid = self.tree.identify_row(event.y)
        if not iid or iid not in self._iids:
            return
        i = self._iids.index(iid)

        if col == "#2":  # colonna ▶Imp — toggle importo se non bloccato
            if self._is_split[i] or not self._imp_diverso[i]:
                return   # bloccato: split, segno invertito, o importo già uguale
            self._chk_importo[i].set(not self._chk_importo[i].get())
        else:            # qualsiasi altra colonna → toggle data
            if not self._data_diversa[i]:
                return   # data già uguale: niente da aggiornare
            self._chk_data[i].set(not self._chk_data[i].get())

        self._aggiorna_riga_chk(i)

    def _aggiorna_riga_chk(self, i: int):
        """Aggiorna i simboli nella riga i della Treeview dopo un toggle."""
        iid = self._iids[i]
        vals = list(self.tree.item(iid, "values"))
        # Colonna data: ☑/☐ se la data è diversa, — altrimenti
        vals[0] = ("☑" if self._chk_data[i].get() else "☐") if self._data_diversa[i] else "—"
        # Colonna importo: 🔒 split, ⚠️ segno inv., — se uguale, ☑/☐ altrimenti
        if self._is_split[i] and not self._imp_diverso[i]:
            vals[1] = "🔒"
        elif self._is_split[i]:
            vals[1] = "🔒"
        elif not self._imp_diverso[i]:
            vals[1] = "—"
        else:
            vals[1] = "☑" if self._chk_importo[i].get() else "☐"
        self.tree.item(iid, values=vals)

    def _sel_tutti(self, campo: str, valore: bool):
        """Seleziona/deseleziona tutti i checkbox per data o importo."""
        for i, v in enumerate(self._chk_data if campo == "data" else self._chk_importo):
            if campo == "importo" and self._is_split[i]:
                continue   # importo bloccato
            if campo == "importo" and not self._imp_diverso[i]:
                continue   # importo non diverso: non applicabile
            if campo == "data" and not self._data_diversa[i]:
                continue   # data non diversa: non applicabile
            v.set(valore)
            self._aggiorna_riga_chk(i)

    # ── Applicazione aggiornamenti ───────────────────────────────────────

    def _applica(self):
        """Verifica Moneyspire chiuso, backup, scrittura aggiornamenti."""
        import subprocess

        n_data = sum(1 for v in self._chk_data    if v.get())
        n_imp  = sum(1 for v in self._chk_importo if v.get())
        n_tot  = sum(1 for i in range(len(self.results))
                     if self._chk_data[i].get() or self._chk_importo[i].get())

        if n_tot == 0:
            messagebox.showinfo("Nessuna selezione",
                                "Seleziona almeno una contabile da aggiornare.",
                                parent=self)
            return

        # Verifica Moneyspire non aperto
        r = subprocess.run(["pgrep", "-f", "Moneyspire"],
                           capture_output=True, text=True)
        if r.returncode == 0:
            messagebox.showerror(
                "Moneyspire aperto",
                "Chiudi Moneyspire prima di procedere.",
                parent=self)
            return

        msg = (f"Stai per applicare {n_tot} aggiornamenti su Moneyspire:\n"
               f"  - Data aggiornata:    {n_data}\n"
               f"  - Importo aggiornato: {n_imp}\n\n"
               f"Verrà creato un backup automatico del file .ffd.\n\n"
               f"Procedere?")
        if not messagebox.askyesno("Conferma aggiornamento", msg, parent=self):
            return

        # Backup + copia di lavoro
        try:
            work_path, backup_path = prepara_db_scrittura(self.ffd_path, _PROFILO)
        except Exception as e:
            messagebox.showerror("Errore backup", str(e), parent=self)
            return

        self.app.log(f"  [Fuzzy] Backup: {Path(backup_path).name}")
        self.app.log(f"  [Fuzzy] Copia di lavoro: {Path(work_path).name}")

        # Scrittura
        try:
            writer = MoneyWriter(work_path)
            ok = err = 0
            righe_audit: list[dict] = []

            for i, r in enumerate(self.results):
                mt = r["money_txns"][0] if r["money_txns"] else None
                if not mt:
                    continue
                txn_id = mt["id"]
                bt = r["bank_txn"]

                agg_data    = self._chk_data[i].get()
                agg_importo = self._chk_importo[i].get()

                if not agg_data and not agg_importo:
                    continue

                # Sicurezza: non toccare mai l'importo di una transazione split
                if agg_importo and self._is_split[i]:
                    mt_amount = mt["amount"]
                    bt_signed = bt["deposit"] - bt["withdrawal"]
                    segno_inv = (mt_amount != 0 and bt_signed != 0
                                 and (bt_signed > 0) != (mt_amount > 0))
                    motivo = "segno invertito" if segno_inv else "split"
                    self.app.log(
                        f"  [Fuzzy] ⚠️ Importo BLOCCATO ({motivo}): "
                        f"{bt.get('descrizione','')[:40]} — correggere manualmente in Moneyspire")
                    righe_audit.append({
                        "tipo": "WARN", "txn_id": txn_id,
                        "data": mt["date"], "importo": mt["amount"],
                        "categoria": mt.get("category", ""), "payee": mt.get("payee", ""),
                        "memo": mt.get("memo", ""),
                        "note": "importo bloccato — split, correggere in Moneyspire"
                    })
                    agg_importo = False

                try:
                    note_parts = []
                    if agg_data and bt["date"] != mt["date"]:
                        writer.correggi_data(txn_id, bt["date"])
                        self.app.log(
                            f"  [Fuzzy] Data  {mt['date']} → {bt['date']}  "
                            f"{bt.get('descrizione','')[:35]}"
                            + (" [split]" if self._is_split[i] else ""))
                        note_parts.append(f"data: {mt['date']} → {bt['date']}")

                    if agg_importo:
                        amt_banca = bt["deposit"] - bt["withdrawal"]
                        if amt_banca >= 0:
                            writer.correggi_importo(txn_id, deposit=amt_banca)
                        else:
                            writer.correggi_importo(txn_id, withdrawal=abs(amt_banca))
                        self.app.log(
                            f"  [Fuzzy] Imp   {fmt_eur(mt['amount'])} → "
                            f"{fmt_eur(amt_banca)}  {bt.get('descrizione','')[:35]}")
                        note_parts.append(f"imp: {fmt_eur(mt['amount'])} → {fmt_eur(amt_banca)}")

                    tipo_audit = "UPD_DATA" if (agg_data and not agg_importo) else \
                                 "UPD_IMP"  if (agg_importo and not agg_data) else "UPD_ENTRAMBI"
                    righe_audit.append({
                        "tipo": tipo_audit, "txn_id": txn_id,
                        "data": bt["date"], "importo": bt["deposit"] - bt["withdrawal"],
                        "categoria": mt.get("category", ""), "payee": mt.get("payee", ""),
                        "memo": mt.get("memo", "") or bt.get("descrizione", ""),
                        "note": "; ".join(note_parts) + (" [split]" if self._is_split[i] else "")
                    })
                    ok += 1
                except Exception as e:
                    err += 1
                    self.app.log(f"  [Fuzzy] ❌ {bt.get('descrizione','')[:35]}: {e}")
                    righe_audit.append({
                        "tipo": "ERR", "txn_id": txn_id,
                        "data": mt["date"], "importo": mt["amount"],
                        "categoria": mt.get("category", ""), "payee": mt.get("payee", ""),
                        "memo": mt.get("memo", ""), "note": str(e)[:60]
                    })

            writer.close()

        except Exception as e:
            messagebox.showerror("Errore scrittura", str(e), parent=self)
            import os as _os; _os.unlink(work_path)
            return

        # Verifica e sostituzione
        from datetime import date as _date
        ok_ver, msg_ver = verifica_scrittura(
            work_path, self.account_id,
            _date(1900, 1, 1), _date(2099, 12, 31), 0)

        self.app.log(f"  [Fuzzy] {msg_ver}")
        self.app.log(f"  [Fuzzy] OK={ok}  Errori={err}")

        msg_ris = (
            f"Aggiornamenti applicati: {ok}\n"
            f"Errori: {err}\n\n"
            f"Verifica: {msg_ver}\n\n"
            f"Backup: {Path(backup_path).name}\n\n"
            "Sostituire il file Moneyspire con la copia aggiornata?")
        if err == 0:
            if messagebox.askyesno("Conferma sostituzione", msg_ris, parent=self):
                finalizza_db(self.ffd_path, work_path)
                self.app.log("  [Fuzzy] ✅ File .ffd sostituito")
                # ── Audit log ────────────────────────────────────────────
                conto_nome = self.app.cfg.get("last_paths", {}).get("conto", "?")
                riepilogo_audit = (f"{ok} aggiornamenti  {err} errori  "
                                   f"— Verifica: {msg_ver}")
                write_audit_log(
                    operazione="AGGIORNA_FUZZY",
                    ffd_path=self.ffd_path,
                    backup_path=backup_path,
                    conto_nome=conto_nome,
                    account_id=self.account_id,
                    righe=righe_audit,
                    riepilogo=riepilogo_audit)
                self.app.log(f"  [Fuzzy] 📋 Audit log: {_audit_log_path().name}")
                # ─────────────────────────────────────────────────────────
                messagebox.showinfo("Completato",
                    f"Aggiornamento completato.\n"
                    f"Backup conservato: {Path(backup_path).name}\n\n"
                    "Ora puoi aprire Moneyspire per verificare.",
                    parent=self)
                self._scrittura_ok = True
                self.destroy()
            else:
                import os as _os; _os.unlink(work_path)
                self.app.log("  [Fuzzy] ⚠️ Sostituzione annullata — copia eliminata")
        else:
            import os as _os; _os.unlink(work_path)
            messagebox.showerror("Errori",
                f"Ci sono stati {err} errori.\n"
                f"Originale non modificato. Copia eliminata.",
                parent=self)


# ══════════════════════════════════════════════════════════════════════
# DIALOG MODIFICA CATEGORIA / PAYEE / MEMO
#
# Aperto da DialogRevisione con doppio click su una riga.
# Funzionalità:
#   - Filtro live sulla listbox categorie
#   - Payee → categoria automatica (da LastCategoryID in Money)
#   - Copia descrizione completa Fineco nel campo Memo
#   - Checkbox Ignora (esclude la transazione dall'inserimento)
# ══════════════════════════════════════════════════════════════════════

class DialogModificaCategoria(tk.Toplevel):
    """
    Dialogo per modificare categoria, payee, memo di una transazione.
    Funzionalità:
    - Filtro live sulla categoria (scrivi lettere → filtra in tempo reale)
    - Campo Payee con autocompletamento dai payee esistenti
    - Bottone "Copia da banca" per copiare la descrizione completa nel memo
    - Checkbox Ignora
    """

    def __init__(self, parent, txn: dict, cats: list[str], on_save,
                 payees: list[str] | None = None,
                 payee_cats: dict | None = None,
                 cats_by_id: dict | None = None,
                 contropartita_default: str = "Da classificare"):
        super().__init__(parent)
        self.title("Modifica transazione")
        self.on_save     = on_save
        self._txn        = txn
        self._cats_all   = cats          # lista completa categorie
        self._payees_all = payees or []
        self._desc2      = txn.get("_descrizione_completa","") or ""
        # Mappa payee_name → LastCategoryID e ID → nome per auto-categoria
        self._payee_cats = payee_cats or {}   # dict[str, int]
        self._cats_by_id = cats_by_id  or {}  # dict[int, str]
        self._cp_default = contropartita_default
        self.resizable(True, False)

        frm = ttk.Frame(self, padding=16)
        frm.pack(fill="both", expand=True)
        frm.columnconfigure(1, weight=1)

        # ── Info transazione (sola lettura) ───────────────────────────
        amt = txn["deposit"] - txn["withdrawal"]
        for r, lbl, val, bold in [
            (0, "Data:",        txn["txn_date"].strftime("%d/%m/%Y"), False),
            (1, "Importo:",     fmt_eur(amt),                          True),
            (2, "Descrizione:", txn.get("_descrizione_banca",""),      False),
        ]:
            ttk.Label(frm, text=lbl, anchor="e", width=14
                      ).grid(row=r, column=0, sticky="e", pady=2)
            ttk.Label(frm, text=val, wraplength=420,
                      font=("Helvetica", 11, "bold") if bold else ("Helvetica", 10)
                      ).grid(row=r, column=1, sticky="w", padx=8)

        # Descrizione completa (sola lettura, selezionabile per copia)
        ttk.Label(frm, text="Descr. completa:", anchor="e", width=14
                  ).grid(row=3, column=0, sticky="ne", pady=3)
        self._txt_desc2 = tk.Text(frm, height=3, wrap="word",
                                   font=("Helvetica", 10), bg="#f0f0f0")
        self._txt_desc2.insert("1.0", self._desc2)
        self._txt_desc2.configure(state="disabled")
        self._txt_desc2.grid(row=3, column=1, sticky="ew", padx=8, pady=3)

        ttk.Separator(frm).grid(row=4, column=0, columnspan=2,
                                sticky="ew", pady=8)

        # ── Categoria con ricerca + Listbox ──────────────────────────────
        ttk.Label(frm, text="Cerca categoria:", anchor="e", width=14
                  ).grid(row=5, column=0, sticky="ne", pady=3)

        cat_frame = ttk.Frame(frm)
        cat_frame.grid(row=5, column=1, sticky="ew", padx=8, pady=3)
        cat_frame.columnconfigure(0, weight=1)

        # Campo ricerca libero — non ruba mai il focus alla lista
        self._var_filtro = tk.StringVar()
        filtro_entry = ttk.Entry(cat_frame, textvariable=self._var_filtro, width=44)
        filtro_entry.grid(row=0, column=0, sticky="ew")
        self._var_filtro.trace_add("write", self._aggiorna_filtro)

        # Listbox con scrollbar — mostra i risultati del filtro
        lb_frame = ttk.Frame(cat_frame)
        lb_frame.grid(row=1, column=0, sticky="ew", pady=(2,0))
        lb_frame.columnconfigure(0, weight=1)

        self._lb_cat = tk.Listbox(lb_frame, height=6, font=("Helvetica", 10),
                                   selectmode="single", exportselection=False)
        lb_vsb = ttk.Scrollbar(lb_frame, orient="vertical",
                                command=self._lb_cat.yview)
        self._lb_cat.configure(yscrollcommand=lb_vsb.set)
        self._lb_cat.grid(row=0, column=0, sticky="ew")
        lb_vsb.grid(row=0, column=1, sticky="ns")

        # Campo categoria selezionata (sola lettura, mostra la scelta corrente)
        ttk.Label(frm, text="Categoria:", anchor="e", width=14
                  ).grid(row=6, column=0, sticky="e", pady=3)
        self._var_cat = tk.StringVar(
            value=txn.get("_categoria_suggerita","") or "")
        lbl_cat_sel = ttk.Label(frm, textvariable=self._var_cat,
                                 font=("Helvetica", 10, "bold"),
                                 foreground="#1a5276")
        lbl_cat_sel.grid(row=6, column=1, sticky="w", padx=8)

        # Popola la Listbox inizialmente
        self._popola_listbox(self._cats_all)
        # Preselezione se c'è già una categoria
        cat_corrente = self._var_cat.get()
        if cat_corrente:
            self._var_filtro.set(cat_corrente.split(":")[-1])  # pre-filtra sull'ultimo segmento

        # Binding selezione
        self._lb_cat.bind("<<ListboxSelect>>", self._on_lb_select)
        self._lb_cat.bind("<Return>",          self._on_lb_select)

        # ── Payee ─────────────────────────────────────────────────────
        ttk.Label(frm, text="Payee:", anchor="e", width=14
                  ).grid(row=7, column=0, sticky="e", pady=3)
        self._var_payee = tk.StringVar(value=txn.get("_payee_suggerito","") or "")
        self._cb_payee = ttk.Combobox(frm, textvariable=self._var_payee,
                                       values=self._payees_all, width=42)
        self._cb_payee.grid(row=7, column=1, sticky="w", padx=8)
        # Quando si seleziona un payee, propone automaticamente la sua categoria default
        self._var_payee.trace_add("write", self._on_payee_changed)

        # ── Memo + bottone Copia da banca ─────────────────────────────
        ttk.Label(frm, text="Memo:", anchor="e", width=14
                  ).grid(row=8, column=0, sticky="e", pady=3)
        memo_frame = ttk.Frame(frm)
        memo_frame.grid(row=8, column=1, sticky="ew", padx=8)
        self._var_memo = tk.StringVar(value=txn.get("memo","") or "")
        ttk.Entry(memo_frame, textvariable=self._var_memo, width=36
                  ).pack(side="left")
        ttk.Button(memo_frame, text="📋 Copia da banca", width=16,
                   command=self._copia_da_banca).pack(side="left", padx=6)

        # ── Ignora ────────────────────────────────────────────────────
        self._var_ignora = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm,
                        text="Ignora questa transazione (non inserire in Money)",
                        variable=self._var_ignora
                        ).grid(row=9, column=1, sticky="w", padx=8, pady=4)

        # ── Pulsanti ──────────────────────────────────────────────────
        btn_frm = ttk.Frame(frm)
        btn_frm.grid(row=10, column=0, columnspan=2, pady=12)
        ttk.Button(btn_frm, text="Salva",   command=self._salva,
                   width=12).pack(side="left", padx=6)
        ttk.Button(btn_frm, text="Annulla", command=self.destroy,
                   width=10).pack(side="left", padx=6)

        filtro_entry.focus_set()
        self.transient(parent)
        self.grab_set()

    # ── Filtro + Listbox ──────────────────────────────────────────────

    def _popola_listbox(self, items: list[str]):
        self._lb_cat.delete(0, "end")
        for item in items:
            self._lb_cat.insert("end", item)

    def _aggiorna_filtro(self, *_):
        """Filtra la Listbox in base al testo — NON tocca il focus."""
        testo = self._var_filtro.get().lower().strip()
        if not testo:
            filtrate = self._cats_all
        else:
            parole   = testo.split()
            filtrate = [c for c in self._cats_all
                        if all(p in c.lower() for p in parole)]
        self._popola_listbox(filtrate)
        # Se una sola opzione, selezionala automaticamente
        if len(filtrate) == 1:
            self._lb_cat.selection_set(0)
            self._var_cat.set(filtrate[0])

    def _on_lb_select(self, _=None):
        """Aggiorna _var_cat con l'elemento selezionato nella Listbox."""
        sel = self._lb_cat.curselection()
        if sel:
            self._var_cat.set(self._lb_cat.get(sel[0]))

    def _aggiorna_filtro_da_cat(self, *_):
        pass   # non serve più con la Listbox

    # ── Payee → Categoria automatica ─────────────────────────────────

    def _on_payee_changed(self, *_):
        """Quando si sceglie un payee, pre-compila la categoria con LastCategoryID."""
        payee = self._var_payee.get().strip()
        if not payee:
            return
        cat_id = self._payee_cats.get(payee)
        if cat_id and cat_id in self._cats_by_id:
            cat_name = self._cats_by_id[cat_id]
            # Aggiorna solo se la categoria è ancora vuota o generica
            cat_attuale = self._var_cat.get()
            if not cat_attuale or cat_attuale in (self._cp_default, "", "—"):
                self._var_cat.set(cat_name)
                # Aggiorna anche il filtro per mostrare la categoria selezionata
                self._var_filtro.set(cat_name.split(":")[-1])

    # ── Copia da banca ────────────────────────────────────────────────

    def _copia_da_banca(self):
        """Copia la descrizione completa Fineco nel campo memo (troncata a 100 car)."""
        desc2 = self._desc2.strip()
        if desc2:
            self._var_memo.set(desc2[:100])

    # ── Salva ─────────────────────────────────────────────────────────

    def _salva(self):
        self.on_save(
            self._var_cat.get().strip(),
            self._var_memo.get().strip(),
            self._var_ignora.get(),
            self._var_payee.get().strip()
        )
        self.destroy()

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2: REGOLE
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# TAB 2 — REGOLE
#
# Gestione manuale di ms_rules.json:
#   aggiungi, modifica, elimina, sposta su/giù, impara da history Money.
# RuleDialog: form di inserimento/modifica singola regola.
# ══════════════════════════════════════════════════════════════════════

class TabRegole(ttk.Frame):
    def __init__(self, parent, app: "App"):
        super().__init__(parent)
        self.app = app
        self._build()

    def _build(self):
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=8, pady=6)
        ttk.Button(btn_bar, text="➕  Nuova",      command=self._add).pack(side="left", padx=4)
        ttk.Button(btn_bar, text="✏️  Modifica",   command=self._edit).pack(side="left", padx=4)
        ttk.Button(btn_bar, text="🗑️  Elimina",    command=self._del).pack(side="left", padx=4)
        ttk.Button(btn_bar, text="⬆  Sposta su",   command=self._up).pack(side="left", padx=4)
        ttk.Button(btn_bar, text="🔄  Ricarica",    command=self._reload).pack(side="left", padx=4)
        ttk.Button(btn_bar, text="🧠  Impara da Money",
                   command=self._learn).pack(side="left", padx=16)

        tbl = ttk.Frame(self)
        tbl.pack(fill="both", expand=True, padx=8, pady=4)
        cols = ("#", "Pattern", "Regex", "Categoria", "Payee", "Fonte", "Hits")
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings", selectmode="browse")
        for col, w in zip(cols, (40, 240, 50, 220, 160, 70, 50)):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="w" if w > 60 else "center")
        vsb = ttk.Scrollbar(tbl, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)
        self._reload()

    def _reload(self):
        if self.app.rules is None:
            return
        self.app.rules.load()
        self.tree.delete(*self.tree.get_children())
        for i, r in enumerate(self.app.rules.rules):
            self.tree.insert("", "end", values=(
                i+1, r.get("pattern",""), "✓" if r.get("regex") else "",
                r.get("category",""), r.get("payee",""),
                r.get("source",""), r.get("hits", 0)
            ))

    def _selected_idx(self) -> int | None:
        sel = self.tree.selection()
        return self.tree.index(sel[0]) if sel else None

    def _add(self):
        RuleDialog(self, self.app, rule=None, on_save=self._reload)

    def _edit(self):
        idx = self._selected_idx()
        if idx is not None:
            RuleDialog(self, self.app, rule=self.app.rules.rules[idx],
                       on_save=self._reload, edit_idx=idx)

    def _del(self):
        idx = self._selected_idx()
        if idx is None:
            return
        pat = self.app.rules.rules[idx].get("pattern", "")
        if messagebox.askyesno("Conferma", f"Eliminare la regola:\n\"{pat}\"?"):
            self.app.rules.remove(idx)
            self._reload()

    def _up(self):
        idx = self._selected_idx()
        if idx is not None:
            self.app.rules.move_up(idx)
            self._reload()

    def _learn(self):
        ffd = self.app.tab_ric.pv_ffd.get()
        if not ffd:
            ffd = filedialog.askopenfilename(
                title="Seleziona .ffd per apprendimento",
                filetypes=[("Moneyspire", "*.ffd")])
        if not ffd:
            return
        try:
            db = MoneyspireDB(ffd)
            # Prende tutte le transazioni del conto principale (ultimi 2 anni)
            acct = list(self.app.cfg.get("conti", {}).values())[0]["ffd_account_id"]
            from datetime import date
            txns = db.get_transactions(acct,
                                       date_from=date(date.today().year - 2, 1, 1))
            db.close()
            n = self.app.rules.learn_from_history(txns)
            self._reload()
            self.app.log(f"  Apprendimento: aggiunte {n} nuove regole automatiche")
            messagebox.showinfo("Apprendimento completato",
                                f"Aggiunte {n} nuove regole automatiche")
        except Exception as e:
            messagebox.showerror("Errore", str(e))


class RuleDialog(tk.Toplevel):
    def __init__(self, parent, app: "App", rule: dict | None,
                 on_save, edit_idx: int | None = None):
        super().__init__(parent)
        self.title("Nuova regola" if rule is None else "Modifica regola")
        self.app      = app
        self.on_save  = on_save
        self.edit_idx = edit_idx
        rule = rule or {}

        frm = ttk.Frame(self, padding=16)
        frm.pack(fill="both", expand=True)
        frm.columnconfigure(1, weight=1)

        self._var_pat   = tk.StringVar(value=rule.get("pattern", ""))
        self._var_regex = tk.BooleanVar(value=rule.get("regex", False))
        self._var_cat   = tk.StringVar(value=rule.get("category", ""))
        self._var_payee = tk.StringVar(value=rule.get("payee", ""))

        def row(r, lbl, widget):
            ttk.Label(frm, text=lbl, width=14, anchor="e"
                      ).grid(row=r, column=0, sticky="e", pady=4)
            widget.grid(row=r, column=1, sticky="ew", padx=8, pady=4)

        row(0, "Pattern:",   ttk.Entry(frm, textvariable=self._var_pat, width=42))
        row(1, "È regex:",   ttk.Checkbutton(frm, variable=self._var_regex))
        cats = sorted(set(app.db_cats.values())) if app.db_cats else []
        row(2, "Categoria:", ttk.Combobox(frm, textvariable=self._var_cat,
                                          values=cats, width=40))
        row(3, "Payee:",     ttk.Entry(frm, textvariable=self._var_payee, width=42))
        ttk.Button(frm, text="Salva", command=self._save
                   ).grid(row=4, column=0, columnspan=2, pady=12)
        self.transient(parent)
        self.grab_set()

    def _save(self):
        pat, cat = self._var_pat.get().strip(), self._var_cat.get().strip()
        if not pat or not cat:
            messagebox.showwarning("Attenzione",
                                   "Pattern e Categoria sono obbligatori", parent=self)
            return
        if self.edit_idx is not None:
            r = self.app.rules.rules[self.edit_idx]
            r.update({"pattern": pat, "regex": self._var_regex.get(),
                      "category": cat, "payee": self._var_payee.get().strip(),
                      "source": "manual"})
            self.app.rules.save()
        else:
            self.app.rules.add_or_update(
                pat, cat, self._var_payee.get().strip(),
                self._var_regex.get(), source="manual")
        self.on_save()
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3: LOG
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# TAB 3 — LOG
#
# Log testuale timestampato di tutte le operazioni della sessione.
# Copia negli appunti o salva su file.
# Riceve messaggi tramite App.log() → TabLog.append().
# ══════════════════════════════════════════════════════════════════════

class TabLog(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self._build()

    def _build(self):
        self.txt = tk.Text(self, wrap="word", font=("Menlo", 10),
                           state="disabled", bg="#1e1e1e", fg="#d4d4d4",
                           insertbackground="white")
        vsb = ttk.Scrollbar(self, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.txt.pack(fill="both", expand=True, padx=4, pady=4)

    def append(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.txt.configure(state="normal")
        self.txt.insert("end", f"{ts}  {msg}\n")
        self.txt.see("end")
        self.txt.configure(state="disabled")

    def _clear(self):
        self.txt.configure(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.configure(state="disabled")

    def _copy(self):
        """Copia il log negli appunti — usa after() per stabilità su macOS."""
        content = self.txt.get("1.0", "end-1c")
        self.txt.clipboard_clear()
        self.txt.clipboard_append(content)
        self.txt.update()
        # Conferma visiva nella status bar (se raggiungibile)
        try:
            self.winfo_toplevel().status.config(text="Log copiato negli appunti")
        except Exception:
            pass

    def _save_file(self):
        """Salva il log su file .txt scelto dall'utente."""
        content = self.txt.get("1.0", "end-1c")
        if not content.strip():
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        p = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Testo", "*.txt"), ("Tutti", "*.*")],
            initialfile=f"ms_reconciler_log_{ts}.txt")
        if p:
            Path(p).write_text(content, encoding="utf-8")
            try:
                self.winfo_toplevel().status.config(text=f"Log salvato: {p}")
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4: FASE 2 — INTEGRAZIONE FILE EXCEL ELABORATI
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# TAB 4 — FASE 2 EXCEL
#
# Mantiene aggiornati i file Excel elaborati annuali:
#   2026.xlsx       → fogli Movimenti, Lombard, USD (con saldo a catena)
#   2026 mc.xlsx    → fogli mensili MC Fineco
#   2026 visa.xlsx  → fogli mensili Visa (Ignazio + Silvia)
# Flusso: confronto → revisione mancanti → backup → scrittura → sostituzione.
# ══════════════════════════════════════════════════════════════════════

class TabFase2(ttk.Frame):
    """
    Confronta file originali Fineco con i file elaborati annuali (2026.xlsx,
    2026 mc.xlsx, 2026 visa.xlsx). Identifica transazioni mancanti e
    aggiornamenti Moneymap possibili.
    """

    def __init__(self, parent, app: "App"):
        super().__init__(parent)
        self.app = app
        self._risultato: dict = {}
        self._build()

    def _build(self):
        # ── File ─────────────────────────────────────────────────────────
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x", padx=8, pady=6)
        top.columnconfigure(1, weight=1)

        def file_row(row_n, label, pathvar, filetypes, persist_key=""):
            ttk.Label(top, text=label, anchor="w", width=32
                      ).grid(row=row_n, column=0, sticky="w", pady=3)
            pe = PathEntry(top, pathvar) if _HAS_PATH_WIDGETS else ttk.Entry(
                top, textvariable=pathvar, state="readonly")
            pe.grid(row=row_n, column=1, sticky="ew", padx=4)
            def browse(pv=pathvar, ft=filetypes, lk=persist_key):
                current = pv.get()
                if current and Path(current).expanduser().parent.exists():
                    start_dir = str(Path(current).expanduser().parent)
                else:
                    start_dir = str(Path.home())
                p = filedialog.askopenfilename(filetypes=ft, initialdir=start_dir)
                if p:
                    pv.set(p)
                    self.app.cfg.setdefault("last_paths", {})[lk] = p
                    save_config(self.app.cfg)
            ttk.Button(top, text="📂", width=3, command=browse
                       ).grid(row=row_n, column=2)

        last = self.app.cfg.get("last_paths", {})
        # Chiavi dedicate Fase 2 — indipendenti dalla Fase 1
        def _norm2(p): return _norm_path(p) if p else ""
        self.pv_orig   = PathVar(value=_norm2(last.get("fase2_orig", "")))
        self.pv_elab   = PathVar(value=_norm2(last.get("xlsx_elaborato", "")))

        file_row(0, "File originale Fineco (movements_*.xlsx):",
                 self.pv_orig, [("Excel", "*.xlsx")], "fase2_orig")
        file_row(1, "File elaborato annuale (2026.xlsx / mc / visa):",
                 self.pv_elab, [("Excel", "*.xlsx")], "xlsx_elaborato")

        # ── Controlli ────────────────────────────────────────────────────
        ctrl = ttk.Frame(top)
        ctrl.grid(row=2, column=0, columnspan=3, pady=8, sticky="w")

        ttk.Label(ctrl, text="Conto:").pack(side="left")
        conti = [
            nome for nome, cfg in self.app.cfg.get("conti", {}).items()
            if cfg.get("includi_in_riconciliazione", True)
        ]
        # Chiave "fase2_conto" — indipendente da "conto" della Fase 1
        self.var_conto = tk.StringVar(value=last.get("fase2_conto", conti[0] if conti else ""))
        self._cb_conto = ttk.Combobox(ctrl, textvariable=self.var_conto,
                     values=conti, state="readonly", width=20
                     )
        self._cb_conto.pack(side="left", padx=6)

        ttk.Label(ctrl, text="Mese:").pack(side="left", padx=(12, 2))
        now = date.today()
        self.var_mese = tk.IntVar(value=now.month - 1 or 12)
        ttk.Spinbox(ctrl, from_=1, to=12, textvariable=self.var_mese,
                    width=4).pack(side="left")

        ttk.Label(ctrl, text="Anno:").pack(side="left", padx=(8, 2))
        self.var_anno = tk.IntVar(value=now.year)
        ttk.Spinbox(ctrl, from_=2020, to=2035, textvariable=self.var_anno,
                    width=6).pack(side="left")

        ttk.Button(ctrl, text="🔍  Confronta",
                   command=self._confronta).pack(side="left", padx=12)
        self.btn_scrivi = ttk.Button(ctrl, text="✅  Scrivi nel file Excel",
                   command=self._scrivi, state="disabled")
        self.btn_scrivi.pack(side="left", padx=4)
        ttk.Button(ctrl, text="📊  Riepilogo mensile",
                   command=self._riepilogo).pack(side="left", padx=4)

        # ── Sommario ─────────────────────────────────────────────────────
        self.lbl_sum = ttk.Label(self, text="", font=("Helvetica", 11))
        self.lbl_sum.pack(padx=8, pady=2, anchor="w")

        # ── Notebook interno: Mancanti / Solo elaborato / Moneymap ───────
        nb2 = ttk.Notebook(self)
        nb2.pack(fill="both", expand=True, padx=8, pady=4)

        self._tree_miss  = self._make_tree(nb2, ("Data", "Importo", "Descrizione",
                                                  "Descrizione completa"),
                                           "Mancanti nel file elaborato")
        self._tree_extra = self._make_tree(nb2, ("Data", "Importo", "Descrizione",
                                                  "Moneymap attuale"),
                                           "Solo nel file elaborato")
        self._tree_mmap  = self._make_tree(nb2, ("Riga", "Data", "Importo",
                                                  "Descrizione", "Moneymap attuale",
                                                  "Moneymap nuova"),
                                           "Moneymap aggiornabile")

    def _make_tree(self, parent_nb, cols, tab_label):
        frm = ttk.Frame(parent_nb)
        parent_nb.add(frm, text=f"  {tab_label}  ")
        tree = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        widths = {"Data": 88, "Importo": 90, "Riga": 55,
                  "Descrizione": 260, "Descrizione completa": 300,
                  "Moneymap attuale": 180, "Moneymap nuova": 180}
        _sort_state: dict = {}
        for col in cols:
            tree.heading(col, text=col,
                         command=lambda c=col, t=tree, s=_sort_state: treeview_sort(
                             t, c, s,
                             date_cols=("Data",), number_cols=("Importo",)))
            tree.column(col, width=widths.get(col, 120),
                        anchor="center" if col in ("Data","Importo","Riga") else "w")
        vsb = ttk.Scrollbar(frm, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        return tree

    def _aggiorna_conti(self):
        """Ripopola il combobox conti dopo la scelta del profilo."""
        conti = [
            nome for nome, cfg in self.app.cfg.get("conti", {}).items()
            if cfg.get("includi_in_riconciliazione", True)
        ]
        last = self.app.cfg.get("last_paths", {})
        self._cb_conto["values"] = conti
        val = last.get("fase2_conto", conti[0] if conti else "")
        self.var_conto.set(val if val in conti else (conti[0] if conti else ""))
        def _norm(p): return _norm_path(p) if p else ""
        if last.get("fase2_orig"):      self.pv_orig.set(_norm(last["fase2_orig"]))
        if last.get("xlsx_elaborato"):  self.pv_elab.set(_norm(last["xlsx_elaborato"]))

    def _confronta(self):
        orig_path  = self.pv_orig.get()
        elab_path  = self.pv_elab.get()
        conto_name = self.var_conto.get()
        mese       = self.var_mese.get()
        anno       = self.var_anno.get()

        if not orig_path or not os.path.exists(orig_path):
            messagebox.showerror("Errore", "Seleziona il file originale Fineco"); return
        if not elab_path or not os.path.exists(elab_path):
            messagebox.showerror("Errore", "Seleziona il file elaborato annuale"); return

        conto_cfg = self.app.cfg.get("conti", {}).get(conto_name, {})
        tipo = conto_cfg.get("tipo", "conto_corrente")

        self.app.log(f"▶ Fase 2: confronto {conto_name} — {MESI_IT_NOMI[mese-1].capitalize()} {anno}")

        try:
            # Leggi transazioni originali banca
            tipo_file_analisi = conto_cfg.get("tipo_file_analisi", "originale_fineco")
            if tipo == "carta_credito":
                num_carta = conto_cfg.get("numeri_carta", conto_cfg.get("numero_carta"))
                bank_txns = parse_fineco_cc(orig_path, num_carta, mese, anno)
                tipo_elab = "cc_mensile"
            elif tipo_file_analisi == "originale_unicredit":
                # Unicredit non ha file elaborato Fineco — Fase 2 non applicabile
                messagebox.showinfo(
                    "Fase 2 non disponibile",
                    "Il conto Unicredit CCM non ha un file elaborato associato.\n"
                    "La Fase 2 (integrazione Excel) è disponibile solo per i conti Fineco."
                )
                return
            elif conto_cfg.get("valuta") == "USD":
                # Conto USD: parser dedicato (header a riga 7, date gg/mm/yyyy)
                from ms_engine import parse_fineco_conto_usd
                bank_txns = parse_fineco_conto_usd(orig_path, mese, anno)
                tipo_elab = "cc"
            else:
                bank_txns = parse_fineco_conto_originale(orig_path, mese, anno)
                tipo_elab = "cc"

            self.app.log(f"  Originale: {len(bank_txns)} transazioni")

            # Individua il foglio nel file elaborato
            from openpyxl import load_workbook as _lwb
            wb_check = _lwb(elab_path, read_only=True)
            fogli_elab = wb_check.sheetnames
            wb_check.close()

            integr = ExcelIntegrator(elab_path, tipo_elab)

            if tipo_elab == "cc_mensile":
                # File MC o Visa: foglio mensile
                # Costruisci sheet_map da fogli disponibili
                sheet_map = {}
                for i, nome in enumerate(MESI_IT_NOMI):
                    for foglio in fogli_elab:
                        if foglio.lower().strip() == nome:
                            sheet_map[i+1] = foglio
                            break
                self._risultato = integr.confronta_cc_mensile(
                    bank_txns, mese, anno, sheet_map)
            else:
                # File CC: foglio con nome da config (es. "Movimenti", "Lombard", "USD")
                sheet_name = conto_cfg.get("excel_sheet", "Movimenti")
                if sheet_name not in fogli_elab:
                    sheet_name = fogli_elab[0]
                self._risultato = integr.confronta_cc(
                    bank_txns, sheet_name, mese, anno)

            self._aggiorna_tabelle()
            n_miss  = len(self._risultato.get("presenti_solo_originale", []))
            n_extra = len(self._risultato.get("presenti_solo_elaborato", []))
            n_mmap  = len(self._risultato.get("moneymap_aggiornabile", []))
            n_abb   = self._risultato.get("n_abbinati", 0)
            self.lbl_sum.config(
                text=f"Abbinate: {n_abb}   "
                     f"❌ Mancanti nel elaborato: {n_miss}   "
                     f"⚠️ Solo elaborato: {n_extra}   "
                     f"🗂 Moneymap aggiornabile: {n_mmap}")
            self.app.log(f"  Abbinate: {n_abb}  Mancanti: {n_miss}  "
                         f"Solo elaborato: {n_extra}  Moneymap: {n_mmap}")
            # Salva path e conto Fase 2 nelle last_paths (chiavi dedicate)
            lp = self.app.cfg.setdefault("last_paths", {})
            lp["xlsx_elaborato"] = elab_path
            lp["fase2_conto"]    = conto_name
            save_config(self.app.cfg)
            # Abilita scrittura solo se ci sono mancanti e il file è CC (non mensile)
            self._sheet_name = sheet_name if tipo_elab == "cc" else None
            self._tipo_elab  = tipo_elab
            self._mese_elab  = mese
            self.btn_scrivi.config(
                state="normal" if n_miss > 0 else "disabled")

        except Exception as e:
            import traceback
            self.app.log(f"  ERRORE Fase 2: {e}")
            self.app.log(traceback.format_exc())
            messagebox.showerror("Errore Fase 2", str(e))

    def _scrivi(self):
        """Backup + scrivi transazioni mancanti nel file elaborato + verifica + finalizza."""
        mancanti   = self._risultato.get("presenti_solo_originale", [])
        if not mancanti:
            messagebox.showinfo("Nessun dato", "Nessuna transazione mancante da inserire")
            return

        elab_path  = self.pv_elab.get()
        tipo_elab  = getattr(self, "_tipo_elab", "cc")
        sheet_name = getattr(self, "_sheet_name", "Movimenti") or "Movimenti"
        mese       = getattr(self, "_mese_elab", 1)
        n_miss     = len(mancanti)

        if tipo_elab == "cc_mensile":
            foglio_desc = MESI_IT_NOMI[mese-1].capitalize()
        else:
            foglio_desc = sheet_name

        msg = (f"Stai per inserire {n_miss} transazioni nel file:\n"
               f"{Path(elab_path).name}\n"
               f"Foglio: {foglio_desc}\n\n"
               f"Verrà creato un backup automatico del file Excel.\n\n"
               f"Procedere?")
        if not messagebox.askyesno("Conferma scrittura Excel", msg, parent=self):
            return

        try:
            integr = ExcelIntegrator(elab_path, tipo_elab)
            work_path, backup_path = integr.prepara_excel_scrittura(_PROFILO)
        except Exception as e:
            messagebox.showerror("Errore backup", str(e), parent=self)
            return

        self.app.log(f"  Backup: {Path(backup_path).name}")
        self.app.log(f"  Copia di lavoro: {Path(work_path).name}")

        try:
            if tipo_elab == "cc_mensile":
                n_ins, msg_ins = integr.scrivi_transazioni_cc_mensile(
                    work_path, mancanti, mese)
            else:
                n_ins, msg_ins = integr.scrivi_transazioni_cc(
                    work_path, mancanti, sheet_name, rules=self.app.rules)
            self.app.log(f"  Scrittura: {msg_ins}")
        except Exception as e:
            import traceback, os as _os
            self.app.log(f"  ❌ Errore scrittura: {e}")
            self.app.log(traceback.format_exc())
            _os.unlink(work_path)
            messagebox.showerror("Errore scrittura", str(e), parent=self)
            return

        # Verifica: rileggi il file di lavoro e conta le nuove righe
        try:
            from openpyxl import load_workbook as _lwb
            wb_v = _lwb(work_path, read_only=True, data_only=True)
            foglio_ver = foglio_desc if tipo_elab == "cc_mensile" else sheet_name
            # Cerca il foglio case-insensitive
            found_v = next((s for s in wb_v.sheetnames
                            if s.lower() == foglio_ver.lower()), None)
            ws_v = wb_v[found_v] if found_v else wb_v.active
            col_check = 1  # colonna A per CC, colonna data_op per mensile
            n_dopo = sum(1 for r in ws_v.iter_rows(min_row=2, values_only=True)
                         if r[col_check-1] is not None)
            wb_v.close()
            ok_ver = n_ins > 0
            msg_ver = f"Verifica OK: {n_dopo} righe nel foglio '{found_v}'" if ok_ver else "Verifica FALLITA"
        except Exception as e:
            ok_ver, msg_ver = False, f"Errore verifica: {e}"

        self.app.log(f"  {msg_ver}")

        msg_ris = (f"Inserite: {n_ins}\n\n"
                   f"{msg_ver}\n\n"
                   f"Backup: {Path(backup_path).name}\n\n"
                   f"Sostituire il file originale con la versione aggiornata?")

        import os as _os
        if ok_ver and messagebox.askyesno("Conferma sostituzione", msg_ris, parent=self):
            integr.finalizza_excel(elab_path, work_path)
            self.app.log(f"  ✅ File Excel aggiornato — backup conservato")
            messagebox.showinfo("Completato",
                f"File aggiornato: {Path(elab_path).name}\n"
                f"Inserite {n_ins} transazioni nel foglio '{foglio_desc}'.\n"
                f"Backup: {Path(backup_path).name}",
                parent=self)
            self.btn_scrivi.config(state="disabled")
            self._confronta()
        else:
            _os.unlink(work_path)
            self.app.log("  ⚠️ Sostituzione annullata — copia di lavoro eliminata")

    def _aggiorna_tabelle(self):
        # Mancanti
        self._tree_miss.delete(*self._tree_miss.get_children())
        for bt in self._risultato.get("presenti_solo_originale", []):
            self._tree_miss.insert("", "end", values=(
                bt["date"].strftime("%d/%m/%Y"),
                fmt_eur(bt["amount"]),
                bt.get("descrizione", "")[:50],
                bt.get("descrizione_completa", "")[:80],
            ))
        # Solo elaborato
        self._tree_extra.delete(*self._tree_extra.get_children())
        for et in self._risultato.get("presenti_solo_elaborato", []):
            self._tree_extra.insert("", "end", values=(
                et["date"].strftime("%d/%m/%Y"),
                fmt_eur(et["amount"]),
                et.get("descrizione", "")[:50],
                et.get("moneymap", ""),
            ))
        # Moneymap aggiornabile
        self._tree_mmap.delete(*self._tree_mmap.get_children())
        for ag in self._risultato.get("moneymap_aggiornabile", []):
            self._tree_mmap.insert("", "end", values=(
                ag.get("row_idx", ""),
                ag["date"].strftime("%d/%m/%Y"),
                fmt_eur(ag["amount"]),
                ag.get("descrizione", "")[:50],
                ag.get("moneymap_attuale", ""),
                ag.get("moneymap_nuova", ""),
            ))

    def _esporta(self):
        mancanti = self._risultato.get("presenti_solo_originale", [])
        if not mancanti:
            messagebox.showinfo("Nessun dato", "Nessuna transazione mancante da esportare")
            return
        p = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialfile="fase2_mancanti_elaborato.csv")
        if not p:
            return
        import csv as _csv
        with open(p, "w", newline="", encoding="utf-8-sig") as f:
            w = _csv.DictWriter(f, fieldnames=[
                "Data", "Importo", "Descrizione", "Descrizione_Completa"], delimiter=";")
            w.writeheader()
            for bt in mancanti:
                w.writerow({
                    "Data":                 bt["date"].strftime("%d/%m/%Y"),
                    "Importo":              f"{bt['amount']:.2f}".replace(".", ","),
                    "Descrizione":          bt.get("descrizione", ""),
                    "Descrizione_Completa": bt.get("descrizione_completa", ""),
                })
        messagebox.showinfo("Esportato", f"{len(mancanti)} righe → {p}")
        self.app.log(f"  Fase 2 export mancanti: {len(mancanti)} righe → {p}")

    def _riepilogo(self):
        elab_path = self.pv_elab.get()
        if not elab_path or not os.path.exists(elab_path):
            messagebox.showerror("Errore", "Seleziona il file elaborato annuale"); return
        conto_name = self.var_conto.get()
        conto_cfg  = self.app.cfg.get("conti", {}).get(conto_name, {})
        tipo       = conto_cfg.get("tipo", "conto_corrente")
        anno       = self.var_anno.get()

        if tipo == "carta_credito":
            messagebox.showinfo("Riepilogo",
                "Riepilogo mensile disponibile solo per conti CC (non carte).")
            return
        sheet_name = conto_cfg.get("excel_sheet", "Movimenti")
        try:
            integr  = ExcelIntegrator(elab_path, "cc")
            righe   = integr.riepilogo_mensile(sheet_name, anno)
            if not righe:
                messagebox.showinfo("Riepilogo", "Nessun dato trovato"); return

            dlg = tk.Toplevel(self)
            dlg.title(f"Riepilogo mensile {anno} — {conto_name}")
            dlg.geometry("780x400")
            cols = ("Mese", "Saldo inizio", "Entrate", "Uscite",
                    "Netto", "Saldo fine", "N")
            tree = ttk.Treeview(dlg, columns=cols, show="headings")
            for col, w in zip(cols, (100, 110, 110, 110, 110, 110, 45)):
                tree.heading(col, text=col)
                tree.column(col, width=w,
                            anchor="w" if col == "Mese" else "e")
            vsb = ttk.Scrollbar(dlg, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y")
            tree.pack(fill="both", expand=True, padx=8, pady=8)
            # Colore: prefisso simbolo nel campo Netto per chiarezza visiva
            # (background ttk su macOS è poco contrastato)
            # Verde/rosso sul saldo finale con background più saturo
            tree.tag_configure("pos", background="#b8ddb8")   # verde medio
            tree.tag_configure("neg", background="#f4aaaa")   # rosso medio
            tree.tag_configure("neu", background="")          # neutro
            for r in righe:
                netto = r["netto"]
                if netto > 0.005:
                    tag = "pos"
                    netto_str = f"▲ {fmt_eur(netto)}"
                elif netto < -0.005:
                    tag = "neg"
                    netto_str = f"▼ {fmt_eur(netto)}"
                else:
                    tag = "neu"
                    netto_str = fmt_eur(netto)
                tree.insert("", "end", tags=(tag,), values=(
                    r["nome_mese"],
                    fmt_eur(r["saldo_iniziale"]),
                    fmt_eur(r["entrate"]),
                    fmt_eur(-r["uscite"]),
                    netto_str,
                    fmt_eur(r["saldo_finale"]),
                    r["n_transazioni"],
                ))
        except Exception as e:
            messagebox.showerror("Errore", str(e))


# ─────────────────────────────────────────────────────────────────────────────
# APP PRINCIPALE
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# APPLICAZIONE PRINCIPALE
#
# Finestra root tkinter. Crea i 4 tab e gestisce:
#   - Caricamento/salvataggio config (ms_config.json)
#   - Caricamento categorie e payee dal .ffd (db_cats, db_payee_cats)
#   - Log testuale condiviso (App.log → TabLog.append)
# Entry point: App().mainloop()
# ══════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    # Profili disponibili: codice → nome display (uguale a Budget GUI)
    PROFILI = {"IRC": "IRC — Ignazio", "SC": "SC — Silvia"}

    def __init__(self):
        super().__init__()
        self.title("Moneyspire Reconciler  v2.1.3")
        self.geometry("1340x820")
        self.minsize(960, 640)

        self.cfg   = {}
        self.rules = None
        self.db_cats:       dict[int, str] = {}
        self.db_payees:     dict[int, str] = {}
        self.db_payee_cats: dict[str, int] = {}

        self._build()

        # Avvia con l'ultimo profilo usato, default IRC
        ultimo = self._leggi_ultimo_profilo()
        self._var_profilo.set(self.PROFILI.get(ultimo, self.PROFILI["IRC"]))
        self._attiva_profilo(ultimo)

    def _leggi_ultimo_profilo(self) -> str:
        """Legge l'ultimo profilo usato da un file dedicato, default IRC."""
        try:
            import json as _json
            p = _CFG_DIR / "ultimo_profilo.json"
            if p.exists():
                return _json.loads(p.read_text(encoding="utf-8")).get("profilo", "IRC")
        except Exception:
            pass
        return "IRC"

    def _salva_ultimo_profilo(self, codice: str):
        """Salva il profilo corrente in un file dedicato per il prossimo avvio."""
        try:
            import json as _json
            p = _CFG_DIR / "ultimo_profilo.json"
            p.write_text(_json.dumps({"profilo": codice}, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _build(self):
        # ── Barra inferiore (PRIMA del notebook!) ─────────────────────────
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", side="bottom")

        ttk.Button(bottom, text="✖  Esci", command=self.quit,
                   width=10).pack(side="right", padx=(4, 8), pady=3)
        ttk.Separator(bottom, orient="vertical").pack(side="right", fill="y", pady=4)
        ttk.Button(bottom, text="🗑  Pulisci",
                   command=lambda: self.tab_log._clear()
                   ).pack(side="right", padx=4, pady=3)
        ttk.Button(bottom, text="💾  Salva su file",
                   command=lambda: self.tab_log._save_file()
                   ).pack(side="right", padx=4, pady=3)
        ttk.Button(bottom, text="📋  Copia negli appunti",
                   command=lambda: self.tab_log._copy()
                   ).pack(side="right", padx=4, pady=3)
        ttk.Separator(bottom, orient="vertical").pack(side="right", fill="y", pady=4)

        self.status = ttk.Label(bottom, text="Caricamento…",
                                anchor="w", relief="sunken", padding=(6, 2))
        self.status.pack(side="left", fill="x", expand=True)

        # ── Notebook ──────────────────────────────────────────────────────
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill="both", expand=True, padx=6, pady=(4, 0))

        self.tab_ric    = TabRiconcilia(self._nb, self)
        self.tab_regole = TabRegole(self._nb, self)
        self.tab_fase2  = TabFase2(self._nb, self)
        self.tab_log    = TabLog(self._nb)

        self._nb.add(self.tab_ric,    text="  📊  Riconcilia  ")
        self._nb.add(self.tab_regole, text="  📋  Regole  ")
        self._nb.add(self.tab_fase2,  text="  📁  Fase 2 Excel  ")
        self._nb.add(self.tab_log,    text="  📝  Log  ")

    def _on_profilo_change(self, _=None):
        """Cambio profilo dalla combobox."""
        label = self._var_profilo.get()
        codice = next((k for k, v in self.PROFILI.items() if v == label), "IRC")
        self._attiva_profilo(codice)

    def _attiva_profilo(self, codice: str):
        """Carica config e rules del profilo scelto, aggiorna la UI."""
        _init_profilo_da_argomento(codice)

        self.title(APP_TITLE)
        try:
            self.cfg = load_config()
        except FileNotFoundError as _e:
            from tkinter import messagebox
            messagebox.showerror("Configurazione mancante", str(_e))
            return
        self.rules = RulesEngine(str(RULES_PATH))
        self._nb.select(0)

        n_regole = len(self.rules.rules)
        prefisso = self.cfg.get("prefisso_output", f"{codice}_")
        self._lbl_prefisso.config(text=prefisso)
        self._lbl_regole.config(text=f"{n_regole} regole caricate")

        # Pulizia tabella risultati e stato analisi
        self.tab_ric.results = []
        self.tab_ric._last_ffd = ""
        self.tab_ric._last_account_id = 0
        self.tab_ric.tree.delete(*self.tab_ric.tree.get_children())
        # Pulizia log
        self.tab_log._clear()

        self.tab_ric._aggiorna_conti()
        self.tab_fase2._aggiorna_conti()
        self.tab_regole._reload()

        self.log(f"Avvio {APP_TITLE}")
        self.log(f"Config: {CFG_PATH}")
        self.log(f"Regole: {RULES_PATH}  ({n_regole} caricate)")

        ffd_saved = self.cfg.get("last_paths", {}).get("ffd", "")
        if ffd_saved and os.path.exists(ffd_saved):
            self.after(100, lambda: self.load_db_cats(ffd_saved))

        self.status.config(text=f"Ambiente {codice} — Pronto.")
        self._salva_ultimo_profilo(codice)

    def log(self, msg: str):
        if hasattr(self, "tab_log"):
            self.tab_log.append(msg)
        if hasattr(self, "status"):
            msg_lower = msg.lower()
            if not any(k in msg_lower for k in ("config:", "regole:", "avvio ", "profilo attivo:")):
                self.status.config(text=msg[:120] + ("…" if len(msg) > 120 else ""))
        self.update_idletasks()

    def load_db_cats(self, ffd_path: str):
        try:
            db = MoneyspireDB(ffd_path)
            self.db_cats          = db.get_categories()
            self.db_payees        = db.get_payees()
            self.db_payee_cats    = db.get_payee_categories()
            db.close()
            self.log(f"  Categorie caricate dal .ffd: {len(self.db_cats)}")
        except Exception as e:
            self.log(f"  Avviso: impossibile caricare categorie — {e}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
