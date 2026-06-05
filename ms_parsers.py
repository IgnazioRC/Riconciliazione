"""
ms_parsers.py — Parser file Excel Fineco e lettura saldi
               parse_fineco_conto_originale / _usd / _cc
               leggi_saldo_fineco / leggi_variazione_mensile_*
Parte di: Moneyspire Reconciler
"""

import re
import csv
from pathlib import Path
from datetime import date, datetime
from difflib import SequenceMatcher

import pandas as pd

from ms_constants import _to_date, _to_float, fmt_eur, DEFAULT_CONFIG, _amt_eq, _date_ok

# ─────────────────────────────────────────────────────────────────────────────
# PARSER FILE BANCA
# ─────────────────────────────────────────────────────────────────────────────
#
# Ogni conto ha DUE file distinti:
#
# FILE ORIGINALE FINECO  → fonte dati per l'analisi/riconciliazione
#   CC:   movements_YYYYMMDD.xlsx  (header riga 13, note iniziali, IBAN in riga 1)
#   Carte: estrattoconto_N_.xlsx   (header riga 1, colonna Numero carta)
#
# FILE ELABORATO (tuo)   → archivio storico da aggiornare
#   CC:   2026.xlsx  (fogli Movimenti/Lombard/USD, formule saldo, fine mese giallo)
#   MC:   2026_mc.xlsx  (12 fogli mensili)
#   Visa: 2026_visa.xlsx (12 fogli mensili, entrambe le carte)
#
# Il motore di analisi usa SOLO i file originali Fineco.
# L'integrazione dei file elaborati è un modulo separato (fase 2).
# ─────────────────────────────────────────────────────────────────────────────


def leggi_numero_conto_fineco(xlsx_path: str) -> str:
    """
    Legge il numero conto/IBAN dalla riga 1 del file CC originale Fineco.
    Es: "Conto Corrente: 5874182" → "5874182"
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    first_row = next(ws.iter_rows(max_row=1, values_only=True), (None,))
    wb.close()
    val = str(first_row[0] or "")
    # Cerca numero dopo i due punti
    if ":" in val:
        return val.split(":", 1)[1].strip()
    return val.strip()


def parse_fineco_conto_originale(xlsx_path: str,
                                  mese: int | None = None,
                                  anno: int | None = None) -> list[dict]:
    """
    Parsa il file CC originale esportato da Fineco (movements_YYYYMMDD.xlsx).
    Struttura:
      Righe 1-12: intestazione (numero conto, periodo, note)
      Riga 13:    header colonne
      Riga 14+:   dati
    Colonne: Data_Operazione, Data_Valuta, Entrate, Uscite,
             Descrizione, Descrizione_Completa, Stato, Moneymap
    Filtro: solo righe "Contabilizzato"; filtra per mese/anno se specificati.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Supporta sia un foglio unico che il nome "Movimenti"
    sheet_name = "Movimenti" if "Movimenti" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    col_idx: dict[str, int] = {}
    header_found = False
    result = []

    for row in ws.iter_rows(values_only=True):
        # Cerca la riga header (contiene "Data_Operazione")
        if not header_found:
            if any(str(c or "").strip() == "Data_Operazione" for c in row):
                col_idx = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
                header_found = True
            continue

        def get(col_name):
            idx = col_idx.get(col_name)
            return row[idx] if idx is not None and idx < len(row) else None

        stato = str(get("Stato") or "").strip().lower()
        if stato != "contabilizzato":
            continue

        dt_op = _to_date(get("Data_Operazione"))
        if dt_op is None:
            continue
        if mese is not None and dt_op.month != mese:
            continue
        if anno is not None and dt_op.year != anno:
            continue

        entrate = abs(_to_float(get("Entrate")))
        uscite  = abs(_to_float(get("Uscite")))
        if entrate == 0 and uscite == 0:
            continue

        desc  = str(get("Descrizione")         or "").strip()
        desc2 = str(get("Descrizione_Completa") or "").strip()

        result.append({
            "date":                 dt_op,
            "date_valuta":          _to_date(get("Data_Valuta")),
            "amount":               entrate - uscite,
            "deposit":              entrate,
            "withdrawal":           uscite,
            "descrizione":          desc,
            "descrizione_completa": desc2,
            "raw_text":             f"{desc} {desc2}".strip(),
            "moneymap":             str(get("Moneymap") or "").strip(),
            "source":               sheet_name,
            "source_type":          "conto_corrente"
        })

    wb.close()
    return result


def parse_fineco_conto(xlsx_path: str, sheet_name: str, col_map: dict) -> list[dict]:
    """
    Parsa il file CC ELABORATO da Ignazio (2026.xlsx).
    Usa openpyxl data_only=True perché contiene formule per il saldo.
    Colonne: Data, Entrate, Uscite, Descrizione, Descrizione_Completa, Moneymap, Saldo
    Mantenuto per compatibilità con la fase di integrazione Excel.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    c_data  = col_map.get("data",                 "Data")
    c_entr  = col_map.get("entrate",              "Entrate")
    c_usc   = col_map.get("uscite",               "Uscite")
    c_desc  = col_map.get("descrizione",          "Descrizione")
    c_desc2 = col_map.get("descrizione_completa", "Descrizione_Completa")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx: dict[str, int] = {str(h).strip(): i
                               for i, h in enumerate(header_row) if h is not None}

    def get(row_vals, col_name):
        idx = col_idx.get(col_name)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    result = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        dt = _to_date(get(row_vals, c_data))
        if dt is None:
            continue
        entrate = abs(_to_float(get(row_vals, c_entr)))
        uscite  = abs(_to_float(get(row_vals, c_usc)))
        if entrate == 0 and uscite == 0:
            continue
        desc  = str(get(row_vals, c_desc)  or "").strip()
        desc2 = str(get(row_vals, c_desc2) or "").strip()
        result.append({
            "date": dt, "amount": entrate - uscite,
            "deposit": entrate, "withdrawal": uscite,
            "descrizione": desc, "descrizione_completa": desc2,
            "raw_text": f"{desc} {desc2}".strip(),
            "source": sheet_name, "source_type": "conto_corrente"
        })

    wb.close()
    return result



def _is_file_usd(xlsx_path: str) -> bool:
    """
    Riconosce il file movimenti del conto USD Fineco.
    Criterio: presenza dell'header 'Data' (senza 'Data_Operazione')
    insieme a 'Descrizione_Completa', oppure intestazione conto USD.
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c or "").strip() for c in row]
            if i < 6 and any("Conto Corrente: USD" in c for c in cells):
                wb.close(); return True
            if "Data" in cells and "Descrizione_Completa" in cells \
                    and "Data_Operazione" not in cells:
                wb.close(); return True
            if i > 12:
                break
        wb.close()
    except Exception:
        return False
    return False


def parse_fineco_conto_usd(xlsx_path: str,
                            mese: int | None = None,
                            anno: int | None = None) -> list[dict]:
    """
    Parsa il file movimenti del conto USD Fineco.
    Formato diverso dal CC: header a riga 7, date gg/mm/yyyy, importi in USD.
    Non filtra per Stato (colonna assente).
    """
    from openpyxl import load_workbook
    from datetime import date as _date

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Trova header
    header_row = None
    for i, row in enumerate(rows):
        if any(str(v or "").strip() == "Data" for v in row):
            header_row = i
            break
    if header_row is None:
        return []

    result = []
    for row in rows[header_row + 1:]:
        dt_raw = row[0]
        if not dt_raw:
            continue
        # Parsing data: gg/mm/yyyy oppure datetime
        try:
            if isinstance(dt_raw, str) and "/" in dt_raw:
                parts = dt_raw.strip().split("/")
                dt = _date(int(parts[2]), int(parts[1]), int(parts[0]))
            elif hasattr(dt_raw, "date"):
                dt = dt_raw.date()
            else:
                continue
        except Exception:
            continue

        if mese and dt.month != mese:
            continue
        if anno and dt.year != anno:
            continue

        entrate = float(row[2] or 0)
        uscite  = abs(float(row[3] or 0))
        desc    = str(row[4] or "").strip()
        desc2   = str(row[5] or "").strip()

        amount = entrate - uscite
        if amount == 0:
            continue

        raw = f"{desc} {desc2}"
        result.append({
            "date":                 dt,
            "amount":               amount,
            "deposit":              entrate,
            "withdrawal":           uscite,
            "descrizione":          desc,
            "descrizione_completa": desc2,
            "raw_text":             raw,
            "intestatario":         "",
            "valuta":               "USD",
            "_is_usd":              True,   # flag per il matcher
        })
    return result

def parse_fineco_cc(xlsx_path: str,
                    numero_carta: str | list[str] | None = None,
                    mese: int | None = None,
                    anno: int | None = None,
                    solo_contabilizzati: bool = True) -> list[dict]:
    """
    Parsa il file estratto conto carta Fineco (formato nativo).
    Funziona per MC (una carta) e Visa (due carte: Ignazio + Silvia).
    numero_carta: stringa singola, lista di stringhe, o None (tutte).
    Filtra per ultime 4 cifre del numero carta e per mese/anno.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    col_idx: dict[str, int] = {}
    header_found = False
    result = []

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if any(str(c or "").strip() == "Data operazione" for c in row):
                col_idx = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
                header_found = True
            continue

        def get(col_name):
            idx = col_idx.get(col_name)
            return row[idx] if idx is not None and idx < len(row) else None

        num_carta_val = str(get("Numero carta") or "").strip()
        stato         = str(get("Stato operazione") or "").strip().lower()

        # Filtra per numero carta: accetta stringa singola o lista
        if numero_carta:
            carte_ok = numero_carta if isinstance(numero_carta, list) else [numero_carta]
            if not any(c in num_carta_val for c in carte_ok):
                continue
        if solo_contabilizzati and stato not in ("contabilizzato", ""):
            continue

        dt_op  = _to_date(get("Data operazione"))
        dt_reg = _to_date(get("Data registrazione"))
        if dt_op is None and dt_reg is None:
            continue

        # Filtro mese: include la transazione se:
        # - la data REGISTRAZIONE è nel mese richiesto (caso normale), OPPURE
        # - la data OPERAZIONE è nel mese richiesto e la data registrazione è
        #   nel mese successivo (transazioni di fine mese addebitate dopo)
        # In questo modo l'estratto di febbraio include le transazioni del
        # 26-28/02 anche se registrate il 2-4 marzo.
        if mese is not None:
            nel_mese_per_reg = (dt_reg and dt_reg.month == mese
                                and (anno is None or dt_reg.year == anno))
            nel_mese_per_op  = (dt_op  and dt_op.month  == mese
                                and (anno is None or dt_op.year  == anno))
            if not nel_mese_per_reg and not nel_mese_per_op:
                continue
        elif anno is not None:
            dt_filtro = dt_reg or dt_op
            if dt_filtro and dt_filtro.year != anno:
                continue

        importo = _to_float(get("Importo"))
        desc    = str(get("Descrizione") or "").strip()
        result.append({
            "date":                 dt_op or dt_reg,
            "date_registrazione":   dt_reg,
            "amount":               importo,
            "deposit":              max(importo, 0),
            "withdrawal":           abs(min(importo, 0)),
            "descrizione":          desc,
            "descrizione_completa": desc,
            "raw_text":             desc,
            "numero_carta":         num_carta_val,
            "intestatario":         str(get("Intestatario carta") or "").strip(),
            "stato":                stato,
            "source":               "CC:" + (num_carta_val[-4:] if num_carta_val else "?"),
            "source_type":          "carta_credito"
        })

    wb.close()
    return result

# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# PARSER CC DA CARTELLA (multi-file)
# ─────────────────────────────────────────────────────────────────────────────

def parse_fineco_cc_cartella(cartella: str,
                             numero_carta: str | list[str] | None = None,
                             mese: int | None = None,
                             anno: int | None = None,
                             solo_contabilizzati: bool = True) -> list[dict]:
    """
    Parsa tutti i file .xlsx di estratto conto Fineco CC presenti in una cartella,
    li unifica e rimuove i duplicati (stessa data operazione + importo + descrizione).

    Utile per passare più estratti mensili in una volta sola, evitando il problema
    delle transazioni di fine mese che compaiono nell'estratto del mese successivo.

    Il filtraggio per mese/anno viene applicato su ogni file individualmente.
    Se mese=None, vengono incluse tutte le transazioni trovate nei file.

    Restituisce la lista ordinata per data operazione.
    """
    from pathlib import Path as _Path

    cartella_p = _Path(cartella).expanduser()
    if not cartella_p.is_dir():
        raise ValueError(f"La cartella non esiste: '{cartella}'")

    # Legge tutti i file .xlsx nella cartella (non ricorsivo)
    files = sorted(cartella_p.glob("*.xlsx"))
    if not files:
        raise ValueError(f"Nessun file .xlsx trovato in: '{cartella}'")

    all_txns: list[dict] = []
    for f in files:
        try:
            txns = parse_fineco_cc(str(f), numero_carta, mese, anno, solo_contabilizzati)
            for t in txns:
                t["_source_file"] = f.name   # traccia il file di origine
            all_txns.extend(txns)
        except Exception:
            # File non riconosciuto come estratto CC — lo ignora silenziosamente
            pass

    # Deduplicazione: rimuove righe identiche per (data, importo, descrizione, carta)
    # che possono comparire in più estratti mensili sovrapposti
    seen: set[tuple] = set()
    unique: list[dict] = []
    for t in all_txns:
        key = (
            t.get("date"),
            t.get("amount"),
            t.get("descrizione", ""),
            t.get("numero_carta", ""),
        )
        if key not in seen:
            seen.add(key)
            unique.append(t)

    # Ordina per data operazione
    unique.sort(key=lambda t: t.get("date") or t.get("date_registrazione"))
    return unique


# ─────────────────────────────────────────────────────────────────────────────
# PARSER UNICREDIT
# ─────────────────────────────────────────────────────────────────────────────

def parse_unicredit_ccm(xls_path: str,
                        mese: int | None = None,
                        anno: int | None = None) -> list[dict]:
    """
    Parsa il file movimenti esportato da Unicredit (formato .xls).

    Struttura del file:
      Riga 1:   intestazione conto  (es. "Rapporto IT 26 Q ... - RUSCONI CLERICI IGNAZIO")
      Riga 2:   "Saldo contabile € X A Credito"
      Riga 3:   "Saldo disponibile € X A Credito"
      Riga 4:   vuota
      Riga 5:   header colonne → Data Registrazione | Data valuta | Causale | Descrizione | Importo (EUR)
      Riga 6+:  dati

    Colonne usate:
      - Data Registrazione  → data della transazione (datetime o stringa gg/mm/yyyy)
      - Descrizione         → testo descrittivo (può essere molto lungo, contiene causale bancaria)
      - Importo (EUR)       → positivo = entrata, negativo = uscita
      - Causale             → codice numerico (ignorato nel matching, conservato per debug)

    Filtro mese/anno applicato sulla Data Registrazione.
    """
    import pandas as pd
    from datetime import date as _date

    try:
        # Legge tutto il file senza interpretare l'header automaticamente
        df_raw = pd.read_excel(xls_path, header=None, engine="xlrd")
    except Exception as e:
        raise ValueError(f"Impossibile aprire il file Unicredit '{xls_path}': {e}")

    # Cerca la riga header (quella che contiene "Data Registrazione")
    header_row_idx = None
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v)]
        if any("Data Registrazione" in v or "Data registrazione" in v for v in vals):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(
            f"Header non trovato nel file Unicredit '{xls_path}'. "
            "Atteso 'Data Registrazione' nella riga 5."
        )

    # Rilegge con l'header corretto
    df = pd.read_excel(xls_path, header=header_row_idx, engine="xlrd")

    # Normalizza nomi colonne (strip spazi, case-insensitive)
    df.columns = [str(c).strip() for c in df.columns]

    # Mappa colonne flessibile
    col_data   = next((c for c in df.columns if "data registrazione" in c.lower()), None)
    col_desc   = next((c for c in df.columns if c.lower() == "descrizione"), None)
    col_import = next((c for c in df.columns if "importo" in c.lower()), None)
    col_causale = next((c for c in df.columns if "causale" in c.lower()), None)

    if not col_data or not col_import:
        raise ValueError(
            f"Colonne obbligatorie non trovate nel file Unicredit. "
            f"Trovate: {list(df.columns)}"
        )

    result = []
    for _, row in df.iterrows():
        # Parsing data
        dt_raw = row.get(col_data)
        if pd.isna(dt_raw) or dt_raw is None:
            continue

        try:
            if isinstance(dt_raw, str):
                dt_raw = dt_raw.strip()
                # Prova formati comuni
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                    try:
                        dt = datetime.strptime(dt_raw[:len(fmt)], fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    continue
            elif hasattr(dt_raw, "date"):
                dt = dt_raw.date()
            elif hasattr(dt_raw, "year"):
                dt = _date(dt_raw.year, dt_raw.month, dt_raw.day)
            else:
                continue
        except Exception:
            continue

        # Filtro mese/anno
        if mese is not None and dt.month != mese:
            continue
        if anno is not None and dt.year != anno:
            continue

        # Importo: positivo = entrata, negativo = uscita
        try:
            importo = _to_float(row.get(col_import))
        except Exception:
            continue
        if importo == 0:
            continue

        # Descrizione: può contenere newline e spazi multipli — normalizza
        desc_raw = str(row.get(col_desc) or "").strip()
        # Unicredit inserisce il testo su più righe con doppi spazi — compatta
        desc = " ".join(desc_raw.split())

        causale = str(row.get(col_causale) or "").strip() if col_causale else ""

        entrate  = max(importo, 0.0)
        uscite   = abs(min(importo, 0.0))

        result.append({
            "date":                 dt,
            "amount":               importo,
            "deposit":              entrate,
            "withdrawal":           uscite,
            "descrizione":          desc,
            "descrizione_completa": desc,
            "raw_text":             desc,
            "causale_unicredit":    causale,   # conservato per debug, non usato nel match
            "source":               "Unicredit CCM",
            "source_type":          "conto_corrente",
        })

    return result


def leggi_intestazione_unicredit(xls_path: str) -> str:
    """
    Legge la riga 1 del file Unicredit per estrarre l'intestazione del conto.
    Usata dalla validazione per verificare che il file sia Unicredit.
    Es: "Rapporto IT 26 Q 02008 35261 000013117508EUR - RUSCONI CLERICI IGNAZIO"
    """
    import pandas as pd
    try:
        df = pd.read_excel(xls_path, header=None, nrows=1, engine="xlrd")
        val = str(df.iloc[0, 0] or "").strip()
        return val
    except Exception:
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# MATCHING ENGINE
# ─────────────────────────────────────────────────────────────────────────────

# _amt_eq e _date_ok importate da ms_constants




def leggi_saldo_fineco(xlsx_path: str, mese: int, anno: int) -> float | None:
    """
    Legge il saldo di fine mese dal file originale Fineco.
    Strategia: legge il Saldo Iniziale e somma solo le transazioni
    fino alla fine del mese richiesto (il file può coprire più mesi).
    Se il mese richiesto è l'ultimo del file, usa direttamente il Saldo Finale.
    """
    from openpyxl import load_workbook
    import calendar, re

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        saldo_iniziale = None
        saldo_finale   = None
        header_found   = False
        ultimo_giorno  = calendar.monthrange(anno, mese)[1]

        totale_entrate = 0.0
        totale_uscite  = 0.0
        ultima_data_file = None

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            cell0 = str(row[0]).strip()

            # Riga header: "Saldo Iniziale: X     Saldo Finale: Y"
            if "Saldo Iniziale" in cell0 and saldo_iniziale is None:
                # Estrai Saldo Iniziale
                m1 = re.search(r'Saldo Iniziale:\\s*([\d\.\,]+)', cell0)
                if m1:
                    saldo_iniziale = float(
                        m1.group(1).replace(".", "").replace(",", "."))
                # Estrai Saldo Finale se presente
                m2 = re.search(r'Saldo Finale:\\s*([\d\.\,]+)', cell0)
                if m2:
                    saldo_finale = float(
                        m2.group(1).replace(".", "").replace(",", "."))
                continue

            if "Data_Operazione" in cell0:
                header_found = True
                continue

            if not header_found:
                continue

            # Riga dati
            try:
                dt_raw = row[0]
                if hasattr(dt_raw, "date"):
                    dt = dt_raw.date()
                else:
                    from datetime import date as _d
                    dt = _d.fromisoformat(str(dt_raw)[:10])
            except Exception:
                continue

            if ultima_data_file is None or dt > ultima_data_file:
                ultima_data_file = dt

            # Somma solo le transazioni fino alla fine del mese richiesto
            from datetime import date as _d2
            fine_mese = _d2(anno, mese, ultimo_giorno)
            if dt > fine_mese:
                continue

            stato = str(row[6] or "").lower() if len(row) > 6 else ""
            if stato and stato != "contabilizzato":
                continue

            totale_entrate += float(row[2] or 0)
            totale_uscite  += abs(float(row[3] or 0))

        wb.close()

        if saldo_iniziale is None:
            return None

        saldo_calcolato = round(saldo_iniziale + totale_entrate - totale_uscite, 2)

        # Se il mese richiesto è l'ultimo del file e abbiamo il Saldo Finale,
        # usa quello (più preciso — include eventuali aggiustamenti Fineco)
        if (saldo_finale is not None and ultima_data_file and
                ultima_data_file.month == mese and ultima_data_file.year == anno):
            return saldo_finale

        return saldo_calcolato

    except Exception:
        return None



def leggi_variazione_mensile_fineco(xlsx_path: str, mese: int, anno: int) -> float | None:
    """Calcola la variazione netta Fineco nel mese (entrate - uscite contabilizzate)."""
    try:
        if _is_file_usd(xlsx_path):
            txns = parse_fineco_conto_usd(xlsx_path, mese, anno)
        else:
            txns = parse_fineco_conto_originale(xlsx_path, mese, anno)
        return round(sum(t["amount"] for t in txns), 2)
    except Exception:
        return None


def leggi_saldo_money(db: "MoneyspireDB", account_id: int,
                      mese: int, anno: int) -> float | None:
    """
    Calcola il saldo Money a fine mese usando InitialBalance + movimenti storici.
    Se InitialBalance = 0 e il conto ha storia lunga, usa la variazione mensile
    rispetto al saldo Fineco per il confronto (modalità delta).
    """
    import sqlite3, calendar
    from datetime import date

    try:
        ultimo = calendar.monthrange(anno, mese)[1]
        fine   = date(anno, mese, ultimo).isoformat()
        conn   = sqlite3.connect(db.path)
        cur    = conn.cursor()

        # Leggi InitialBalance del conto
        cur.execute("SELECT COALESCE(InitialBalance,0) FROM Accounts WHERE ID=?",
                    (account_id,))
        row = cur.fetchone()
        initial = float(row[0] or 0) if row else 0.0

        # Somma tutte le transazioni fino alla fine del mese
        cur.execute("""
            SELECT COALESCE(SUM(Deposit),0) - COALESCE(SUM(Withdrawal),0)
            FROM Transactions
            WHERE AccountID=? AND DATE(TransactionDate)<=?
        """, (account_id, fine))
        mov = float(cur.fetchone()[0] or 0)

        conn.close()
        return round(initial + mov, 2)
    except Exception:
        return None


def leggi_variazione_mensile_money(db: "MoneyspireDB", account_id: int,
                                    mese: int, anno: int) -> float | None:
    """
    Calcola la variazione netta Money nel mese (dep - wit).
    Usato per confronto quando il saldo assoluto non è affidabile.
    """
    import sqlite3, calendar
    from datetime import date

    try:
        primo  = date(anno, mese, 1).isoformat()
        ultimo = calendar.monthrange(anno, mese)[1]
        fine   = date(anno, mese, ultimo).isoformat()
        conn   = sqlite3.connect(db.path)
        cur    = conn.cursor()
        cur.execute("""
            SELECT COALESCE(SUM(Deposit),0) - COALESCE(SUM(Withdrawal),0)
            FROM Transactions
            WHERE AccountID=? AND DATE(TransactionDate)>=? AND DATE(TransactionDate)<=?
        """, (account_id, primo, fine))
        var = float(cur.fetchone()[0] or 0)
        conn.close()
        return round(var, 2)
    except Exception:
        return None
