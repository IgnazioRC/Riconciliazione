"""
ms_parsers_silvia.py
====================
Parser aggiuntivi per il profilo Silvia (SC):
  - parse_bper()             : movimenti BPER (ex Banca Popolare di Sondrio), formato .xls
  - parse_nexi_xlsx()        : movimenti carta Nexi, formato .xlsx esportato dal portale Nexi Pay
  - leggi_intestazione_bper(): validazione e metadati file BPER
  - leggi_intestazione_nexi(): validazione e metadati file Nexi xlsx

Integrazione in ms_parsers.py (aggiungere in cima):
    from ms_parsers_silvia import (parse_bper, parse_nexi_xlsx,
                                   leggi_intestazione_bper, leggi_intestazione_nexi)

Integrazione in ms_engine.py / ms_matching.py (aggiungere nei case):
    case "originale_bper":
        movimenti = parse_bper(path, mese, anno)
    case "originale_nexi_xlsx":
        movimenti = parse_nexi_xlsx(path, mese, anno)

────────────────────────────────────────────────────────────
Struttura file BPER .xls
────────────────────────────────────────────────────────────
  Righe 0-5  : vuote
  Riga 6     : "Modena, DD mese AAAA"
  Righe 7-11 : intestazione conto (IBAN, saldi, intestatario)
  Righe 12-15: info filtri applicati
  Riga 16    : HEADER colonne →
               Col 1: "Data operazione"
               Col 2: "Data valuta"
               Col 3: "Descrizione"
               Col 4: "Entrate"   (positivo o vuoto)
               Col 5: "Uscite"    (negativo o vuoto)
               Col 6: "Categoria"
               Col 7: "Stato"     ("Contabilizzato" | "Da contabilizzare")
  Riga 17+   : movimenti
  Ultima riga: totali (skippata automaticamente)

────────────────────────────────────────────────────────────
Struttura file Nexi .xlsx
────────────────────────────────────────────────────────────
  Righe 0-6  : intestazione (titolo, periodo, numero carta...)
  Riga 9     : HEADER colonne →
               Col 1: "Mese"         (es. "Marzo")
               Col 2: "Data"         (es. "17/03/2026")
               Col 3: "Riferimento"
               Col 4: "Categorie"
               Col 5: "Descrizione"
               Col 6: "Stato"        ("" | "Non Contabilizzato")
               Col 9: "Importo (€)"  (float positivo = uscita)
  Riga 10+   : movimenti
"""

import re
import warnings
from datetime import date, datetime
from typing import Optional

# ──────────────────────────────────────────────────────────
# Costanti
# ──────────────────────────────────────────────────────────

MESI_IT_LONG = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}


# ──────────────────────────────────────────────────────────
# Utilità condivise
# ──────────────────────────────────────────────────────────

def _parse_data_it(s) -> Optional[date]:
    """
    Parsa date in vari formati italiani:
      - oggetti date/datetime già parsati da openpyxl
      - "05 maggio 2026"
      - "05/03/2026" o "05-03-2026"
    Restituisce None se non riesce.
    """
    if s is None:
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).strip().strip('"')
    if not s or s in ("-", "–"):
        return None

    # Formato DD/MM/YYYY o DD-MM-YYYY
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None

    # Formato "DD mese YYYY"
    parts = s.lower().split()
    if len(parts) == 3:
        mese_n = MESI_IT_LONG.get(parts[1])
        if mese_n:
            try:
                return date(int(parts[2]), mese_n, int(parts[0]))
            except ValueError:
                return None

    return None


def _parse_importo(v) -> Optional[float]:
    """
    Converte un valore cella in float.
    Gestisce: float/int, "1.234,56", "-0.36", celle vuote → None.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "").strip()
    if not s or s in ("-", "–"):
        return None
    # Formato italiano "1.234,56"
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# ──────────────────────────────────────────────────────────
# BPER
# ──────────────────────────────────────────────────────────

def leggi_intestazione_bper(path: str) -> dict:
    """
    Legge i metadati dal file BPER .xls.
    Restituisce dict con: titolare, iban, saldo_disponibile.
    Lancia ValueError se il file non è riconosciuto come BPER.
    """
    try:
        import xlrd
    except ImportError:
        raise ImportError("Libreria 'xlrd' non trovata. Installa con: pip install xlrd")

    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)

    info = {}
    found = False

    for r in range(min(ws.nrows, 20)):
        row = [str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)]
        joined = " ".join(row).lower()

        if "saldo disponibile" in joined or "saldo contabile" in joined:
            found = True
            # Saldo disponibile è in col 5
            for c in range(ws.ncols):
                v = str(ws.cell_value(r, c)).strip()
                if "saldo disponibile" in v.lower():
                    saldo_raw = str(ws.cell_value(r, c + 1)).strip() if c + 1 < ws.ncols else ""
                    info["saldo_disponibile"] = _parse_importo(saldo_raw)

        if "intestatari" in joined or "intestatario" in joined:
            for c in range(1, ws.ncols):
                v = str(ws.cell_value(r, c)).strip()
                if v and "intestat" not in v.lower():
                    info["titolare"] = v
                    break

        if "iban" in joined:
            for c in range(1, ws.ncols):
                v = str(ws.cell_value(r, c)).strip()
                if v.upper().startswith("IT") and len(v) > 10:
                    info["iban"] = v
                    break

    if not found:
        raise ValueError(f"Il file non sembra un export BPER: '{path}'")

    return info


def parse_bper(path: str,
               mese: Optional[int] = None,
               anno: Optional[int] = None,
               includi_non_contabilizzati: bool = True) -> list[dict]:
    """
    Parsa un file export BPER (.xls).

    Restituisce lista di dict:
        data_operazione : date
        data_valuta     : date | None
        descrizione     : str
        importo         : float  (+ entrata, - uscita)
        categoria_banca : str
        stato           : str   ("Contabilizzato" | "Da contabilizzare")
        fonte           : "BPER"
    """
    try:
        import xlrd
    except ImportError:
        raise ImportError("Libreria 'xlrd' non trovata. Installa con: pip install xlrd")

    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)

    # Trova riga header
    header_row = None
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            if str(ws.cell_value(r, c)).strip().lower() == "data operazione":
                header_row = r
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError(f"Header colonne non trovato nel file BPER: '{path}'")

    # Mappa colonne
    col = {}
    for c in range(ws.ncols):
        h = str(ws.cell_value(header_row, c)).strip().lower()
        if h == "data operazione":   col["data_op"]   = c
        elif h == "data valuta":     col["data_val"]  = c
        elif h == "descrizione":     col["desc"]      = c
        elif h == "entrate":         col["entrate"]   = c
        elif h == "uscite":          col["uscite"]    = c
        elif h == "categoria":       col["categoria"] = c
        elif h == "stato":           col["stato"]     = c

    for k in ("data_op", "desc", "entrate", "uscite"):
        if k not in col:
            raise ValueError(f"Colonna '{k}' mancante nel file BPER: '{path}'")

    movimenti = []

    for r in range(header_row + 1, ws.nrows):
        desc = str(ws.cell_value(r, col["desc"])).strip()
        # Salta righe vuote, totali e note finali
        if not desc or desc.lower() in ("totale",) or desc.lower().startswith("dati aggiornati"):
            continue

        data_op = _parse_data_it(str(ws.cell_value(r, col["data_op"])).strip())
        if data_op is None:
            continue

        # Filtri periodo
        if mese is not None and data_op.month != mese:
            continue
        if anno is not None and data_op.year != anno:
            continue

        data_val = _parse_data_it(str(ws.cell_value(r, col["data_val"])).strip()) \
                   if "data_val" in col else None

        entrata = _parse_importo(ws.cell_value(r, col["entrate"]))
        uscita  = _parse_importo(ws.cell_value(r, col["uscite"]))

        if entrata is not None and entrata != 0:
            importo = abs(entrata)
        elif uscita is not None and uscita != 0:
            importo = -abs(uscita)
        else:
            importo = 0.0

        stato = str(ws.cell_value(r, col["stato"])).strip() if "stato" in col else ""
        if not includi_non_contabilizzati and stato.lower() == "da contabilizzare":
            continue

        categoria_banca = str(ws.cell_value(r, col["categoria"])).strip() \
                          if "categoria" in col else ""

        movimenti.append({
            # Campi standard attesi dal motore di matching
            "date":        data_op,
            "amount":      importo,
            "deposit":     max(importo, 0),
            "withdrawal":  abs(min(importo, 0)),
            "raw_text":    desc,
            "descrizione": desc,
            "descrizione_completa": desc,
            "source":      "BPER",
            "source_type": "conto_corrente",
            # Campi specifici BPER
            "data_operazione": data_op,
            "data_valuta":     data_val,
            "importo":         importo,
            "categoria_banca": categoria_banca,
            "stato":           stato,
            "fonte":           "BPER",
        })

    return movimenti


# ──────────────────────────────────────────────────────────
# Nexi xlsx
# ──────────────────────────────────────────────────────────

def leggi_intestazione_nexi(path: str) -> dict:
    """
    Legge i metadati dal file Nexi .xlsx.
    Restituisce dict con: numero_carta, tipo_carta, periodo.
    Lancia ValueError se il file non è riconosciuto come Nexi.
    """
    try:
        import openpyxl
        warnings.filterwarnings("ignore", category=UserWarning)
    except ImportError:
        raise ImportError("Libreria 'openpyxl' non trovata. Installa con: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    info = {}
    found = False

    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
        row_str = " ".join(str(v) for v in row if v is not None)
        row_lower = row_str.lower()

        if "movimenti carta" in row_lower or "carta di credito" in row_lower:
            found = True
        for v in row:
            if v is None:
                continue
            s = str(v)
            if "Movimenti Carta" in s:
                info["numero_carta"] = s.replace("Movimenti Carta", "").strip()
            if "Periodo" in s and "dal" in s.lower():
                info["periodo"] = s.strip()
        if "carta di credito" in row_lower:
            info["tipo_carta"] = "Carta di Credito"

    wb.close()

    if not found:
        raise ValueError(f"Il file non sembra un export Nexi: '{path}'")

    return info


def parse_nexi_xlsx(path: str,
                    mese: Optional[int] = None,
                    anno: Optional[int] = None,
                    includi_non_contabilizzati: bool = True) -> list[dict]:
    """
    Parsa un file export Nexi (.xlsx) scaricato dal portale Nexi Pay.

    Restituisce lista di dict:
        data_operazione : date
        mese_estratto   : str    (es. "Marzo")
        descrizione     : str
        importo         : float  (negativo = uscita, positivo = rimborso)
        categoria_nexi  : str
        riferimento     : str
        stato           : str   ("" | "Non Contabilizzato")
        fonte           : "Nexi"

    Nota: tutti i movimenti Nexi sono uscite (importo negativo).
    I rimborsi eventuali hanno importo positivo nel file → restituiti positivi.
    La quota annuale carta appare come uscita normale.
    """
    try:
        import openpyxl
        warnings.filterwarnings("ignore", category=UserWarning)
    except ImportError:
        raise ImportError("Libreria 'openpyxl' non trovata. Installa con: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    # Trova riga header
    header_row_idx = None
    col = {}

    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        vals_lower = [str(v).strip().lower() if v is not None else "" for v in row]
        if "mese" in vals_lower and "data" in vals_lower and "descrizione" in vals_lower:
            header_row_idx = i
            for c, h in enumerate(vals_lower):
                if h == "mese":                             col["mese"]       = c
                elif h == "data":                           col["data"]       = c
                elif h == "riferimento":                    col["riferimento"]= c
                elif h == "categorie":                      col["categoria"]  = c
                elif h == "descrizione":                    col["desc"]       = c
                elif h == "stato":                          col["stato"]      = c
                elif "importo (€)" in h or h == "importo": col["importo"]    = c
            break

    if header_row_idx is None:
        raise ValueError(f"Header non trovato nel file Nexi: '{path}'")

    # Fallback colonna importo (col J = indice 9 nel file reale)
    if "importo" not in col:
        col["importo"] = 9

    if "desc" not in col:
        raise ValueError(f"Colonna 'Descrizione' non trovata nel file Nexi: '{path}'")

    movimenti = []

    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 2, values_only=True)):
        if all(v is None for v in row):
            continue

        row_list = list(row)

        # Data
        data_raw = row_list[col["data"]] if col.get("data") is not None else None
        data_op = _parse_data_it(data_raw)
        if data_op is None:
            continue

        # Filtri periodo
        if mese is not None and data_op.month != mese:
            continue
        if anno is not None and data_op.year != anno:
            continue

        # Descrizione
        desc_v = row_list[col["desc"]]
        desc = str(desc_v).strip() if desc_v is not None else ""
        if not desc:
            continue

        # Importo: nel file Nexi i valori sono positivi per le uscite
        importo_raw = row_list[col["importo"]] if col.get("importo") is not None else None
        importo = _parse_importo(importo_raw)
        if importo is None:
            importo = 0.0
        # Converti: positivo nel file = uscita → negativo
        #           negativo nel file = rimborso → positivo
        importo = -importo

        # Campi accessori
        mese_str    = str(row_list[col["mese"]]).strip()       if col.get("mese")        is not None and row_list[col["mese"]]        is not None else ""
        stato       = str(row_list[col["stato"]]).strip()      if col.get("stato")       is not None and row_list[col["stato"]]       is not None else ""
        riferimento = str(row_list[col["riferimento"]]).strip() if col.get("riferimento") is not None and row_list[col["riferimento"]] is not None else ""
        categoria   = str(row_list[col["categoria"]]).strip()  if col.get("categoria")   is not None and row_list[col["categoria"]]   is not None else ""

        if not includi_non_contabilizzati and stato.lower() == "non contabilizzato":
            continue

        movimenti.append({
            # Campi standard attesi dal motore di matching
            "date":        data_op,
            "amount":      importo,
            "deposit":     max(importo, 0),
            "withdrawal":  abs(min(importo, 0)),
            "raw_text":    desc,
            "descrizione": desc,
            "descrizione_completa": desc,
            "source":      "Nexi",
            "source_type": "carta_credito",
            # Campi specifici Nexi
            "data_operazione": data_op,
            "mese_estratto":   mese_str,
            "importo":         importo,
            "categoria_nexi":  categoria,
            "riferimento":     riferimento,
            "stato":           stato,
            "fonte":           "Nexi",
        })

    wb.close()
    return movimenti


# ──────────────────────────────────────────────────────────
# Test da riga di comando
# ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Uso: python ms_parsers_silvia.py <file.xls|file.xlsx> [mese] [anno]")
        sys.exit(1)

    path = sys.argv[1]
    mese_arg = int(sys.argv[2]) if len(sys.argv) > 2 else None
    anno_arg = int(sys.argv[3]) if len(sys.argv) > 3 else None

    if path.lower().endswith(".xls"):
        print(f"\n=== BPER: {path} ===")
        try:
            info = leggi_intestazione_bper(path)
            print(f"Intestazione: {info}")
        except Exception as e:
            print(f"Errore intestazione: {e}")
        movs = parse_bper(path, mese_arg, anno_arg)
        print(f"Movimenti trovati: {len(movs)}")
        for m in movs[:8]:
            print(f"  {m['data_operazione']}  {m['importo']:>10.2f}  {m['descrizione'][:45]}")

    elif path.lower().endswith(".xlsx"):
        print(f"\n=== Nexi xlsx: {path} ===")
        try:
            info = leggi_intestazione_nexi(path)
            print(f"Intestazione: {info}")
        except Exception as e:
            print(f"Errore intestazione: {e}")
        movs = parse_nexi_xlsx(path, mese_arg, anno_arg)
        print(f"Movimenti trovati: {len(movs)}")
        for m in movs[:8]:
            print(f"  {m['data_operazione']}  {m['importo']:>10.2f}  {m['descrizione'][:45]}")


